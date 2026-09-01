"""Diagnostic-only audit of ORB Long V8's own V6 Risk Event mechanism -
"is it actually being evaluated and applied?" - NOT a strategy comparison,
NOT an optimization, NOT a parameter search. See src/db.py's own V8
comment ("Dynamic Risk Reduction Based On V6 Detection") and src/
backtest_engine.py's _evaluate_v6_risk_event/_v6_audit_record, which this
script reads from unchanged.

What this script does and does NOT do:
  - Runs ONE real backtest of V8's own, ALREADY-DEFINED preset (src/
    backtest_runner.run_one_strategy - the exact same simulation every
    dashboard backtest goes through), with its rules_json read from the
    DB and never modified in any way. This is the same "run the existing
    strategy once to see what happened" pattern analyze_v42_hard_stop.py/
    analyze_v41_no_stop.py/analyze_v5_ablation.py already use - it is NOT
    a parameter sweep (that's run_optimization.py, never touched here)
    and no threshold anywhere is changed.
  - Never creates, runs, or references V9.
  - Every trade V8 produces gets exactly one row in the exported "V6
    Audit Report" sheet - triggered or not - straight off that trade's
    own pair["v6_audit"] dict (see _v6_audit_record's own docstring for
    the full field list this reads from). Nothing here re-derives or
    recomputes any V6 condition - it only reports what backtest_engine.py
    itself already decided, so this can't disagree with the actual
    simulation by construction.

Read-only: never modifies the DB, never writes rules_json.

Two ways to run it:
  1. --backtest-id <id> (RECOMMENDED on the production server - cheap,
     read-only): points at an ALREADY-FINISHED multi-strategy backtest
     that included ORB Long V8 (created via the dashboard's "New
     Backtest" form with the "Run on remote worker" box checked - see
     docs/worker.md) and just reads its already-computed pairs/v6_audit
     straight out of the DB. No simulation runs here at all in this mode
     - the actual heavy computation already happened on the WORKER
     machine (your own laptop/desktop), not the small always-on server.
  2. --start-date/--end-date (no --backtest-id): runs the V8 simulation
     ITSELF, right here, right now. Fine on your own machine or a worker
     box - do NOT run this on the small production server (see docs/
     worker.md's own "very little memory headroom" note): a full-universe,
     multi-month run with V6's own per-bar RSI/EMA recomputation is
     exactly the kind of CPU/memory-heavy job that setup warns about, and
     can make the whole box (including the live trading engine/Gateway)
     unresponsive.

Usage:
    # Recommended - read an already-finished remote-worker backtest:
    python3 audit_v6_risk_event.py --backtest-id 42

    # Runs the simulation locally right now - NOT on the production server:
    python3 audit_v6_risk_event.py --start-date 2024-01-01 --end-date 2024-06-30
    python3 audit_v6_risk_event.py --start-date 2024-01-01 --end-date 2024-06-30 --account-id 1 --output v6_audit.xlsx
"""
import argparse
import json
import re
from datetime import date, datetime
from pathlib import Path

from analyze_strategy import find_strategy, section
from analyze_v5_ablation import _run
from src import backtest_data, backtest_engine, db

PROJECT_DIR = Path(__file__).resolve().parent

DEFAULT_PORTFOLIO_VALUE = 100_000.0
DEFAULT_MAX_RISK_PCT = 1.0
DEFAULT_MAX_TRADES_PER_DAY = 5
DEFAULT_COMMISSION_PER_TRADE = 1.5
DEFAULT_OUTPUT = "v6_audit_report.xlsx"

V8_STRATEGY_NEEDLE = "ORB Long V8"


def _find_strategy_id_in_results(results: dict, pattern: str) -> str | None:
    """strategy_id (str) of the first result in one multi-strategy
    backtest's own `results` dict whose strategy_name matches `pattern` -
    same whole-word, case-insensitive convention backtest.html's own
    _findStrategyIdByPattern already uses (so a whole-word "v8" pattern
    never accidentally matches "v4.2"), reimplemented here in Python
    since this script has no access to that JS. None if no result
    matches (or errored)."""
    needle = re.compile(pattern, re.IGNORECASE)
    for sid, r in results.items():
        if isinstance(r, dict) and not r.get("error") and needle.search(r.get("strategy_name") or ""):
            return sid
    return None


def _v6a(pair: dict) -> dict:
    return pair.get("v6_audit") or {}


def _split_iso(value) -> tuple:
    """(date_str, time_str) off an ISO timestamp string - None/None if
    value is falsy (e.g. a trade that never reached a stop-hit at all)."""
    if not value:
        return (None, None)
    try:
        dt = datetime.fromisoformat(value)
    except (TypeError, ValueError):
        return (value, None)
    return (dt.date().isoformat(), dt.time().isoformat())


def build_audit_rows(pairs: list[dict]) -> list[dict]:
    """One row per EVERY V8 trade (not just triggered ones), sorted by
    entry timestamp ascending - the exact column set the audit spec asks
    for, each value read straight off pair["v6_audit"]/pair itself, never
    recomputed."""
    ordered = sorted(pairs, key=lambda p: p.get("buy_time") or "")
    rows = []
    for i, p in enumerate(ordered, start=1):
        v6a = _v6a(p)
        entry_date, entry_time = _split_iso(p.get("buy_time"))
        exit_date, exit_time = _split_iso(p.get("sell_time"))
        hit_date, hit_time = _split_iso(v6a.get("adjusted_stop_hit_timestamp"))

        stop_before_r = v6a.get("stop_before_v6_r")
        stop_after_r = v6a.get("stop_after_v6_r")
        stop_changed = bool(v6a.get("v6_stop_changed"))

        net_pnl = None
        if p.get("pnl_usd") is not None:
            net_pnl = round(float(p["pnl_usd"]) - float(p.get("commission_usd") or 0), 2)

        rows.append({
            "Trade ID": i,
            "Symbol": p.get("symbol"),
            "Entry Date": entry_date, "Entry Time": entry_time,
            "Exit Date": exit_date, "Exit Time": exit_time,

            "Trade Open At 10m?": v6a.get("trade_open_at_v6_evaluation"),
            "V6 Evaluated?": v6a.get("v6_evaluated"),
            "V6 Evaluation Timestamp": v6a.get("v6_actual_evaluation_timestamp"),

            "Trailing Not Active?": v6a.get("condition_trailing_not_active"),
            "Current R At 10m": v6a.get("v6_current_r"),
            "Current R <= -0.40R ?": v6a.get("condition_current_r"),
            "RSI Delta At 10m": v6a.get("v6_rsi_delta"),
            "RSI Delta <= -5 ?": v6a.get("condition_rsi_delta"),
            "Returned Inside Opening Range ?": v6a.get("condition_returned_inside_or"),
            "Lost EMA9 ?": v6a.get("condition_lost_ema9"),
            "10m MFE R So Far": v6a.get("v6_mfe_r_so_far"),
            "MFE <= 0.30R ?": v6a.get("condition_mfe"),

            "All Conditions Passed": v6a.get("all_conditions_passed"),
            "V6 Triggered": v6a.get("v6_triggered"),

            # Signed (e.g. -2.5/-2.0), matching the spec's own examples -
            # stop_before_v6_r/stop_after_v6_r are stored as UNSIGNED
            # magnitudes internally (see _v6_audit_record's own docstring
            # on why), negated here for display only.
            "Original Hard Stop R": -stop_before_r if stop_before_r is not None else None,
            "Adjusted Stop Applied?": stop_changed,
            "Adjusted Stop R": (-stop_after_r if (stop_changed and stop_after_r is not None) else None),
            "Stop Changed?": stop_changed,

            "Adjusted Stop Hit?": v6a.get("adjusted_stop_hit"),
            "Adjusted Stop Hit Timestamp": v6a.get("adjusted_stop_hit_timestamp"),
            "Adjusted Stop Exit R": v6a.get("adjusted_stop_exit_r_after_costs"),

            "Actual Exit Reason": p.get("exit_reason"),
            "Actual Final R": p.get("final_r"),
            "Actual Net P&L": net_pnl,
        })
    return rows


def build_summary(rows: list[dict]) -> dict:
    total = len(rows)
    evaluated = sum(1 for r in rows if r["V6 Evaluated?"])
    triggered = sum(1 for r in rows if r["V6 Triggered"])
    stop_changed = sum(1 for r in rows if r["Stop Changed?"])
    stop_hit = sum(1 for r in rows if r["Adjusted Stop Hit?"])
    triggered_and_stop_changed = sum(1 for r in rows if r["V6 Triggered"] and r["Stop Changed?"])
    stop_hit_and_exit_hard_stop = sum(1 for r in rows if r["Adjusted Stop Hit?"] and r["Actual Exit Reason"] == "hard_stop")
    pct = lambda n: round(n / total * 100, 1) if total else None
    return {
        "Total Trades": total,
        "Trades Evaluated": evaluated,
        "Trades Not Evaluated": total - evaluated,
        "V6 Trigger Count": triggered,
        "V6 Trigger %": pct(triggered),
        "Stop Changed Count": stop_changed,
        "Stop Changed %": pct(stop_changed),
        "Adjusted Stop Hit Count": stop_hit,
        "Adjusted Stop Hit %": pct(stop_hit),
        "Triggered AND Stop Changed Count": triggered_and_stop_changed,
        "Adjusted Stop Hit Count (cross-check)": stop_hit,
        "Adjusted Stop Hit AND Exit=hard_stop Count": stop_hit_and_exit_hard_stop,
    }


def affected_trade_ids(rows: list[dict]) -> list[int]:
    """"Did V8 change any historical trade outcome" - true exactly for a
    trade whose actual exit really was the tightened stop (adjusted_stop_
    hit) - anything else (never triggered, triggered but not tighter,
    triggered but the tightened level was never touched) rode to the same
    exit path V4.2's own untouched logic would have produced."""
    return [r["Trade ID"] for r in rows if r["Adjusted Stop Hit?"]]


def print_final_answer(rows: list[dict], summary: dict, orig_r: float, adjusted_r: float):
    triggered = summary["V6 Trigger Count"]
    stop_changed = summary["Stop Changed Count"]
    stop_hit = summary["Adjusted Stop Hit Count"]
    affected = affected_trade_ids(rows)

    section("VALIDATION")
    print(f"Number of trades where stop actually moved from {orig_r:g}R to {adjusted_r:g}R: {stop_changed}")
    print(f"Number of trades where the new stop changed the final outcome: {stop_hit}")

    section("FINAL ANSWER")
    print(f"1. Was V6 ever triggered?                {'YES' if triggered > 0 else 'NO'}")
    print(f"2. How many times?                       {triggered}")
    print(f"3. Was the stop ever changed?             {'YES' if stop_changed > 0 else 'NO'}")
    print(f"4. How many times?                        {stop_changed}")
    print(f"5. Was the adjusted stop ever hit?         {'YES' if stop_hit > 0 else 'NO'}")
    print(f"6. How many times?                         {stop_hit}")
    print(f"7. Did V8 change any historical trade outcome? {'YES' if affected else 'NO'}")
    print(f"8. Affected Trade IDs:                     {affected if affected else '(none)'}")


def export_xlsx(rows: list[dict], summary: dict, orig_r: float, adjusted_r: float, scope_label: str, output_path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color="EEF1F5", end_color="EEF1F5", fill_type="solid")
    header_font = Font(bold=True)

    def autosize(ws, ncols):
        for col in range(1, ncols + 1):
            letter = get_column_letter(col)
            width = max((len(str(c.value)) for c in ws[letter] if c.value is not None), default=8)
            ws.column_dimensions[letter].width = min(max(width + 2, 8), 40)

    def write_table(ws, table_rows: list[dict]):
        if not table_rows:
            ws.append(["No data"])
            return
        cols = list(table_rows[0].keys())
        ws.append(cols)
        for cell in ws[ws.max_row]:
            cell.font, cell.fill = header_font, header_fill
        for row in table_rows:
            ws.append([row.get(c) for c in cols])
        ws.freeze_panes = "A2"
        autosize(ws, len(cols))

    def write_kv(ws, kv_rows: list[tuple]):
        for label, value in kv_rows:
            ws.append([label, value])
        autosize(ws, 2)

    wb = Workbook()
    audit_ws = wb.active
    audit_ws.title = "V6 Audit Report"
    write_table(audit_ws, rows)

    summary_ws = wb.create_sheet("V6 Audit Summary")
    summary_ws.append(["V6 Risk Event Audit Summary - ORB Long V8"])
    summary_ws["A1"].font = Font(bold=True, size=14)
    summary_ws.append([f"Scope: {scope_label} | Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    summary_ws.append([])
    write_kv(summary_ws, list(summary.items()))
    summary_ws.append([])

    summary_ws.append(["Validation"])
    summary_ws[summary_ws.max_row][0].font = Font(bold=True)
    write_kv(summary_ws, [
        (f"Trades where stop moved from {orig_r:g}R to {adjusted_r:g}R", summary["Stop Changed Count"]),
        ("Trades where the new stop changed the final outcome", summary["Adjusted Stop Hit Count"]),
    ])
    summary_ws.append([])

    affected = affected_trade_ids(rows)
    summary_ws.append(["Final Answer"])
    summary_ws[summary_ws.max_row][0].font = Font(bold=True)
    write_kv(summary_ws, [
        ("1. Was V6 ever triggered?", "YES" if summary["V6 Trigger Count"] > 0 else "NO"),
        ("2. How many times?", summary["V6 Trigger Count"]),
        ("3. Was the stop ever changed?", "YES" if summary["Stop Changed Count"] > 0 else "NO"),
        ("4. How many times?", summary["Stop Changed Count"]),
        ("5. Was the adjusted stop ever hit?", "YES" if summary["Adjusted Stop Hit Count"] > 0 else "NO"),
        ("6. How many times?", summary["Adjusted Stop Hit Count"]),
        ("7. Did V8 change any historical trade outcome?", "YES" if affected else "NO"),
        ("8. Affected Trade IDs", ", ".join(str(x) for x in affected) if affected else "(none)"),
    ])

    wb.save(output_path)


def _v8_pairs_from_one_backtest(backtest_id: int) -> tuple[dict, dict]:
    """(backtest, v8_result) for one already-finished backtest - shared by
    both the single-id and pooled-multi-id paths below, so a bad id/
    status/missing-V8 error is reported identically either way."""
    backtest = db.get_backtest(backtest_id)
    if backtest is None:
        raise SystemExit(f"No backtest with id {backtest_id}.")
    if backtest["status"] != "done":
        raise SystemExit(f"Backtest {backtest_id} is not done yet (status={backtest['status']}).")
    results = backtest["results"] or {}
    v8_sid = _find_strategy_id_in_results(results, r"\bv8\b")
    if v8_sid is None:
        available = [r.get("strategy_name") for r in results.values() if isinstance(r, dict)]
        raise SystemExit(f"Backtest {backtest_id} has no ORB Long V8 result. Strategies in this backtest: {available}")
    return backtest, results[v8_sid]


def _load_pairs_from_backtests(backtest_ids: list[int]) -> tuple[list[dict], str]:
    """(pooled pairs, scope_label) off one or more ALREADY-FINISHED
    backtest rows - zero simulation here, just reads what a remote worker
    (or the server itself, if it was run locally) already computed and
    stored (see src/db.get_backtest). This is the cheap, read-only path
    meant for the small production server - see the module's own
    docstring.

    Pools every given id's own V8 pairs together (e.g. a run of weekly-
    chunked backtests covering a full multi-month range, exactly what
    the dashboard's own remote-worker history tends to look like), with
    the SAME "exact (start_date, end_date) dedup, newest created_at wins"
    rule analyze_strategy.pool_strategy_pairs already uses - so auditing
    the same week's backtest twice by accident (e.g. a retry that got a
    new id) never double-counts that week's trades."""
    latest_by_range = {}
    v8_label = None
    for bid in backtest_ids:
        backtest, v8_result = _v8_pairs_from_one_backtest(bid)
        v8_label = v8_label or v8_result.get("strategy_name", "ORB Long V8")
        params = backtest["params"] or {}
        date_key = (params.get("start_date"), params.get("end_date"))
        existing = latest_by_range.get(date_key)
        if existing is None or backtest["created_at"] > existing["created_at"]:
            latest_by_range[date_key] = {
                "created_at": backtest["created_at"],
                "pairs": v8_result.get("pairs") or [],
                "symbol_count": len(params.get("symbols") or []),
            }

    pooled_pairs, symbol_counts = [], []
    for date_key, entry in sorted(latest_by_range.items()):
        pooled_pairs.extend(entry["pairs"])
        symbol_counts.append(entry["symbol_count"])

    starts = [d[0] for d in latest_by_range if d[0]]
    ends = [d[1] for d in latest_by_range if d[1]]
    date_range = f"{min(starts)} to {max(ends)}" if starts and ends else "unknown range"
    max_symbols = max(symbol_counts) if symbol_counts else 0
    id_desc = f"backtest #{backtest_ids[0]}" if len(backtest_ids) == 1 else f"{len(backtest_ids)} backtests pooled ({backtest_ids[0]}-{backtest_ids[-1]})"
    scope_label = f"{v8_label} | {id_desc} | {date_range} | up to {max_symbols} symbol(s)"
    return pooled_pairs, scope_label


def list_v8_backtests(account_id: int) -> None:
    """--list: prints every 'done' backtest (newest first) that carries an
    ORB Long V8 result, so you don't need raw SQL/sqlite3 on the server
    just to find a backtest_id - reuses db.list_backtests/db.get_backtest,
    the exact same functions the dashboard's own History page uses, so
    there's no separate query logic to keep in sync."""
    summaries = [b for b in db.list_backtests(account_id, limit=200) if b["status"] == "done"]
    if not summaries:
        print("No 'done' backtests found for this account.")
        return
    print(f"{'ID':>6}  {'Date range':<23}  {'Symbols':>7}  {'V8 trades':>9}  Created")
    found_any = False
    for summary in summaries:
        backtest = db.get_backtest(summary["id"])
        results = backtest["results"] or {}
        v8_sid = _find_strategy_id_in_results(results, r"\bv8\b")
        if v8_sid is None:
            continue
        found_any = True
        params = backtest["params"] or {}
        symbol_count = len(params.get("symbols") or [])
        trade_count = len(results[v8_sid].get("pairs") or [])
        date_range = f"{params.get('start_date')} to {params.get('end_date')}"
        print(f"{summary['id']:>6}  {date_range:<23}  {symbol_count:>7}  {trade_count:>9}  {summary['created_at']}")
    if not found_any:
        print("No 'done' backtest in this account's history carries an ORB Long V8 result yet.")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--list", action="store_true",
                         help="List every 'done' backtest that carries an ORB Long V8 result (with its id, date "
                              "range, symbol/trade counts) and exit - no raw SQL needed to find a --backtest-id.")
    parser.add_argument("--backtest-id", default=None,
                         help="One id, or a comma-separated list of ids (e.g. 870,871,872), of already-finished "
                              "backtest(s) whose own V8 result to read and pool together (recommended on the "
                              "production server - see docs/worker.md). Mutually exclusive with --start-date/"
                              "--end-date, which run the simulation right here instead.")
    parser.add_argument("--start-date", help="YYYY-MM-DD (ignored with --backtest-id)")
    parser.add_argument("--end-date", help="YYYY-MM-DD (ignored with --backtest-id)")
    parser.add_argument("--account-id", type=int, default=None)
    parser.add_argument("--portfolio-value", type=float, default=DEFAULT_PORTFOLIO_VALUE)
    parser.add_argument("--max-risk-pct", type=float, default=DEFAULT_MAX_RISK_PCT)
    parser.add_argument("--max-trades-per-day", type=int, default=DEFAULT_MAX_TRADES_PER_DAY)
    parser.add_argument("--commission-per-trade", type=float, default=DEFAULT_COMMISSION_PER_TRADE)
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Path to write the .xlsx report to")
    args = parser.parse_args()

    db.init_db(seed_rules_path=PROJECT_DIR / "rules.json")
    account_id = args.account_id if args.account_id is not None else db.get_default_account_id()

    if args.list:
        list_v8_backtests(account_id)
        return

    if args.backtest_id is None and not (args.start_date and args.end_date):
        raise SystemExit("Either --backtest-id, --list, or both --start-date and --end-date, are required.")

    # hard_stop_R/new_hard_stop_R come straight off V8's own rules_json
    # regardless of which path below is taken - read-only, never modified,
    # and the SAME preset a remote-worker backtest would have used too
    # (rules_json is never overridden per-run for a plain multi-strategy
    # backtest, only for an Optimization Lab sweep - not this script).
    v8_strategy = find_strategy(account_id, V8_STRATEGY_NEEDLE)
    v8_rules = json.loads(db.get_strategy(v8_strategy["id"])["rules_json"])
    drr_cfg = v8_rules.get("exit", {}).get("dynamic_risk_reduction") or {}
    orig_r = -float(v8_rules["exit"]["hard_stop_R"])
    adjusted_r = -float(drr_cfg["new_hard_stop_R"])

    if args.backtest_id is not None:
        backtest_ids = [int(x) for x in str(args.backtest_id).split(",") if x.strip()]
        section(f"V6 Risk Event Audit — reading {len(backtest_ids)} backtest(s) (no simulation runs here)")
        pairs, scope_label = _load_pairs_from_backtests(backtest_ids)
    else:
        start_date = date.fromisoformat(args.start_date)
        end_date = date.fromisoformat(args.end_date)
        direction = v8_strategy["direction"]
        symbols = backtest_data.cached_symbols(backtest_engine.BAR_SIZE)
        if not symbols:
            raise SystemExit("No symbols have cached historical bars yet - run fetch_backtest_data.py first.")

        scope_label = f"{v8_strategy['name']} | {start_date} to {end_date} | {len(symbols)} symbol(s)"
        section(f"V6 Risk Event Audit — {scope_label}")
        print("WARNING: this runs V8's full simulation LOCALLY, right now - do NOT do this on the small "
              "production server (see docs/worker.md). Use --backtest-id with a remote-worker result instead.")
        print(f"Running {v8_strategy['name']} (unmodified, as already configured)...")
        result = _run(
            v8_strategy["name"], direction, v8_rules, symbols, start_date, end_date,
            args.portfolio_value, args.max_risk_pct, args.max_trades_per_day, args.commission_per_trade,
        )
        pairs = result["pairs"]

    print(f"{len(pairs)} V8 trade(s) in this scope.")

    rows = build_audit_rows(pairs)
    summary = build_summary(rows)

    section("V6 AUDIT SUMMARY")
    for k, v in summary.items():
        print(f"  {k:<45s} {v}")

    print_final_answer(rows, summary, orig_r, adjusted_r)

    export_xlsx(rows, summary, orig_r, adjusted_r, scope_label, args.output)
    print(f"\nWritten: {args.output}")


if __name__ == "__main__":
    main()
