"""Trade Telemetry Dashboard - a PASSIVE, read-only research framework that
collects, stores and analyzes technical-indicator behavior in the minutes
right after a backtest trade's own entry, to help spot early warning signs
that distinguish winners from losers, hard-stop trades, poor-Capture-%
trades, and trades that never reach trailing activation.

NEVER read by any strategy's entry/exit decision, position sizing, risk
model, hard-stop model, or trailing algorithm - this module only ever runs
AFTER a backtest has already produced its own trades (see generate_
telemetry_for_backtest, always called from a "Generate Telemetry" action
against an already-finished backtest row), purely to observe and report.
Nothing here can change what a strategy already did.

Architecture: generic over any backtest whose strategy is ORB-shaped (a
rules_json carrying "opening_range" - see backtest_runner.run_one_strategy's
own dispatch) or, for every non-structural indicator (RSI/MACD/Stochastic/
ADX/VWAP/EMA/volume/volatility/market-context/trade-progress), over ANY
strategy at all - a symbol, an entry timestamp/price/side and an initial
risk width (from perf.initial_risk_per_share) are all this module actually
needs. Only the "Trade Structure Events" section (opening-range/breakout/
retest references) is ORB-specific and is simply omitted (None) for a
non-ORB strategy. Attaching this to V4/V4.1/V4.3 or a future ORB variant
therefore needs no changes here at all - only a different strategy_id/
backtest_id passed into generate_telemetry_for_backtest.

Snapshot resolution: this module's only price/volume source is the SAME
5-minute intraday bar cache backtest_engine.py itself reads (src.
backtest_data) - there is no separate finer-grained data pipeline. Every
snapshot offset (+1m/+3m/+5m/+10m/+15m) resolves to "the last already-
closed 5-minute bar at or before entry_time + offset" (see _snapshot_bar) -
a +1m or +3m snapshot is very often IDENTICAL to the entry snapshot's own
bar (not yet a full 5-minute bar has closed since entry), which is the
honest answer at this data resolution, not a bug. Every stored snapshot
carries its own real bar timestamp and `minutes_elapsed` (which can differ
from the nominal offset for exactly this reason) so nothing here silently
overstates its own precision - see the "adapt to the existing 5-minute
resolution" scoping decision this feature was built against.
"""
from __future__ import annotations

import json
from datetime import date, datetime

import numpy as np
import pandas as pd

from src import backtest_data, backtest_engine, db, es_filter, orb, perf, technicals

SNAPSHOT_OFFSETS = [("entry", 0), ("1m", 1), ("3m", 3), ("5m", 5), ("10m", 10), ("15m", 15)]
_NON_ENTRY_LABELS = [label for label, _ in SNAPSHOT_OFFSETS if label != "entry"]
_OFFSET_MINUTES = dict(SNAPSHOT_OFFSETS)  # "10m" -> 10, "15m" -> 15, ... - evaluate_rule's own
# rule_evaluation_timestamp lookup (see _rule_applicable_mask) needs the nominal
# offset in minutes for whichever offset a given rule evaluates at.

# Same EWM-convergence reasoning as src/entry_metrics.py's own _EWM_WARMUP_
# BARS (RSI(14)/EMA(9/20/50)/MACD/ADX all need enough trailing history to
# have actually converged) - an independent constant rather than importing
# entry_metrics' own private one, since this module's own indicator set
# (MACD/Stochastic/ADX/ATR/OBV) is entirely its own.
_WARMUP_BARS = 400

# Same "typical day" trailing-lookback convention entry_metrics.py's own
# _volume_multiple_vs_avg_daily already uses, reused here for RVOL (via
# orb._compute_rvol, the SAME formula the live strategy's own I3 filter and
# entry_metrics.py's own "rvol" already use - not a new invented one) and
# for this module's own "Volume vs Average" proxy.
RVOL_LOOKBACK_DAYS = 20

MARKET_CONTEXT_SYMBOLS = {"es": es_filter.ES_SYMBOL, "spy": "SPY", "qqq": "QQQ"}

# Default trade-grouping predicates for the Analysis Module / Comparison
# Engine (see the spec's own "Allow grouping trades into" section) - each
# takes one flattened trade record (see flatten_trades) and an optional
# cfg dict of thresholds. Deliberately kept as simple, named, independently
# togglable predicates over the SAME trade set rather than one rigid
# mutually-exclusive classification, since the spec's own "Winners"/
# "Losers" examples already aren't exhaustive (e.g. a trade closing exactly
# at breakeven falls in neither) - a dashboard can combine any of these
# (e.g. "Hard Stop" AND "Losers") rather than being forced to pick one.
def _early_failure_candidate(t, cfg) -> bool:
    """Research-only DERIVED group - "Trades that would have triggered the
    proposed 10-minute Early Failure rule." This is purely a retrospective
    LABEL computed from already-recorded +10m telemetry (see SNAPSHOT_
    OFFSETS/_indicator_snapshot/_structure_snapshot/_derive_since_entry_
    fields) - it does NOT touch, gate, or feed back into any strategy's
    own entry/exit logic, backtest run, or stored results in any way; a
    trade already closed exactly the way its own backtest run decided,
    long before this label is ever computed.

    Membership requires ALL of, evaluated strictly at the +10m snapshot:
      - trailing had NOT activated YET as of +10m (see build_trade_
        telemetry's own "trail_activated_as_of" - a POINT-IN-TIME fact,
        deliberately NOT the trade's overall trail_activated field: that
        field is true for a trade's entire life once it ever activates,
        even if that happens well after +10m, which would make "not yet
        trailing at +10m" indistinguishable from "never trails at all" -
        and silently make this rule unable to ever flag a trade that
        later recovers and becomes a Trailing Winner, exactly the
        trades the Rule Evaluation tab's own Net Benefit Engine needs to
        find to price the cost of a false-positive early exit)
      - 10m current_r <= -0.40R
      - 10m RSI Delta <= -5
      - EITHER 10m Returned Inside Opening Range OR 10m Lost EMA9

    `t.get(...)` (not `t[...]`) throughout - a trade whose own +10m
    snapshot came back None (see _snapshot_bar's own "no bar yet" case)
    simply has no 10m_* columns to read at all for THAT row; missing
    numeric values as NaN safely fail every "<=" comparison below (never
    raise), and the boolean checks use "== 1"/"!= 0" rather than bool(...)
    for the same reason - flatten_trades stores a bool snapshot field as
    int 0/1, and bool(float('nan')) is True in Python, which would
    otherwise silently treat "unknown" as "condition met"."""
    if t.get("10m_trail_activated_as_of") != 0:
        return False
    current_r = t.get("10m_current_r")
    rsi_delta = t.get("10m_rsi_delta")
    if current_r is None or not (current_r <= -0.40):
        return False
    if rsi_delta is None or not (rsi_delta <= -5):
        return False
    return t.get("10m_returned_inside_opening_range") == 1 or t.get("10m_lost_ema9") == 1


def _early_failure_v2(t, cfg) -> bool:
    """Rule B - "Early Failure V2" (OR -> AND version): identical to
    _early_failure_candidate (Rule A / "Early Failure V1") except the two
    structural-failure conditions must BOTH hold rather than either -
    tests whether requiring both increases Precision (see the "Add
    Additional Early Failure Rule Evaluations" spec's own Rule B)."""
    if t.get("10m_trail_activated_as_of") != 0:
        return False
    current_r = t.get("10m_current_r")
    rsi_delta = t.get("10m_rsi_delta")
    if current_r is None or not (current_r <= -0.40):
        return False
    if rsi_delta is None or not (rsi_delta <= -5):
        return False
    return t.get("10m_returned_inside_opening_range") == 1 and t.get("10m_lost_ema9") == 1


def _early_failure_v3(t, cfg) -> bool:
    """Rule C - "Early Failure V3" (structural failure version): same
    momentum gate as V1/V2 (not yet trailing, Current R <= -0.40R, RSI
    Delta <= -5), Returned Inside Opening Range REQUIRED (not just one of
    two options), plus a real structure break (Broke Retest Low OR Broke
    Breakout Candle Low) - tests whether stronger structure failure
    improves results."""
    if t.get("10m_trail_activated_as_of") != 0:
        return False
    current_r = t.get("10m_current_r")
    rsi_delta = t.get("10m_rsi_delta")
    if current_r is None or not (current_r <= -0.40):
        return False
    if rsi_delta is None or not (rsi_delta <= -5):
        return False
    if t.get("10m_returned_inside_opening_range") != 1:
        return False
    return t.get("10m_broke_retest_low") == 1 or t.get("10m_broke_breakout_candle_low") == 1


def _early_failure_v4(t, cfg) -> bool:
    """Rule D - "Early Failure V4" (delayed confirmation version): a
    Stage 1 WARNING at +10m (not yet trailing, Current R <= -0.40R, RSI
    Delta <= -5, Returned Inside Opening Range = True) must be followed by
    a Stage 2 CONFIRMATION at +15m (still not yet trailing, Current R <=
    -0.75R, RSI Delta <= -8, Price Below EMA9, no new intraday high since
    entry) - gives a recovering trade an extra 5 minutes before flagging
    it, rather than acting on the +10m warning alone. "Price Below EMA9"
    reuses the already-stored above_ema9 snapshot field (== 0, i.e. NOT
    above - same strict '>' comparison _indicator_snapshot already applies
    everywhere else in this module) rather than a new stored field.
    evaluation_offset is "15m" (see RULES) - the trade must have still
    been open at +15m for this rule to even be evaluable at all, exactly
    what Fix #2/_rule_applicable_mask is for."""
    if t.get("10m_trail_activated_as_of") != 0:
        return False
    r10 = t.get("10m_current_r")
    rsi10 = t.get("10m_rsi_delta")
    if r10 is None or not (r10 <= -0.40):
        return False
    if rsi10 is None or not (rsi10 <= -5):
        return False
    if t.get("10m_returned_inside_opening_range") != 1:
        return False

    if t.get("15m_trail_activated_as_of") != 0:
        return False
    r15 = t.get("15m_current_r")
    rsi15 = t.get("15m_rsi_delta")
    if r15 is None or not (r15 <= -0.75):
        return False
    if rsi15 is None or not (rsi15 <= -8):
        return False
    if t.get("15m_above_ema9") != 0:  # "Price Below EMA9 = True"
        return False
    return t.get("15m_new_intraday_high_since_entry") == 0  # "New Intraday High Since Entry = False"


def _v5_warning_at_10m(t, cfg) -> bool:
    """Rule E ("Early Failure V5: Recovery-Aware") Stage 1 - the exact
    same +10m warning condition V4 shares (not yet trailing, Current R
    <= -0.40R, RSI Delta <= -5, Returned Inside Opening Range = True), but
    factored out so both the V5 predicate and its own missing-data check
    (_early_failure_v5_missing_data below) agree on when a trade even
    REACHES Stage 2 - a trade that never warned at 10m has nothing to be
    "missing data" about at 15m."""
    if t.get("10m_trail_activated_as_of") != 0:
        return False
    r10 = t.get("10m_current_r")
    if r10 is None or not (r10 <= -0.40):
        return False
    rsi10 = t.get("10m_rsi_delta")
    if rsi10 is None or not (rsi10 <= -5):
        return False
    return t.get("10m_returned_inside_opening_range") == 1


def _early_failure_v5(t, cfg) -> bool:
    """Rule E - "Early Failure V5" (recovery-aware version): a +10m
    warning (see _v5_warning_at_10m) only converts to a flag if, by +15m,
    the trade has NOT begun recovering - both R Recovery (15m Current R
    minus 10m Current R) and RSI Recovery (15m RSI minus 10m RSI, the raw
    indicator value, not RSI Delta) must be <= 0, alongside stricter -0.75R/
    -8 RSI thresholds and price still below EMA9 at +15m. A trade that
    starts improving between +10m and +15m is deliberately spared -
    that's the whole point of this rule versus V4's simpler two-stage
    check (V4 never looks at the DIRECTION of the 10m->15m move, only at
    the +15m snapshot's own absolute values). evaluation_offset is "15m" -
    see RULES."""
    if not _v5_warning_at_10m(t, cfg):
        return False
    if t.get("15m_trail_activated_as_of") != 0:
        return False
    r15 = t.get("15m_current_r")
    if r15 is None or not (r15 <= -0.75):
        return False
    rsi15_delta = t.get("15m_rsi_delta")
    if rsi15_delta is None or not (rsi15_delta <= -8):
        return False
    r10 = t.get("10m_current_r")
    rsi10_raw, rsi15_raw = t.get("10m_rsi"), t.get("15m_rsi")
    if r10 is None or rsi10_raw is None or rsi15_raw is None:
        return False
    r_recovery = r15 - r10
    if not (r_recovery <= 0):
        return False
    rsi_recovery = rsi15_raw - rsi10_raw
    if not (rsi_recovery <= 0):
        return False
    return t.get("15m_above_ema9") == 0


def _early_failure_v5_missing_data(t, cfg) -> bool:
    """A trade that already warned at +10m (see _v5_warning_at_10m) but
    whose +15m snapshot is missing one of the fields the Recovery
    calculation needs (e.g. the cached bar history simply ends before
    +15m, a real if rare data gap - NOT the same as the trade having
    chronologically already exited, which _rule_applicable_mask already
    excludes separately) must not be silently scored as "did not flag" -
    it genuinely could not be evaluated. A trade that never warned at all
    has nothing here to be missing."""
    if not _v5_warning_at_10m(t, cfg):
        return False
    required = ("10m_current_r", "15m_current_r", "10m_rsi", "15m_rsi", "15m_rsi_delta", "15m_above_ema9")
    # pd.isna(), not `is None` - a DataFrame row (the actual `t` this is
    # called with via evaluate_rule's df.apply) represents a missing
    # numeric cell as float NaN, never Python None (see _early_failure_
    # candidate's own docstring on this exact distinction) - `is None`
    # alone would silently never catch a real gap here.
    return any(pd.isna(t.get(k)) for k in required)


def _early_failure_v6(t, cfg) -> bool:
    """Rule F - "Early Failure V6" (no positive progress): a trade that,
    by +10m, is both losing (Current R <= -0.40R, RSI Delta <= -5,
    structurally weak - Returned Inside Opening Range AND Lost EMA9) AND
    never showed meaningful favorable movement at all (MFE R So Far <=
    +0.30R, i.e. it was never even briefly ahead by more than 0.30R). 10m-
    only (evaluation_offset "10m" - see RULES)."""
    if t.get("10m_trail_activated_as_of") != 0:
        return False
    r10 = t.get("10m_current_r")
    if r10 is None or not (r10 <= -0.40):
        return False
    rsi10 = t.get("10m_rsi_delta")
    if rsi10 is None or not (rsi10 <= -5):
        return False
    mfe10 = t.get("10m_mfe_r_so_far")
    if mfe10 is None or not (mfe10 <= 0.30):
        return False
    if t.get("10m_returned_inside_opening_range") != 1:
        return False
    return t.get("10m_lost_ema9") == 1


def _early_failure_v6_missing_data(t, cfg) -> bool:
    """10m_mfe_r_so_far is required for V6's own "never showed favorable
    movement" condition - a trade chronologically open at +10m whose own
    10m snapshot nonetheless has no mfe_r_so_far (a real data gap, e.g. no
    risk width could be computed for that trade - see _indicator_snapshot,
    where current_r/mfe_r_so_far/mae_r_so_far are all gated on the same
    `risk` value) must be excluded, never silently treated as "no
    favorable movement" (which would wrongly favor flagging it). pd.isna()
    - see _early_failure_v5_missing_data's own comment on why `is None`
    alone is not enough for a pandas DataFrame row."""
    return bool(pd.isna(t.get("10m_mfe_r_so_far")))


def _v7_progress_ratio(t):
    """Rule G ("Early Failure V7: Progress Ratio")'s own MFE/MAE ratio -
    factored out so the predicate and its own missing-data check
    (_early_failure_v7_missing_data) read it identically. None (not 0 or
    any other sentinel) when either input is missing - "Do not impute
    missing values" per the rule's own spec."""
    mfe10 = t.get("10m_mfe_r_so_far")
    mae10 = t.get("10m_mae_r_so_far")
    if pd.isna(mfe10) or pd.isna(mae10):
        return None
    return mfe10 / max(mae10, 0.05)


def _early_failure_v7(t, cfg) -> bool:
    """Rule G - "Early Failure V7" (progress ratio): like V6, a losing/
    RSI-weak trade at +10m (Current R <= -0.40R, RSI Delta <= -5,
    Returned Inside Opening Range = True), but instead of an absolute MFE
    cap, compares favorable progress AGAINST adverse movement: Progress
    Ratio = 10m MFE R So Far / max(10m MAE R So Far, 0.05) <= 0.30 - the
    0.05 floor avoids a division blow-up for a trade with almost no
    adverse excursion yet. Not rounded before the comparison (see _v7_
    progress_ratio's own docstring) - only _round() at display/export
    time, same convention _round() itself documents throughout this
    module. evaluation_offset "10m" - see RULES."""
    if t.get("10m_trail_activated_as_of") != 0:
        return False
    r10 = t.get("10m_current_r")
    if r10 is None or not (r10 <= -0.40):
        return False
    rsi10 = t.get("10m_rsi_delta")
    if rsi10 is None or not (rsi10 <= -5):
        return False
    ratio = _v7_progress_ratio(t)
    if ratio is None or not (ratio <= 0.30):
        return False
    return t.get("10m_returned_inside_opening_range") == 1


def _early_failure_v7_missing_data(t, cfg) -> bool:
    """Same reasoning as _early_failure_v6_missing_data - 10m_mfe_r_so_far
    and 10m_mae_r_so_far are both required to even compute a Progress
    Ratio at all; a trade missing either must be excluded, not scored as
    if Progress Ratio were 0 (which would wrongly always flag it) or
    infinite (which would wrongly never flag it). pd.isna() - see
    _early_failure_v5_missing_data's own comment on why `is None` alone is
    not enough for a pandas DataFrame row."""
    return bool(pd.isna(t.get("10m_mfe_r_so_far")) or pd.isna(t.get("10m_mae_r_so_far")))


DEFAULT_GROUPS = {
    "winners": lambda t, cfg: t["final_r"] is not None and t["final_r"] > 0,
    "losers": lambda t, cfg: t["final_r"] is not None and t["final_r"] < 0,
    "hard_stop": lambda t, cfg: t["exit_reason"] == "hard_stop",
    "trailing_winners": lambda t, cfg: bool(t["trail_activated"]) and t["final_r"] is not None and t["final_r"] > 0,
    "never_trailed": lambda t, cfg: not bool(t["trail_activated"]),
    "capture_disaster": lambda t, cfg: (
        t["capture_pct"] is not None and t["capture_pct"] < cfg.get("capture_disaster_threshold", -1000)
    ),
    "early_failure_candidate": _early_failure_candidate,
}

# Rule Evaluation tab registry (see evaluate_rule below) - deliberately
# separate from DEFAULT_GROUPS/apply_group even though "Early Failure
# Candidate" is also a Compare Groups option: Compare Groups treats it as
# just another cohort to line up against a second cohort of the analyst's
# own choosing (Group A/B), while Rule Evaluation treats it as a proposed
# DECISION RULE with one fixed ground truth (did the trade actually become
# a Hard Stop) - a confusion matrix, precision/recall, and a Net Benefit
# R/$ estimate, not a side-by-side comparison. Same predicate function
# either way (single source of truth for what "triggers" means), but the
# two tabs ask genuinely different questions of it. "evaluation_offset"
# names which snapshot offset's own current_r/rsi_delta the rule's
# Early Exit R and Net Benefit accounting read - not hardcoded to "10m"
# in evaluate_rule itself, so a future rule can be defined against a
# different offset (e.g. a 5-minute rule) by adding an entry here alone.
RULES = {
    "early_failure_v1": {
        "label": "Early Failure V1",
        "description": (
            "At 10 minutes: not yet trailing, Current R <= -0.40R, RSI Delta <= -5, "
            "AND (Returned Inside Opening Range OR Lost EMA9)."
        ),
        "predicate": _early_failure_candidate,
        "evaluation_offset": "10m",
    },
    "early_failure_v2": {
        "label": "Early Failure V2",
        "description": (
            "OR -> AND version of V1: at 10 minutes, not yet trailing, Current R <= -0.40R, "
            "RSI Delta <= -5, AND Returned Inside Opening Range = True AND Lost EMA9 = True. "
            "Tests whether requiring BOTH structural failures increases Precision."
        ),
        "predicate": _early_failure_v2,
        "evaluation_offset": "10m",
    },
    "early_failure_v3": {
        "label": "Early Failure V3",
        "description": (
            "Structural failure version: at 10 minutes, not yet trailing, Current R <= -0.40R, "
            "RSI Delta <= -5, Returned Inside Opening Range = True, AND (Broke Retest Low OR "
            "Broke Breakout Candle Low). Tests whether stronger structure failure improves results."
        ),
        "predicate": _early_failure_v3,
        "evaluation_offset": "10m",
    },
    "early_failure_v4": {
        "label": "Early Failure V4",
        "description": (
            "Delayed confirmation version: a Stage 1 warning at 10 minutes (not yet trailing, "
            "Current R <= -0.40R, RSI Delta <= -5, Returned Inside Opening Range = True) must be "
            "followed by a Stage 2 exit confirmation at 15 minutes (still not yet trailing, "
            "Current R <= -0.75R, RSI Delta <= -8, Price Below EMA9 = True, New Intraday High "
            "Since Entry = False). Allows recovering trades more time before flagging."
        ),
        "predicate": _early_failure_v4,
        "evaluation_offset": "15m",
    },
    "early_failure_v5": {
        "label": "Early Failure V5",
        "description": (
            "Recovery-aware version: a 10-minute warning (not yet trailing, Current R <= -0.40R, "
            "RSI Delta <= -5, Returned Inside Opening Range = True) only confirms at 15 minutes if "
            "the trade has NOT begun recovering (R Recovery and RSI Recovery from 10m to 15m both "
            "<= 0), Current R <= -0.75R, RSI Delta <= -8, and price still below EMA9. Tests whether "
            "waiting and checking the DIRECTION of the 10m->15m move avoids cutting recovering trades."
        ),
        "predicate": _early_failure_v5,
        "evaluation_offset": "15m",
        "missing_data_check": _early_failure_v5_missing_data,
    },
    "early_failure_v6": {
        "label": "Early Failure V6",
        "description": (
            "No positive progress: at 10 minutes, not yet trailing, Current R <= -0.40R, RSI Delta "
            "<= -5, Returned Inside Opening Range = True, Lost EMA9 = True, AND the trade never "
            "showed meaningful favorable movement at all (10m MFE R So Far <= +0.30R). Tests whether "
            "a trade with no positive progress at all is more likely to become a Hard Stop."
        ),
        "predicate": _early_failure_v6,
        "evaluation_offset": "10m",
        "missing_data_check": _early_failure_v6_missing_data,
    },
    "early_failure_v7": {
        "label": "Early Failure V7",
        "description": (
            "Progress ratio: at 10 minutes, not yet trailing, Current R <= -0.40R, RSI Delta <= -5, "
            "Returned Inside Opening Range = True, AND Progress Ratio (10m MFE R So Far divided by "
            "max(10m MAE R So Far, 0.05)) <= 0.30. Compares favorable progress against adverse "
            "movement rather than using either one alone."
        ),
        "predicate": _early_failure_v7,
        "evaluation_offset": "10m",
        "missing_data_check": _early_failure_v7_missing_data,
    },
}

# Trade-level columns flatten_trades always carries alongside the per-
# snapshot metric columns - never treated as a predictive FEATURE by
# predictive_ranking/suggested_candidate_filters (they're the labels being
# predicted, or identifying metadata, not indicator inputs).
NON_FEATURE_COLUMNS = {
    "telemetry_id", "backtest_id", "strategy_id", "strategy_name", "symbol", "side",
    "entry_time", "exit_time", "final_r", "exit_reason", "trail_activated", "capture_pct", "risk_dollars",
}

# "1m"/"3m" are excluded from every STATISTICAL computation (Comparison,
# Predictive Ranking/Feature Importance/Mutual Information, Suggested
# Filters, and the heatmap metric picker - see feature_columns below) -
# NOT from storage or the per-trade drill-down view, which still show all
# 6 snapshots. Reason: at this module's own 5-minute bar resolution (see
# the module docstring's own "Snapshot resolution" section), a +1m or +3m
# target very often resolves to the SAME bar as the entry snapshot itself
# - sometimes it's a real, later bar (an entry near the end of its own
# 5-minute window), sometimes it's an exact duplicate of the entry row,
# and there is no way to tell which case a given row is in without
# checking its own minutes_elapsed by hand. Feeding that inconsistent mix
# into a Random Forest/Mutual Information comparison risks (a) implying a
# real measurement was taken 1/3 minutes after entry when it frequently
# wasn't, and (b) padding the feature set with near-duplicates of the
# entry columns, diluting the entry columns' own measured importance
# rather than adding real signal. +5m/+10m/+15m don't have this problem -
# a 5-minute bar has always fully closed by then.
EXCLUDED_SNAPSHOT_PREFIXES = ("1m_", "3m_")

# Same wording shown on the /telemetry page (next to the analysis tabs)
# and written into the Analysis Excel export's own Summary sheet - one
# string, so the explanation can never drift between the two surfaces.
EXCLUDED_SNAPSHOT_NOTE = (
    "1m and 3m telemetry snapshots are unavailable because the backtest data resolution "
    "does not support these offsets. They are excluded from all statistical analysis."
)


def _round(value, digits=4):
    if value is None:
        return None
    try:
        f = float(value)
    except (TypeError, ValueError):
        return None
    if np.isnan(f) or np.isinf(f):
        return None
    return round(f, digits)


def _bool_or_none(value) -> bool | None:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None
    return bool(value)


def _normalize_bars(bars: pd.DataFrame) -> pd.DataFrame:
    """Defensive read-time normalization for a cached bars file this
    module doesn't itself own - dedupes by timestamp (keep last, same
    convention src.backtest_data.save_cached_bars already applies on
    WRITE, but an older cache file can predate that logic, or a bug
    elsewhere could still write one) and coerces OHLCV to numeric
    (coercing an unparseable value to NaN rather than leaving a column
    object-dtype). Both a duplicate-timestamp index and an object-dtype
    OHLCV column make pandas' own EWM/rolling series calls (RSI/EMA/MACD/
    ADX/ATR/OBV) raise a hard-to-diagnose "DataError: No numeric types to
    aggregate" - confirmed by direct testing, not a guess. Every other
    reader of this same cache (backtest_engine.py) only ever replays
    narrow, single-day-sliced windows and is far less likely to ever
    surface either issue than this module, which is the first to compute
    indicator series across a symbol's ENTIRE cached history in one call."""
    bars = bars[~bars.index.duplicated(keep="last")].sort_index()
    for col in ("Open", "High", "Low", "Close", "Volume"):
        bars[col] = pd.to_numeric(bars[col], errors="coerce")
    return bars


def _session_vwap(bars: pd.DataFrame) -> pd.Series:
    """Session VWAP over `bars`' full multi-day span, reset every trading
    day - reuses orb._compute_vwap_series (session-scoped) per calendar
    day and concatenates, the same convention backtest_engine.py's own
    per-day VWAP use already establishes."""
    return pd.concat([orb._compute_vwap_series(day_bars) for _, day_bars in bars.groupby(bars.index.date)])


def _coerce_numeric_frame(frame: pd.DataFrame) -> pd.DataFrame:
    """Final defensive pass over an assembled indicator frame - coerces
    every column to numeric (pd.to_numeric, errors="coerce"), converting
    any lingering pd.NA/object-dtype column back to clean float64 NaN.
    Needed because orb._compute_vwap_series (reused here, not owned by
    this module) divides by cum_vol.replace(0, pd.NA) - which itself
    degrades that column to object dtype in current pandas (same
    DataError: No numeric types to aggregate technicals.py's own
    replace(0, np.nan) fix was for) - a plain float64 column is a no-op
    here, so this is always safe to apply."""
    for col in frame.columns:
        frame[col] = pd.to_numeric(frame[col], errors="coerce")
    return frame


def _indicator_frame(bars: pd.DataFrame) -> pd.DataFrame:
    """Every per-bar indicator this module tracks for a TRADED symbol,
    computed ONCE over its full continuous multi-day bar history (not per
    trade/snapshot) - same "compute the series, not per query" discipline
    orb.py's own RVOL cache already established. Aligned to bars.index."""
    closes, highs, lows, volumes = bars["Close"], bars["High"], bars["Low"], bars["Volume"]
    frame = pd.DataFrame(index=bars.index)
    frame["close"], frame["high"], frame["low"], frame["volume"] = closes, highs, lows, volumes
    frame["rsi"] = orb._compute_rsi_series(closes, 14)
    frame["ema9"] = orb._compute_ema_series(closes, 9)
    frame["ema20"] = orb._compute_ema_series(closes, 20)
    frame["ema50"] = orb._compute_ema_series(closes, 50)
    macd = technicals.macd_series(closes)
    frame["macd"], frame["macd_signal"], frame["macd_hist"] = macd["macd"], macd["signal"], macd["histogram"]
    stoch = technicals.stochastic_series(highs, lows, closes)
    frame["stoch_k"], frame["stoch_d"] = stoch["percent_k"], stoch["percent_d"]
    adx = technicals.adx_series(highs, lows, closes)
    frame["adx"], frame["plus_di"], frame["minus_di"] = adx["adx"], adx["plus_di"], adx["minus_di"]
    frame["atr"] = technicals.atr_series(highs, lows, closes)
    frame["obv"] = technicals.obv_series(closes, volumes)
    frame["vwap"] = _session_vwap(bars)
    return _coerce_numeric_frame(frame)


def _market_context_frame(bars: pd.DataFrame, include_adx: bool) -> pd.DataFrame:
    """The lighter indicator set the spec wants for ES/SPY/QQQ market
    context (Price, RSI, [ADX for ES only], Above VWAP, Distance From
    VWAP) - a subset of _indicator_frame, kept separate so a market-context
    symbol's cache doesn't pay for MACD/Stochastic/OBV it's never asked for."""
    closes = bars["Close"]
    frame = pd.DataFrame(index=bars.index)
    frame["price"] = closes
    frame["rsi"] = orb._compute_rsi_series(closes, 14)
    frame["vwap"] = _session_vwap(bars)
    if include_adx:
        frame["adx"] = technicals.adx_series(bars["High"], bars["Low"], closes)["adx"]
    return _coerce_numeric_frame(frame)


def _snapshot_bar(frame: pd.DataFrame, entry_ts, offset_minutes: int):
    """The last row of `frame` at or before entry_ts + offset_minutes - see
    this module's own docstring on why this, not an exact-offset lookup, is
    the honest resolution for 5-minute data. None if no bar exists yet at
    or before that target (e.g. a very late-session entry's own +15m
    target running past the last cached bar)."""
    target = entry_ts + pd.Timedelta(minutes=offset_minutes)
    eligible = frame[frame.index <= target]
    return eligible.iloc[-1] if not eligible.empty else None


def _market_context_snapshot(mkt_frame: pd.DataFrame | None, entry_ts, offset_minutes: int) -> dict:
    empty = {"price": None, "rsi": None, "adx": None, "above_vwap": None, "dist_vwap_pct": None}
    if mkt_frame is None or mkt_frame.empty:
        return empty
    row = _snapshot_bar(mkt_frame, entry_ts, offset_minutes)
    if row is None:
        return empty
    price, vwap = float(row["price"]), row.get("vwap")
    has_vwap = vwap is not None and not pd.isna(vwap)
    return {
        "price": _round(price, 4),
        "rsi": _round(row.get("rsi"), 2),
        "adx": _round(row.get("adx"), 2) if "adx" in mkt_frame.columns else None,
        "above_vwap": _bool_or_none(price > vwap) if has_vwap else None,
        "dist_vwap_pct": _round((price - vwap) / vwap * 100, 3) if has_vwap and vwap else None,
    }


def _confirm_bar_low(today_bars_upto_entry: pd.DataFrame, or_high: float, model: str | None):
    """The breakout/confirm candle's own Low - for "breakout" model this is
    the entry bar itself (breakout only ever fires on current_ts ==
    confirm_ts, see orb.evaluate_orb_entry); for "retest" it's the earlier
    bar whose Close first cleared or_high, found the same way entry_
    metrics.py's own confirm_bar detection already does (not imported from
    there - that function computes several OTHER things this module
    doesn't need, and duplicating just the bar-lookup here keeps this
    module independent of entry_metrics' own ORB-Long-only assumptions).
    Long-only (or_high, not or_low) since every strategy this feature is
    scoped to (V4/V4.1/V4.2/V4.3) is ORB Long. None if today_bars_upto_
    entry is empty."""
    if today_bars_upto_entry.empty:
        return None, None
    if model == "breakout":
        confirm = today_bars_upto_entry.iloc[-1]
        return confirm.name, float(confirm["Low"])
    or_range = orb.compute_opening_range(today_bars_upto_entry)
    post_or = today_bars_upto_entry[today_bars_upto_entry.index > or_range["or_end_ts"]] if or_range else today_bars_upto_entry.iloc[0:0]
    candidates = post_or[post_or["Close"] > or_high]
    confirm = candidates.iloc[0] if not candidates.empty else today_bars_upto_entry.iloc[-1]
    return confirm.name, float(confirm["Low"])


def _structure_context(today_bars: pd.DataFrame, entry_ts, model: str | None) -> dict:
    """Everything _structure_snapshot needs that's constant across a
    trade's own snapshots (opening range, breakout candle low, retest low)
    - computed ONCE per trade, not per snapshot. None fields throughout if
    the opening range hasn't even formed by entry_ts (shouldn't happen for
    a real ORB entry, but a defensively-handled edge case, not an error)."""
    today_upto_entry = today_bars[today_bars.index <= entry_ts]
    or_range = orb.compute_opening_range(today_upto_entry)
    if or_range is None:
        return {"or_high": None, "or_low": None, "confirm_low": None, "retest_low": None}
    confirm_ts, confirm_low = _confirm_bar_low(today_upto_entry, or_range["or_high"], model)
    retest_low = None
    if model == "retest" and confirm_ts is not None:
        retest_bars = today_upto_entry[(today_upto_entry.index > confirm_ts)]
        retest_low = float(retest_bars["Low"].min()) if not retest_bars.empty else confirm_low
    return {"or_high": or_range["or_high"], "or_low": or_range["or_low"], "confirm_low": confirm_low, "retest_low": retest_low}


def _structure_snapshot(struct_ctx: dict, path: pd.DataFrame, bar_row) -> dict:
    """Trade Structure Events at one snapshot - `path` is the traded
    symbol's own OHLC frame from entry through this snapshot's bar
    (inclusive), used for the "since entry" running-high tracking."""
    if struct_ctx["or_high"] is None:
        return {
            "returned_inside_opening_range": None, "broke_breakout_candle_low": None,
            "broke_retest_low": None, "new_intraday_high_since_entry": None, "minutes_since_last_new_high": None,
        }
    price = float(bar_row["close"])
    running_max_high = path["high"].cummax()
    is_new_high = bool(path["high"].iloc[-1] >= running_max_high.iloc[-1])
    new_high_bars = path.index[path["high"] >= running_max_high]
    last_new_high_ts = new_high_bars[-1] if len(new_high_bars) else path.index[0]
    return {
        "returned_inside_opening_range": _bool_or_none(struct_ctx["or_low"] <= price <= struct_ctx["or_high"]),
        "broke_breakout_candle_low": (
            _bool_or_none(float(path["low"].min()) < struct_ctx["confirm_low"]) if struct_ctx["confirm_low"] is not None else None
        ),
        "broke_retest_low": (
            _bool_or_none(float(path["low"].min()) < struct_ctx["retest_low"]) if struct_ctx["retest_low"] is not None else None
        ),
        "new_intraday_high_since_entry": is_new_high,
        "minutes_since_last_new_high": round((bar_row.name - last_new_high_ts).total_seconds() / 60, 1),
    }


def _indicator_snapshot(frame: pd.DataFrame, entry_ts, entry_price: float, side: str, risk: float | None, bar_row) -> dict:
    """The Momentum/Trend/VWAP/EMA/Volatility/Trade-Progress fields at one
    snapshot bar - everything that does NOT depend on comparing this
    snapshot to the entry snapshot (see _derive_since_entry_fields for
    those - RSI Delta, VWAP/EMA crossing events, ADX Rising/Falling, ATR
    Expansion) or on ORB structure (see _structure_snapshot)."""
    price = float(bar_row["close"])
    sign = 1 if side == "long" else -1
    path = frame.loc[entry_ts:bar_row.name]
    high_since_entry, low_since_entry = float(path["high"].max()), float(path["low"].min())
    mfe_so_far = (high_since_entry - entry_price) if side == "long" else (entry_price - low_since_entry)
    mae_so_far = (entry_price - low_since_entry) if side == "long" else (high_since_entry - entry_price)
    total_volume_since_entry = float(path["volume"].sum())
    bullish_volume = float(path.loc[path["close"] >= path["close"].shift(1).fillna(path["close"]), "volume"].sum())
    bearish_volume = total_volume_since_entry - bullish_volume

    vwap, ema9, ema20, ema50 = bar_row.get("vwap"), bar_row.get("ema9"), bar_row.get("ema20"), bar_row.get("ema50")
    has_vwap = vwap is not None and not pd.isna(vwap)

    return {
        "bar_time": bar_row.name.isoformat(),
        "price": _round(price, 4),
        "rsi": _round(bar_row.get("rsi"), 2),
        "macd": _round(bar_row.get("macd"), 4), "macd_signal": _round(bar_row.get("macd_signal"), 4),
        "macd_hist": _round(bar_row.get("macd_hist"), 4),
        "stoch_k": _round(bar_row.get("stoch_k"), 2), "stoch_d": _round(bar_row.get("stoch_d"), 2),
        "adx": _round(bar_row.get("adx"), 2), "plus_di": _round(bar_row.get("plus_di"), 2), "minus_di": _round(bar_row.get("minus_di"), 2),
        "above_vwap": _bool_or_none(price > vwap) if has_vwap else None,
        "dist_vwap_pct": _round((price - vwap) / vwap * 100, 3) if has_vwap and vwap else None,
        "dist_vwap_r": _round((price - vwap) / risk, 3) if has_vwap and risk else None,
        "ema9": _round(ema9, 4),
        "dist_ema9_pct": _round((price - ema9) / ema9 * 100, 3) if ema9 and not pd.isna(ema9) else None,
        "above_ema9": _bool_or_none(price > ema9) if ema9 is not None and not pd.isna(ema9) else None,
        "ema20": _round(ema20, 4),
        "dist_ema20_pct": _round((price - ema20) / ema20 * 100, 3) if ema20 and not pd.isna(ema20) else None,
        "above_ema20": _bool_or_none(price > ema20) if ema20 is not None and not pd.isna(ema20) else None,
        "ema50": _round(ema50, 4),
        "dist_ema50_pct": _round((price - ema50) / ema50 * 100, 3) if ema50 and not pd.isna(ema50) else None,
        "above_ema50": _bool_or_none(price > ema50) if ema50 is not None and not pd.isna(ema50) else None,
        "obv": _round(bar_row.get("obv"), 1),
        "atr": _round(bar_row.get("atr"), 4),
        "bullish_volume": _round(bullish_volume, 1),
        "bearish_volume": _round(bearish_volume, 1),
        "total_volume_since_entry": _round(total_volume_since_entry, 1),
        "current_r": _round((price - entry_price) / risk * sign, 3) if risk else None,
        "mfe_r_so_far": _round(mfe_so_far / risk, 3) if risk else None,
        "mae_r_so_far": _round(mae_so_far / risk, 3) if risk else None,
        "current_profit_pct": _round((price - entry_price) / entry_price * 100 * sign, 3) if entry_price else None,
        "current_drawdown_pct": _round((high_since_entry - price) / high_since_entry * 100, 3) if side == "long" and high_since_entry else (
            _round((price - low_since_entry) / price * 100, 3) if side == "short" and price else None
        ),
    }


def _minutes_on_side(path: pd.DataFrame, above: bool) -> float:
    """Cumulative minutes (bar-to-bar) `path`'s own close spent above (or
    below) its own vwap column - walks consecutive bar gaps rather than
    just counting bars * 5, so this stays correct even where cached bars
    have a gap (a thin-volume symbol, a half day)."""
    mask = (path["close"] > path["vwap"]) if above else (path["close"] < path["vwap"])
    mask = mask.fillna(False)
    if len(path) < 2:
        return 0.0
    gaps = path.index.to_series().diff().dt.total_seconds().fillna(0) / 60
    return round(float(gaps[mask.values].sum()), 1)


def _derive_since_entry_fields(entry_snapshot: dict, snapshot: dict, frame: pd.DataFrame, entry_ts, bar_ts) -> dict:
    """Everything defined relative to the ENTRY snapshot's own values -
    RSI Delta, VWAP/EMA crossing events + minutes-on-each-side, ADX Rising/
    Falling, ATR Expansion. Applied to every non-entry snapshot; the entry
    snapshot itself gets a documented "baseline, not yet meaningful" all-
    None version (see build_trade_telemetry)."""
    path = frame.loc[entry_ts:bar_ts]
    entry_above_vwap, now_above_vwap = entry_snapshot.get("above_vwap"), snapshot.get("above_vwap")
    entry_above_ema9, now_above_ema9 = entry_snapshot.get("above_ema9"), snapshot.get("above_ema9")
    entry_above_ema20, now_above_ema20 = entry_snapshot.get("above_ema20"), snapshot.get("above_ema20")
    entry_ema9_gt_ema20 = (entry_snapshot["ema9"] > entry_snapshot["ema20"]) if entry_snapshot.get("ema9") is not None and entry_snapshot.get("ema20") is not None else None
    now_ema9_gt_ema20 = (snapshot["ema9"] > snapshot["ema20"]) if snapshot.get("ema9") is not None and snapshot.get("ema20") is not None else None
    entry_adx, now_adx = entry_snapshot.get("adx"), snapshot.get("adx")
    entry_atr, now_atr = entry_snapshot.get("atr"), snapshot.get("atr")

    return {
        "rsi_delta": _round(snapshot["rsi"] - entry_snapshot["rsi"], 2) if snapshot.get("rsi") is not None and entry_snapshot.get("rsi") is not None else None,
        "crossed_above_vwap": _bool_or_none(entry_above_vwap is False and now_above_vwap is True) if entry_above_vwap is not None and now_above_vwap is not None else None,
        "crossed_below_vwap": _bool_or_none(entry_above_vwap is True and now_above_vwap is False) if entry_above_vwap is not None and now_above_vwap is not None else None,
        "minutes_above_vwap": _minutes_on_side(path, above=True) if "vwap" in path.columns else None,
        "minutes_below_vwap": _minutes_on_side(path, above=False) if "vwap" in path.columns else None,
        "lost_ema9": _bool_or_none(entry_above_ema9 is True and now_above_ema9 is False) if entry_above_ema9 is not None and now_above_ema9 is not None else None,
        "lost_ema20": _bool_or_none(entry_above_ema20 is True and now_above_ema20 is False) if entry_above_ema20 is not None and now_above_ema20 is not None else None,
        "ema9_crossed_ema20": _bool_or_none(entry_ema9_gt_ema20 != now_ema9_gt_ema20) if entry_ema9_gt_ema20 is not None and now_ema9_gt_ema20 is not None else None,
        "adx_rising": _bool_or_none(now_adx > entry_adx) if now_adx is not None and entry_adx is not None else None,
        "adx_falling": _bool_or_none(now_adx < entry_adx) if now_adx is not None and entry_adx is not None else None,
        "atr_expansion": _round(now_atr / entry_atr, 3) if now_atr is not None and entry_atr not in (None, 0) else None,
    }


def build_trade_telemetry(
    frame: pd.DataFrame, today_bars: pd.DataFrame, market_frames: dict, prior_day_bars: dict,
    entry_ts, entry_price: float, side: str, risk: float | None, model: str | None,
    trail_activated_at_r: float | None = None,
) -> dict:
    """Builds the full {"entry": {...}, "1m": {...}, ...} snapshots dict
    for ONE closed trade - the core of the snapshot engine. `frame` is the
    traded symbol's own _indicator_frame (already computed once for the
    whole run, see generate_telemetry_for_backtest); `today_bars` is that
    same symbol's raw OHLCV bars for entry_ts's own calendar day (for RVOL
    and structure context); `market_frames` is {"es"|"spy"|"qqq":
    _market_context_frame|None}; `prior_day_bars` is {date: that day's raw
    bars}, the same shape backtest_engine.py already threads into orb.
    evaluate_orb_entry, reused here for orb._compute_rvol's own lookback.
    `trail_activated_at_r` is the pair's own "trail_activated_at_r" field
    (the exact MFE R-multiple at which THIS trade's trailing stop actually
    turned on, None if it never did) - used only to derive each snapshot's
    own "trail_activated_as_of" (see the loop below), not stored itself."""
    struct_ctx = _structure_context(today_bars, entry_ts, model)
    intraday_upto = frame  # RSI/EMA/etc. converge over the full multi-day series already
    snapshots = {}
    snapshot_bar_ts = {}  # label -> original tz-aware bar_row.name, reused by the derive pass below
    # (never re-parsed from the stored "bar_time" ISO string - round-
    # tripping a zoneinfo-aware Timestamp through isoformat()/pd.Timestamp()
    # yields an equal-but-not-identical fixed-offset tzinfo, which pandas'
    # own .loc slicing then rejects against frame.index's zoneinfo tzinfo).
    for label, offset_minutes in SNAPSHOT_OFFSETS:
        bar_row = _snapshot_bar(frame, entry_ts, offset_minutes)
        if bar_row is None:
            snapshots[label] = None
            continue
        bar_ts = bar_row.name
        snapshot_bar_ts[label] = bar_ts
        snap = _indicator_snapshot(frame, entry_ts, entry_price, side, risk, bar_row)
        snap["minutes_elapsed"] = round((bar_ts - entry_ts).total_seconds() / 60, 1)
        today_upto_bar = today_bars[today_bars.index <= bar_ts]
        as_of_time = bar_ts.time()
        snap["rvol"] = _round(
            orb._compute_rvol(today_upto_bar, intraday_upto, bar_ts.date(), as_of_time, RVOL_LOOKBACK_DAYS, prior_day_bars)
            if not today_upto_bar.empty else None,
            3,
        )
        # "Volume vs Average": cumulative volume since entry against this
        # symbol's own long-run average per-bar volume over the same
        # number of elapsed bars - deliberately a DIFFERENT, simpler
        # reference than RVOL just above (which is time-of-day-matched
        # against prior days), so the two fields stay genuinely distinct
        # rather than one relabeling the other.
        bars_elapsed = max(1, len(frame.loc[entry_ts:bar_ts]))
        avg_bar_volume = float(frame["volume"].tail(_WARMUP_BARS).mean())
        snap["volume_vs_average"] = (
            _round(snap["total_volume_since_entry"] / (avg_bar_volume * bars_elapsed), 3) if avg_bar_volume else None
        )
        for key, mkt_frame in market_frames.items():
            snap[key] = _market_context_snapshot(mkt_frame, entry_ts, offset_minutes)
        snap.update(_structure_snapshot(struct_ctx, frame.loc[entry_ts:bar_ts], bar_row))
        # Point-in-time fact: had trailing already activated BY this
        # snapshot - NOT the same as the trade's own final trail_activated
        # (true for the trade's WHOLE life once it ever activates, even if
        # that happened well after this snapshot's own timestamp - using
        # it here would make "not yet trailing at +10m" impossible to ever
        # tell apart from "never trails at all", silently making any rule
        # requiring both "not yet trailing" and "later became a Trailing
        # Winner" unsatisfiable by construction). MFE-so-far crossing
        # trail_activated_at_r is the EXACT condition backtest_engine.py's
        # own no_stop_delayed_trail branch uses to flip pos["trail_
        # activated"] True in the first place, so comparing this
        # snapshot's own mfe_r_so_far against the trade's real activation
        # threshold reconstructs the same fact without re-deriving
        # trailing_trigger_R from the strategy's own rules_json at all.
        # False (not None) both when it hadn't activated YET and when it
        # NEVER did - "as of this snapshot" only needs the yes/no answer,
        # not which of those two is true.
        snap["trail_activated_as_of"] = bool(
            trail_activated_at_r is not None and snap["mfe_r_so_far"] is not None
            and snap["mfe_r_so_far"] >= trail_activated_at_r
        )
        snapshots[label] = snap

    entry_snapshot = snapshots.get("entry")
    if entry_snapshot is not None:
        entry_snapshot.update({
            "rsi_delta": 0.0, "crossed_above_vwap": False, "crossed_below_vwap": False,
            "minutes_above_vwap": 0.0, "minutes_below_vwap": 0.0,
            "lost_ema9": False, "lost_ema20": False, "ema9_crossed_ema20": False,
            "adx_rising": False, "adx_falling": False, "atr_expansion": 1.0,
        })
        for label in _NON_ENTRY_LABELS:
            if snapshots.get(label) is not None:
                snapshots[label].update(_derive_since_entry_fields(entry_snapshot, snapshots[label], frame, entry_ts, snapshot_bar_ts[label]))
    return snapshots


def generate_telemetry_for_backtest(account_id: int, backtest_id: int) -> dict:
    """The "Generate Telemetry" job (see run_telemetry.py) - builds and
    stores one trade_telemetry row per closed trade across every ORB
    strategy included in this backtest's own results. Re-running this for
    the same backtest_id first clears its prior telemetry rows (db.
    delete_trade_telemetry_for_backtest), so this is safe to re-run after a
    re-analysis without ever accumulating duplicates.

    Returns a summary dict: {"trades_processed", "trades_skipped",
    "skipped_reasons": {reason: count}}."""
    backtest = db.get_backtest(backtest_id)
    if backtest is None or backtest["account_id"] != account_id:
        raise ValueError(f"Backtest {backtest_id} not found for this account")
    if backtest["status"] != "done":
        raise ValueError(f"Backtest {backtest_id} is not done (status={backtest['status']})")

    db.delete_trade_telemetry_for_backtest(backtest_id)

    symbol_cache: dict[str, dict] = {}  # symbol -> {"frame", "today_bars_by_date", "prior_day_bars"}
    market_frame_cache: dict[str, pd.DataFrame | None] = {}
    for key, mkt_symbol in MARKET_CONTEXT_SYMBOLS.items():
        mkt_bars = backtest_data.load_cached_bars(mkt_symbol, backtest_engine.BAR_SIZE)
        try:
            market_frame_cache[key] = _market_context_frame(_normalize_bars(mkt_bars), include_adx=(key == "es")) if mkt_bars is not None else None
        except Exception:  # noqa: BLE001 - a bad ES/SPY/QQQ cache must not block generation for every trade
            market_frame_cache[key] = None

    trades_processed = 0
    skipped_reasons: dict[str, int] = {}
    sample_errors: list[str] = []  # first few "reason: ExceptionType: message" strings, for diagnosing WHY without reproducing
    rows_to_insert = []

    for strategy_id, result in (backtest["results"] or {}).items():
        if not isinstance(result, dict) or "pairs" not in result:
            continue
        strategy = db.get_strategy(int(strategy_id))
        strategy_name = result.get("strategy_name") or (strategy["name"] if strategy else f"Strategy {strategy_id}")

        for pair in result["pairs"]:
            symbol = pair["symbol"]
            side = pair["side"]
            entry_iso = pair["buy_time"] if side == "long" else pair["sell_time"]
            exit_iso = pair["sell_time"] if side == "long" else pair["buy_time"]
            entry_price = pair["open_price"]
            risk = perf.initial_risk_per_share(pair)

            if symbol not in symbol_cache:
                try:
                    bars = backtest_data.load_cached_bars(symbol, backtest_engine.BAR_SIZE)
                    if bars is not None and not bars.empty:
                        bars = _normalize_bars(bars)
                    symbol_cache[symbol] = None if bars is None or bars.empty else {
                        "frame": _indicator_frame(bars),
                        "prior_day_bars": dict(tuple(bars.groupby(bars.index.date))),
                    }
                except Exception as exc:  # noqa: BLE001 - one symbol's own bad/pathological cache must not fail the whole run
                    symbol_cache[symbol] = "error"
                    if len(sample_errors) < 5:
                        sample_errors.append(f"{symbol}: {type(exc).__name__}: {exc}")
            cached = symbol_cache[symbol]
            if cached is None:
                skipped_reasons["no_cached_bars"] = skipped_reasons.get("no_cached_bars", 0) + 1
                continue
            if cached == "error":
                skipped_reasons["symbol_processing_error"] = skipped_reasons.get("symbol_processing_error", 0) + 1
                continue

            entry_ts = pd.Timestamp(entry_iso)
            if entry_ts.tzinfo is None:
                entry_ts = entry_ts.tz_localize(cached["frame"].index.tz)
            else:
                entry_ts = entry_ts.tz_convert(cached["frame"].index.tz)
            if entry_ts not in cached["frame"].index and cached["frame"][cached["frame"].index <= entry_ts].empty:
                skipped_reasons["entry_before_cached_history"] = skipped_reasons.get("entry_before_cached_history", 0) + 1
                continue

            today_bars = cached["prior_day_bars"].get(entry_ts.date())
            if today_bars is None or today_bars.empty:
                skipped_reasons["no_bars_for_entry_day"] = skipped_reasons.get("no_bars_for_entry_day", 0) + 1
                continue

            # A single trade hitting a pandas edge case (e.g. a symbol's
            # cache carrying a pathological bar shape) must not crash an
            # entire multi-hundred-trade run - same per-item isolation
            # reasoning as fetch_backtest_data.run_fetch's own per-symbol
            # try/except and run_optimization's per-chunk isolation.
            try:
                snapshots = build_trade_telemetry(
                    cached["frame"], today_bars, market_frame_cache, cached["prior_day_bars"],
                    entry_ts, entry_price, side, risk, pair.get("model"), pair.get("trail_activated_at_r"),
                )
            except Exception as exc:  # noqa: BLE001 - see comment above
                skipped_reasons["trade_processing_error"] = skipped_reasons.get("trade_processing_error", 0) + 1
                if len(sample_errors) < 5:
                    sample_errors.append(f"{symbol} @ {entry_iso}: {type(exc).__name__}: {exc}")
                continue
            rows_to_insert.append({
                "account_id": account_id, "backtest_id": backtest_id,
                "strategy_id": int(strategy_id), "strategy_name": strategy_name,
                "symbol": symbol, "side": side, "entry_time": entry_iso, "exit_time": exit_iso,
                "final_r": pair.get("final_r"), "exit_reason": pair.get("exit_reason"),
                "trail_activated": bool(pair.get("trail_activated")), "capture_pct": pair.get("capture_pct"),
                # Dollar risk (risk-per-share x size) for THIS trade - the
                # Rule Evaluation tab's own Net Benefit Engine (evaluate_
                # rule below) uses it to convert an R-multiple delta into
                # dollars. pnl_usd itself is never stored (final_r *
                # risk_dollars already reproduces it exactly - both derive
                # from the same open/close price move, see perf.r_multiple
                # vs pair_trades' own pnl_usd - so storing it separately
                # would just be a redundant, driftable copy).
                "risk_dollars": (risk * pair["size"]) if risk is not None else None,
                "snapshots": snapshots,
            })
            trades_processed += 1

    if rows_to_insert:
        db.insert_trade_telemetry_rows(rows_to_insert)
    return {
        "trades_processed": trades_processed,
        "trades_skipped": sum(skipped_reasons.values()),
        "skipped_reasons": skipped_reasons,
        "sample_errors": sample_errors,
    }


# --------------------------------------------------------------- analysis ---
# Everything below reads back ALREADY-STORED trade_telemetry rows (see db.
# list_trade_telemetry) - no bars/IBKR access, pure pandas over what
# generate_telemetry_for_backtest already computed and saved. Kept in this
# same module (not a separate one) since "analyze the telemetry this system
# itself collected" is squarely still this feature's own scope, distinct
# from backtest analysis proper.

def flatten_trades(rows: list[dict]) -> pd.DataFrame:
    """One row per trade, one column per "<snapshot>_<metric>" (e.g.
    "5m_rsi", "entry_adx", "10m_es_above_vwap") plus the trade-level
    columns in NON_FEATURE_COLUMNS - the flat feature matrix every
    comparison/ranking/heatmap/export function below works from. A
    snapshot that came back None (see _snapshot_bar) simply contributes no
    columns for that trade - NaN there, not a fabricated 0."""
    records = []
    for t in rows:
        rec = {
            "telemetry_id": t["id"], "backtest_id": t["backtest_id"], "strategy_id": t["strategy_id"],
            "strategy_name": t["strategy_name"], "symbol": t["symbol"], "side": t["side"],
            "entry_time": t["entry_time"], "exit_time": t["exit_time"],
            "final_r": t["final_r"], "exit_reason": t["exit_reason"],
            "trail_activated": bool(t["trail_activated"]), "capture_pct": t["capture_pct"],
            "risk_dollars": t.get("risk_dollars"),
        }
        for snap_label, snap in (t.get("snapshots") or {}).items():
            if not snap:
                continue
            for k, v in snap.items():
                if k == "bar_time":
                    continue
                if isinstance(v, dict):  # es/spy/qqq market-context sub-dicts
                    for kk, vv in v.items():
                        rec[f"{snap_label}_{k}_{kk}"] = int(vv) if isinstance(vv, bool) else vv
                elif isinstance(v, bool):
                    rec[f"{snap_label}_{k}"] = int(v)
                else:
                    rec[f"{snap_label}_{k}"] = v
        records.append(rec)
    return pd.DataFrame.from_records(records)


def feature_columns(df: pd.DataFrame, explicit: list[str] | None = None) -> list[str]:
    """Every column comparison_table/predictive_ranking/suggested_
    candidate_filters/the heatmap metric picker treat as a usable
    indicator feature - always excludes EXCLUDED_SNAPSHOT_PREFIXES (1m/3m,
    see its own docstring) regardless of whether `explicit` was passed, so
    no caller (present or future) has to remember to leave them out
    itself."""
    candidates = [c for c in explicit if c in df.columns] if explicit else [
        c for c in df.columns if c not in NON_FEATURE_COLUMNS and pd.api.types.is_numeric_dtype(df[c])
    ]
    return [c for c in candidates if not c.startswith(EXCLUDED_SNAPSHOT_PREFIXES)]


def apply_group(df: pd.DataFrame, group_key: str, cfg: dict | None = None) -> pd.Series:
    """Boolean mask over `df`'s rows for one of DEFAULT_GROUPS - `cfg`
    only matters for "capture_disaster" (its own configurable threshold,
    see the spec's own "Capture Disaster Group... Configurable" note)."""
    predicate = DEFAULT_GROUPS[group_key]
    cfg = cfg or {}
    return df.apply(lambda row: bool(predicate(row, cfg)), axis=1)


def metric_stats(series: pd.Series) -> dict:
    """Mean/median/std/percentiles for one metric column - Section
    "Statistical Ranking". None throughout for an empty (all-NaN) series,
    not a fabricated 0."""
    s = pd.to_numeric(series, errors="coerce").dropna()
    if s.empty:
        return {"count": 0, "mean": None, "median": None, "std": None, "p10": None, "p25": None, "p75": None, "p90": None}
    return {
        "count": int(len(s)), "mean": _round(s.mean(), 4), "median": _round(s.median(), 4),
        "std": _round(s.std(), 4) if len(s) > 1 else 0.0,
        "p10": _round(s.quantile(0.10), 4), "p25": _round(s.quantile(0.25), 4),
        "p75": _round(s.quantile(0.75), 4), "p90": _round(s.quantile(0.90), 4),
    }


def comparison_table(df: pd.DataFrame, mask_a: pd.Series, mask_b: pd.Series, columns: list[str] | None = None) -> list[dict]:
    """Section "Comparison Engine" - one row per metric, each group's own
    mean/median/count side by side."""
    cols = feature_columns(df, columns)
    rows = []
    for col in cols:
        stats_a, stats_b = metric_stats(df.loc[mask_a, col]), metric_stats(df.loc[mask_b, col])
        rows.append({
            "metric": col,
            "group_a_mean": stats_a["mean"], "group_a_median": stats_a["median"], "group_a_count": stats_a["count"],
            "group_b_mean": stats_b["mean"], "group_b_median": stats_b["median"], "group_b_count": stats_b["count"],
        })
    return rows


def predictive_ranking(df: pd.DataFrame, mask_a: pd.Series, mask_b: pd.Series, columns: list[str] | None = None) -> list[dict]:
    """Section "Predictive Power Ranking": Mutual Information + Random
    Forest feature importance for separating group A from group B (e.g.
    Hard Stop vs Trailing Winner). Requires scikit-learn (see requirements.
    txt) - returns [] rather than raising if it isn't installed, since this
    is a research/analysis feature, never load-bearing for anything else in
    the dashboard. Missing values are median-imputed per column (sklearn
    can't take NaN directly) - a column that's ALL NaN for both groups is
    dropped rather than imputed with a meaningless constant."""
    try:
        from sklearn.ensemble import RandomForestClassifier
        from sklearn.feature_selection import mutual_info_classif
    except ImportError:
        return []
    cols = feature_columns(df, columns)
    sub = df.loc[mask_a | mask_b, cols].copy()
    labels = mask_a.loc[sub.index].astype(int)
    sub = sub.dropna(axis=1, how="all")
    cols = list(sub.columns)
    if sub.empty or len(sub) < 6 or labels.nunique() < 2 or not cols:
        return []
    sub = sub.apply(lambda col: col.fillna(col.median()))
    mi = mutual_info_classif(sub.values, labels.values, random_state=0)
    rf = RandomForestClassifier(n_estimators=200, max_depth=4, random_state=0, min_samples_leaf=max(2, len(sub) // 20))
    rf.fit(sub.values, labels.values)
    ranking = [
        {"metric": col, "mutual_information": _round(m, 4), "feature_importance": _round(imp, 4)}
        for col, m, imp in zip(cols, mi, rf.feature_importances_)
    ]
    ranking.sort(key=lambda r: r["feature_importance"], reverse=True)
    return ranking


def heatmap_data(df: pd.DataFrame, metric_col: str, r_col: str = "final_r", metric_bins: int = 6) -> dict:
    """Section "Indicator Heatmaps": a metric-value-bucket x Final-R-bucket
    2D count grid, for a client-side heatmap render. Empty grid (not an
    error) if `metric_col` has too little spread/data to bucket."""
    sub = df[[metric_col, r_col]].dropna()
    if sub.empty or sub[metric_col].nunique() < 2:
        return {"metric": metric_col, "x_labels": [], "y_labels": [], "matrix": []}
    r_edges = [-float("inf"), -2, -1, 0, 1, 2, 3, float("inf")]
    r_labels = ["<=-2R", "-2R..-1R", "-1R..0R", "0R..1R", "1R..2R", "2R..3R", ">3R"]
    try:
        metric_bucket = pd.qcut(sub[metric_col], q=min(metric_bins, sub[metric_col].nunique()), duplicates="drop")
    except ValueError:
        return {"metric": metric_col, "x_labels": [], "y_labels": [], "matrix": []}
    r_bucket = pd.cut(sub[r_col], bins=r_edges, labels=r_labels)
    table = pd.crosstab(metric_bucket, r_bucket, dropna=False)
    return {
        "metric": metric_col,
        "x_labels": [str(i) for i in table.columns],
        "y_labels": [str(i) for i in table.index],
        "matrix": table.values.tolist(),
    }


def suggested_candidate_filters(df: pd.DataFrame, mask_a: pd.Series, mask_b: pd.Series, label_a: str, label_b: str, top_n: int = 5) -> list[dict]:
    """Section "Suggested Candidate Filters": for each of the top-N most
    predictive metrics (by feature importance), a simple median-threshold
    split and the % of each group on the "worse" side of it - the spec's
    own worked-example shape ("ADX < 18 after 10 minutes appears in 72% of
    Hard Stop trades and only 21% of Trailing Winners"). Purely
    descriptive/statistical - "Do NOT modify strategy" per the spec, this
    never writes back to any strategy's rules_json."""
    ranking = predictive_ranking(df, mask_a, mask_b)[:top_n]
    suggestions = []
    for r in ranking:
        col = r["metric"]
        combined = df.loc[mask_a | mask_b, col].dropna()
        if combined.empty:
            continue
        threshold = float(combined.median())
        a_vals, b_vals = df.loc[mask_a, col].dropna(), df.loc[mask_b, col].dropna()
        a_pct = _round(float((a_vals < threshold).mean() * 100), 1) if len(a_vals) else None
        b_pct = _round(float((b_vals < threshold).mean() * 100), 1) if len(b_vals) else None
        suggestions.append({
            "metric": col, "threshold": _round(threshold, 3), "direction": "below",
            f"{label_a}_pct": a_pct, f"{label_b}_pct": b_pct,
            "feature_importance": r["feature_importance"], "mutual_information": r["mutual_information"],
        })
    return suggestions


def early_failure_analysis(df: pd.DataFrame, mask_hard_stop: pd.Series, mask_trailing_winner: pd.Series) -> dict:
    """Section "Early Failure Analysis": at 5/10/15 minutes, the strongest
    indicators separating Hard Stop trades from Trailing Winners - one
    predictive_ranking call per offset, restricted to that offset's own
    columns only (so "strongest at +10m" can't be answered by a +15m
    column leaking in)."""
    result = {}
    for label in ("5m", "10m", "15m"):
        prefix = f"{label}_"
        cols = [c for c in df.columns if c.startswith(prefix) and c not in NON_FEATURE_COLUMNS]
        result[label] = predictive_ranking(df, mask_hard_stop, mask_trailing_winner, cols)[:10]
    return result


OUTCOME_CATEGORY_LABELS = {
    "hard_stop": "Hard Stops", "trailing_winner": "Trailing Winners", "trailing_loser": "Trailing Losers",
    "eod_winner": "End Of Day Winners", "eod_loser": "End Of Day Losers", "flat_trade": "Flat Trades",
    "other": "Other",
}


def _rule_outcome_category(row) -> str:
    """Mutually-exclusive, first-match-wins bucket for the Rule Outcome
    Breakdown - every possible (exit_reason, trail_activated, final_r
    sign) combination a closed ORB trade in this codebase can have lands
    in exactly one of these 7 categories:
      - hard_stop: exit_reason == "hard_stop"
      - flat_trade: closed at EXACTLY breakeven (final_r == 0) - checked
        before the winner/loser splits below so a breakeven trade never
        gets miscounted as either
      - trailing_winner / trailing_loser: trail ever activated (regardless
        of the exact exit_reason - a real trailing-stop exit or an
        eod_close after activating both count, same "trailing_winners"
        definition DEFAULT_GROUPS already uses), split by whether it
        closed positive or not
      - eod_winner / eod_loser: never trailed, held to end of day, split
        by whether it closed positive or not
      - other: the true remainder (any other exit_reason/trail_activated
        combination this codebase can produce)

    Adding trailing_loser/flat_trade here (previously both folded into
    "other") only refines the DESCRIPTIVE outcome-breakdown labels and the
    candidate-trade "classification" string - it changes nothing evaluate_
    rule's own confusion matrix (hard_stop is exit_reason-only), Net
    Benefit R/$ (Fix #1 already sums Delta R across every flagged trade
    regardless of category), or hard_stop_savings/recovery_winner_cost
    (keyed only on "hard_stop"/"trailing_winner", both unchanged here)
    depend on - every headline number for V1-V4 stays bit-identical."""
    if row.get("exit_reason") == "hard_stop":
        return "hard_stop"
    final_r = row.get("final_r")
    if final_r is not None and final_r == 0:
        return "flat_trade"
    trail_activated = bool(row.get("trail_activated"))
    if trail_activated:
        return "trailing_winner" if (final_r is not None and final_r > 0) else "trailing_loser"
    if row.get("exit_reason") == "eod_close":
        return "eod_winner" if (final_r is not None and final_r > 0) else "eod_loser"
    return "other"


_CANDIDATE_CLASSIFICATION_LABELS = {
    "hard_stop": "TP", "trailing_winner": "FP - Trailing Winner", "trailing_loser": "FP - Trailing Loser",
    "eod_winner": "FP - EOD Winner", "eod_loser": "FP - EOD Loser", "flat_trade": "FP - Flat Trade",
    "other": "FP - Other",
}


def _numeric_col(df: pd.DataFrame, col: str) -> pd.Series:
    """pd.to_numeric over df[col], or an all-NaN Series aligned to df's
    own index if `col` doesn't exist at all (e.g. every trade in this
    scope happens to be missing the offset this rule evaluates at) -
    every caller below can then treat a missing column exactly like a
    present-but-empty one, never a crash."""
    if col not in df.columns:
        return pd.Series(np.nan, index=df.index)
    return pd.to_numeric(df[col], errors="coerce")


def _group_stats(sub: pd.DataFrame) -> dict:
    final_r = pd.to_numeric(sub["final_r"], errors="coerce").dropna()
    capture = pd.to_numeric(sub["capture_pct"], errors="coerce").dropna()
    return {
        "avg_final_r": _round(final_r.mean(), 3) if len(final_r) else None,
        "median_final_r": _round(final_r.median(), 3) if len(final_r) else None,
        "avg_capture_pct": _round(capture.mean(), 1) if len(capture) else None,
    }


def _profit_factor(pnl: pd.Series):
    """Same "inf when there are no losers at all" convention perf.
    aggregate's own profit_factor already uses - None only when there's
    literally nothing to divide (no trades)."""
    pnl = pnl.dropna()
    if pnl.empty:
        return None
    wins = pnl[pnl > 0].sum()
    losses = pnl[pnl <= 0].sum()
    if losses == 0:
        return float("inf") if wins > 0 else None
    return float(wins / abs(losses))


def _max_drawdown_from_pnls(items: list[tuple]) -> float:
    """Same walk-in-exit-order peak-to-trough algorithm as perf.compute_
    max_drawdown, generalized to plain (exit_time, pnl) pairs rather than
    real trade-pair dicts - evaluate_rule's own Max Drawdown Impact needs
    to recompute the equity curve against a MODIFIED scenario (flagged
    trades swapped to their early-exit time/pnl), which perf.compute_max_
    drawdown has no way to express. A pair with a missing time or NaN pnl
    is dropped (same "can't place it on the curve" reasoning throughout
    this module) - 0.0 if nothing valid is left."""
    valid = [(t, p) for t, p in items if t and pd.notna(p)]
    if not valid:
        return 0.0
    valid.sort(key=lambda x: x[0])
    equity = peak = max_dd = 0.0
    for _, pnl in valid:
        equity += pnl
        peak = max(peak, equity)
        max_dd = max(max_dd, peak - equity)
    return max_dd


def _rule_applicable_mask(df: pd.DataFrame, offset_minutes: int) -> pd.Series:
    """Fix #2 ("Add Additional Early Failure Rule Evaluations"): a rule
    evaluated at +N minutes can only ever have been evaluated on a trade
    that was STILL OPEN at entry_time + N minutes - a trade that already
    hard-stopped (or otherwise exited) before that point never reached the
    bar the rule reads from (e.g. a 10-minute rule can't be scored against
    a trade that stopped out after 6 minutes: the +10m snapshot data
    didn't exist yet at the moment the trade needed a decision). Scoring
    it anyway would silently credit or blame the rule for information it
    could never have seen. Compares real entry/exit timestamps (not the
    snapshot's own bar-snapped time) so this is exact regardless of the
    module's own 5-minute bar resolution. `True` (applicable) whenever
    entry_time/exit_time can't even be parsed, rather than silently
    dropping a row this function can't otherwise account for - same
    "never silently drop, only visibly exclude" convention as evaluate_
    rule's own excluded_not_applicable count below."""
    entry_ts = pd.to_datetime(df["entry_time"], utc=True, errors="coerce")
    exit_ts = pd.to_datetime(df["exit_time"], utc=True, errors="coerce")
    eval_ts = entry_ts + pd.Timedelta(minutes=offset_minutes)
    not_applicable = entry_ts.notna() & exit_ts.notna() & (exit_ts < eval_ts)
    return ~not_applicable


def evaluate_rule(rows: list[dict], df: pd.DataFrame, rule_key: str) -> dict:
    """The Rule Evaluation tab's full computation - everything is purely
    RETROSPECTIVE/DESCRIPTIVE over already-closed, already-recorded trades
    (see RULES/_early_failure_candidate's own docstrings): a confusion
    matrix and classification metrics against the trade's own actual
    outcome (Hard Stop vs not), a breakdown of what flagged trades
    actually became, the Net Benefit Engine (R and $ that exiting early on
    every flagged trade would have saved/cost, plus the resulting
    Profit Factor/Max Drawdown impact over the WHOLE scoped trade set),
    and the per-trade Candidate Trades table. Nothing here writes back to
    any strategy, backtest run, or stored trade in any way.

    `rows` is the raw list[dict] (see db.list_trade_telemetry) - needed
    alongside the flattened `df` only for each trade's own snapshots[...]
    ["bar_time"] (Max Drawdown Impact's own modified exit-time lookup;
    flatten_trades doesn't carry bar_time forward, on purpose - see its
    own docstring).

    Fix #2 ("Add Additional Early Failure Rule Evaluations"): a trade that
    actually exited BEFORE this rule's own evaluation_offset (see _rule_
    applicable_mask) is excluded from EVERY part of this computation -
    confusion matrix, outcome breakdown, Net Benefit, candidate trades -
    "total" in the confusion matrix is therefore the count of EVALUABLE
    trades, not every trade in scope; `excluded_not_applicable` reports how
    many were left out and why, so that exclusion is never silent.

    A rule can also declare `missing_data_check` (see RULES - V5/V6/V7
    only, V1-V4 have none) - a second, independent exclusion for a trade
    that WAS chronologically open at the evaluation offset but whose own
    telemetry snapshot is missing a field the rule specifically needs
    (e.g. V6/V7's own mfe_r_so_far). Tracked separately from excluded_
    not_applicable (`excluded_missing_required_data`) since the reason is
    different (a real data gap, not "the trade had already closed") - the
    two are mutually exclusive and reconcile exactly against len(df):
    total (evaluated) + excluded_not_applicable + excluded_missing_
    required_data == len(df) always."""
    rule = RULES[rule_key]
    offset = rule["evaluation_offset"]
    offset_minutes = _OFFSET_MINUTES[offset]
    r_col, rsi_col = f"{offset}_current_r", f"{offset}_rsi_delta"
    or_col, ema9_col = f"{offset}_returned_inside_opening_range", f"{offset}_lost_ema9"

    empty = {
        "rule": rule_key, "label": rule["label"], "description": rule["description"],
        "evaluation_offset": offset,
        "confusion_matrix": {"tp": 0, "fp": 0, "fn": 0, "tn": 0, "total": 0},
        "metrics": {"precision": None, "recall": None, "f1_score": None, "accuracy": None, "balanced_accuracy": None},
        "outcome_breakdown": [], "net_benefit": None, "candidate_trades": [],
        "excluded_not_applicable": 0, "excluded_missing_required_data": 0,
    }
    if df.empty:
        return empty

    applicable = _rule_applicable_mask(df, offset_minutes)
    excluded_not_applicable = int((~applicable).sum())

    missing_data_check = rule.get("missing_data_check")
    if missing_data_check is not None:
        missing_data = df.apply(lambda row: bool(missing_data_check(row, {})), axis=1) & applicable
    else:
        missing_data = pd.Series(False, index=df.index)
    excluded_missing_required_data = int(missing_data.sum())
    evaluated = applicable & ~missing_data

    # `triggered` is forced False for every inapplicable/missing-data trade
    # regardless of what the predicate itself would have said (a predicate
    # reading a missing/None offset column often already returns False on
    # its own, but that's incidental to THIS column being missing, not a
    # guarantee every future rule's predicate makes - forcing it here is
    # the actual contract Fix #2 needs).
    triggered = df.apply(lambda row: bool(rule["predicate"](row, {})), axis=1) & evaluated
    actual_hard_stop = df["exit_reason"] == "hard_stop"

    tp = int((triggered & actual_hard_stop).sum())
    fp = int((triggered & ~actual_hard_stop).sum())
    fn = int((~triggered & actual_hard_stop & evaluated).sum())
    tn = int((~triggered & ~actual_hard_stop & evaluated).sum())
    total = tp + fp + fn + tn
    precision = (tp / (tp + fp)) if (tp + fp) else None
    recall = (tp / (tp + fn)) if (tp + fn) else None
    f1 = (2 * precision * recall / (precision + recall)) if (precision and recall and (precision + recall) > 0) else None
    accuracy = ((tp + tn) / total) if total else None
    specificity = (tn / (tn + fp)) if (tn + fp) else None
    balanced_accuracy = ((recall + specificity) / 2) if (recall is not None and specificity is not None) else None

    flagged = df[triggered].copy()
    flagged_count = len(flagged)
    if flagged_count == 0:
        return {
            **empty,
            "confusion_matrix": {"tp": tp, "fp": fp, "fn": fn, "tn": tn, "total": total},
            "metrics": {
                "precision": _round(precision, 3), "recall": _round(recall, 3), "f1_score": _round(f1, 3),
                "accuracy": _round(accuracy, 3), "balanced_accuracy": _round(balanced_accuracy, 3),
            },
            "excluded_not_applicable": excluded_not_applicable,
            "excluded_missing_required_data": excluded_missing_required_data,
        }
    flagged["_category"] = flagged.apply(_rule_outcome_category, axis=1)

    outcome_breakdown = []
    for cat in ("hard_stop", "trailing_winner", "trailing_loser", "eod_winner", "eod_loser", "flat_trade", "other"):
        sub = flagged[flagged["_category"] == cat]
        outcome_breakdown.append({
            "category": cat, "label": OUTCOME_CATEGORY_LABELS[cat],
            "count": len(sub), "pct": _round(len(sub) / flagged_count * 100, 1),
            **_group_stats(sub),
        })

    # --- Net Benefit Engine ---
    # Delta R = Early Exit R (this rule's own evaluation-offset current_r)
    # minus Actual Final R - positive means exiting early would have been
    # BETTER than what actually happened, negative means WORSE.
    #
    # Fix #1 ("Add Additional Early Failure Rule Evaluations"): Net Benefit
    # R/$ is the sum of Delta R/$ across EVERY flagged trade, not just the
    # ones that became a Hard Stop or a Trailing Winner - excluding End Of
    # Day Winners/Losers/Other from the total (the previous behavior)
    # silently understated the rule's real cost/benefit whenever a
    # meaningful share of flagged trades landed in one of those buckets.
    # hard_stop_savings_r/recovery_winner_cost_r/other_outcome_delta_r
    # below are still broken out by outcome category (same categories as
    # the Outcome Breakdown table) purely for transparency into WHERE
    # net_benefit_r's total came from - they are no longer what net_
    # benefit_r itself is computed from.
    flagged["_early_exit_r"] = _numeric_col(flagged, r_col)
    flagged["_delta_r"] = flagged["_early_exit_r"] - pd.to_numeric(flagged["final_r"], errors="coerce")
    hard_stop_rows = flagged[flagged["_category"] == "hard_stop"]
    trailing_winner_rows = flagged[flagged["_category"] == "trailing_winner"]
    other_outcome_rows = flagged[~flagged["_category"].isin(["hard_stop", "trailing_winner"])]

    hard_stop_savings_r = _round(hard_stop_rows["_delta_r"].sum(), 3)
    recovery_winner_cost_r = _round(-trailing_winner_rows["_delta_r"].sum(), 3)
    other_outcome_delta_r = _round(other_outcome_rows["_delta_r"].sum(), 3)
    net_benefit_r = _round(flagged["_delta_r"].sum(), 3)

    flagged["_delta_dollars"] = flagged["_delta_r"] * _numeric_col(flagged, "risk_dollars")
    hard_stop_savings_dollars = _round(flagged.loc[hard_stop_rows.index, "_delta_dollars"].sum(), 2)
    recovery_winner_cost_dollars = _round(-flagged.loc[trailing_winner_rows.index, "_delta_dollars"].sum(), 2)
    other_outcome_delta_dollars = _round(flagged.loc[other_outcome_rows.index, "_delta_dollars"].sum(), 2)
    net_benefit_dollars = _round(flagged["_delta_dollars"].sum(), 2)

    # --- Profit Factor / Max Drawdown impact, over the WHOLE scoped trade
    # set (not just flagged ones) - "if this rule had been live, what
    # would the strategy's own aggregate numbers have looked like" -
    # original_pnl is derived (final_r * risk_dollars reproduces pnl_usd
    # exactly - see generate_telemetry_for_backtest's own risk_dollars
    # comment), modified_pnl swaps a flagged trade's own pnl for its
    # early-exit equivalent (r_col * risk_dollars), everything else
    # unchanged.
    all_risk_dollars = _numeric_col(df, "risk_dollars")
    original_pnl = pd.to_numeric(df["final_r"], errors="coerce") * all_risk_dollars
    modified_pnl = original_pnl.copy()
    modified_pnl.loc[triggered] = (_numeric_col(df, r_col) * all_risk_dollars).loc[triggered]

    original_pf = _profit_factor(original_pnl)
    modified_pf = _profit_factor(modified_pnl)
    profit_factor_impact = (
        _round(modified_pf - original_pf, 3)
        if isinstance(original_pf, (int, float)) and isinstance(modified_pf, (int, float))
        else None
    )

    # `or {}` (not `.get(offset, {})`) - a trade whose own snapshots[offset]
    # key is PRESENT but None (a real, common case - see _snapshot_bar's
    # own "no bar yet" docstring, e.g. a late-session entry with no +15m
    # bar available yet) must fall back to {} too; dict.get's own default
    # only kicks in when the key is ABSENT, not when its value is None, so
    # `.get(offset, {})` alone crashes here (AttributeError: 'NoneType'
    # object has no attribute 'get') on the very case this line exists to
    # handle. Pre-existing bug (predates this rule's own V5/V6/V7 work) -
    # already affected V4 (evaluation_offset "15m") for any real scope
    # containing a trade with no +15m bar; only surfaced now via a wider
    # multi-rule test.
    bar_time_by_id = {r["id"]: ((r.get("snapshots") or {}).get(offset) or {}).get("bar_time") for r in rows}
    original_dd = _max_drawdown_from_pnls(list(zip(df["exit_time"], original_pnl)))
    modified_exit_times = df["exit_time"].copy()
    for idx in df.index[triggered]:
        bar_time = bar_time_by_id.get(df.loc[idx, "telemetry_id"])
        if bar_time:
            modified_exit_times.loc[idx] = bar_time
    modified_dd = _max_drawdown_from_pnls(list(zip(modified_exit_times, modified_pnl)))
    max_drawdown_impact = _round(modified_dd - original_dd, 2)

    # Outcome-category breakdown of the flagged population beyond the
    # hard_stop/trailing_winner subsets above - "Sacrificed" (a real cost:
    # the flagged trade would otherwise have closed positive) vs
    # "Improved" (a real benefit: the flagged trade was heading to an End
    # Of Day loss anyway, and the early exit made that loss smaller, i.e.
    # its own Delta R > 0) vs the true catch-all remainder (trailing_loser/
    # flat_trade/an EOD loser the early exit did NOT improve/literal
    # "other" - none of these get their own Matrix column, so they're
    # folded together here rather than invented a column nobody asked for).
    eod_winner_rows = flagged[flagged["_category"] == "eod_winner"]
    eod_loser_rows = flagged[flagged["_category"] == "eod_loser"]
    eod_losers_improved = int((eod_loser_rows["_delta_r"] > 0).sum())
    eod_winners_sacrificed = len(eod_winner_rows)
    other_outcomes_flagged = flagged_count - tp - len(trailing_winner_rows) - eod_winners_sacrificed - eod_losers_improved

    net_benefit = {
        "trades_flagged": flagged_count,
        "hard_stops_captured": tp, "hard_stops_missed": fn,
        "trailing_winners_sacrificed": len(trailing_winner_rows),
        "eod_winners_sacrificed": eod_winners_sacrificed, "eod_losers_improved": eod_losers_improved,
        "other_outcomes_flagged": other_outcomes_flagged,
        "hard_stop_savings_r": hard_stop_savings_r, "recovery_winner_cost_r": recovery_winner_cost_r,
        "other_outcome_count": len(other_outcome_rows), "other_outcome_delta_r": other_outcome_delta_r,
        "net_benefit_r": net_benefit_r,
        "hard_stop_savings_dollars": hard_stop_savings_dollars, "recovery_winner_cost_dollars": recovery_winner_cost_dollars,
        "other_outcome_delta_dollars": other_outcome_delta_dollars,
        "net_benefit_dollars": net_benefit_dollars,
        "profit_factor_original": _round(original_pf, 3) if isinstance(original_pf, (int, float)) else original_pf,
        "profit_factor_modified": _round(modified_pf, 3) if isinstance(modified_pf, (int, float)) else modified_pf,
        "profit_factor_impact": profit_factor_impact,
        "max_drawdown_original": _round(original_dd, 2), "max_drawdown_modified": _round(modified_dd, 2),
        "max_drawdown_impact": max_drawdown_impact,
    }

    candidate_trades = []
    for idx in flagged.index:
        row = flagged.loc[idx]
        # signal_current_r/estimated_simulated_fill_r: this module has no
        # separate order-fill/slippage model (see backtest_engine.py's own
        # commission_per_trade - a flat per-fill dollar amount tracked
        # SEPARATELY from R, never a slippage adjustment to price) - the
        # evaluation-offset bar's own Current R is both the "signal" a
        # human/rule would have read AND the only executable-price estimate
        # this module can produce, so the two are numerically identical.
        # Exposed under the honest "estimated_simulated_fill_r" name (not
        # "simulated_fill_r") so it's never mistaken for a true order-
        # simulated fill with its own slippage/commission model - see this
        # rule's own spec: "Do not present an estimate as an executed
        # backtest result."
        signal_r = _round(row.get(r_col), 3)
        candidate_trades.append({
            "telemetry_id": int(row["telemetry_id"]), "symbol": row["symbol"], "entry_time": row["entry_time"],
            "evaluation_offset": offset,
            "current_r": signal_r, "signal_current_r": signal_r, "estimated_simulated_fill_r": signal_r,
            "rsi_delta": _round(row.get(rsi_col), 2),
            "returned_inside_opening_range": bool(row.get(or_col) == 1),
            "lost_ema9": bool(row.get(ema9_col) == 1),
            "exit_reason": row["exit_reason"], "final_r": _round(row["final_r"], 3),
            "capture_pct": _round(row["capture_pct"], 1),
            "delta_r": _round(row["_delta_r"], 3),
            "outcome_category": row["_category"],
            "classification": _CANDIDATE_CLASSIFICATION_LABELS.get(row["_category"], "FP - Other"),
        })

    return {
        "rule": rule_key, "label": rule["label"], "description": rule["description"],
        "evaluation_offset": offset,
        "confusion_matrix": {"tp": tp, "fp": fp, "fn": fn, "tn": tn, "total": total},
        "metrics": {
            "precision": _round(precision, 3), "recall": _round(recall, 3), "f1_score": _round(f1, 3),
            "accuracy": _round(accuracy, 3), "balanced_accuracy": _round(balanced_accuracy, 3),
        },
        "outcome_breakdown": outcome_breakdown,
        "net_benefit": net_benefit,
        "candidate_trades": candidate_trades,
        "excluded_not_applicable": excluded_not_applicable,
        "excluded_missing_required_data": excluded_missing_required_data,
    }


def _net_benefit_label(value: float | None) -> str | None:
    """Presentational-only classification of a rule's own Net Benefit R -
    "Ranking Requirements": Positive (>0), Near Break-Even (-1R..1R),
    Negative (<-1R). The spec's own ranges overlap on (0, 1] (both
    "Positive" and "Near Break-Even" match there); Positive is checked
    first so a rule with, say, +0.5R reads as Positive rather than Near
    Break-Even - the more informative of the two. Never changes any
    number, purely a label."""
    if value is None:
        return None
    if value > 0:
        return "Positive Net Benefit"
    if value < -1:
        return "Negative Net Benefit"
    return "Near Break-Even"


# (axis key, path into a per-rule evaluate_rule() result, higher_is_better)
# - "Ranking Requirements": rank all rules by each of these, best = rank 1.
# Drawdown Reduction ranks on max_drawdown_impact ASCENDING (a more
# NEGATIVE impact means drawdown went DOWN, i.e. improved) rather than
# inventing a separate "drawdown_reduction" field nobody else reads.
_RANKING_AXES = [
    ("net_benefit_r", ("net_benefit", "net_benefit_r"), True),
    ("net_benefit_dollars", ("net_benefit", "net_benefit_dollars"), True),
    ("precision", ("metrics", "precision"), True),
    ("recall", ("metrics", "recall"), True),
    ("f1_score", ("metrics", "f1_score"), True),
    ("profit_factor_impact", ("net_benefit", "profit_factor_impact"), True),
    ("drawdown_reduction", ("net_benefit", "max_drawdown_impact"), False),
    ("trailing_winners_sacrificed", ("net_benefit", "trailing_winners_sacrificed"), False),
]


def _read_path(result: dict, path: tuple[str, ...]):
    value = result
    for key in path:
        if not isinstance(value, dict):
            return None
        value = value.get(key)
    return value


def _rank_rule_matrix(results: list[dict]) -> None:
    """Mutates each result in-place, adding "rankings" (1 = best on that
    axis, None if this rule has no value for it - e.g. profit_factor_
    impact when no trades were flagged) and "net_benefit_label". Ties get
    the same rank (standard competition ranking - 1224, not 1234), same
    convention pandas' own .rank(method="min") uses, so two rules with an
    identical Net Benefit R are never arbitrarily ordered against each
    other."""
    for axis_key, path, higher_is_better in _RANKING_AXES:
        scored = [(r, _read_path(r, path)) for r in results]
        available = sorted(
            ((r, v) for r, v in scored if v is not None),
            key=lambda rv: rv[1], reverse=higher_is_better,
        )
        rank_by_id = {}
        for i, (r, v) in enumerate(available):
            if i > 0 and available[i - 1][1] == v:
                rank_by_id[id(r)] = rank_by_id[id(available[i - 1][0])]
            else:
                rank_by_id[id(r)] = i + 1
        for r, _v in scored:
            r.setdefault("rankings", {})[axis_key] = rank_by_id.get(id(r))
    for r in results:
        r["net_benefit_label"] = _net_benefit_label(_read_path(r, ("net_benefit", "net_benefit_r")))


_VS_V3_METRICS = [
    ("net_benefit_r", ("net_benefit", "net_benefit_r")),
    ("net_benefit_dollars", ("net_benefit", "net_benefit_dollars")),
    ("precision", ("metrics", "precision")),
    ("recall", ("metrics", "recall")),
    ("trailing_winners_sacrificed", ("net_benefit", "trailing_winners_sacrificed")),
    ("max_drawdown_impact", ("net_benefit", "max_drawdown_impact")),
]


def _attach_vs_v3(results: list[dict]) -> None:
    """"Comparison Against V3 Baseline" - mutates each result in-place,
    adding "vs_v3" ({metric: this_rule_value - v3_value}). Deliberately
    reads V3's OWN values out of THIS SAME `results` list rather than any
    hardcoded number ("Do not hard-code V3 metric values. Read them from
    the current matrix evaluation results.") - if early_failure_v3 was
    not included in this particular evaluate_rule_matrix call, there is
    nothing to compare against and vs_v3 is None throughout, never a
    fabricated value."""
    v3 = next((r for r in results if r.get("rule") == "early_failure_v3"), None)
    for r in results:
        if v3 is None or r is v3:
            r["vs_v3"] = None
            continue
        comparison = {}
        for key, path in _VS_V3_METRICS:
            mine, theirs = _read_path(r, path), _read_path(v3, path)
            comparison[key] = _round(mine - theirs, 3) if mine is not None and theirs is not None else None
        r["vs_v3"] = comparison


def evaluate_rule_matrix(rows: list[dict], df: pd.DataFrame, rule_keys: list[str]) -> list[dict]:
    """Rule Evaluation Matrix: the SAME evaluate_rule computation run for
    several rules at once, so their confusion matrices / classification
    metrics / Net Benefit numbers can be lined up side by side (see the
    /telemetry page's own "Rule Evaluation Matrix" section) - no separate
    computation path, just evaluate_rule called once per key, so a rule
    can never behave differently in the matrix than it does in its own
    single-rule "Rule Evaluation" tab. Unknown keys are silently skipped
    (the web layer validates keys before calling this).

    Also attaches, on top of each evaluate_rule() result: "rankings" (this
    rule's rank on each of 8 axes among the OTHER rules in this same call
    - see _rank_rule_matrix), "net_benefit_label" (Positive/Near Break-
    Even/Negative, presentational only), and "vs_v3" (delta against
    early_failure_v3's own result in this same call, None if V3 wasn't
    requested - see _attach_vs_v3)."""
    results = [evaluate_rule(rows, df, key) for key in rule_keys if key in RULES]
    _rank_rule_matrix(results)
    _attach_vs_v3(results)
    return results


def export_dataframe(df: pd.DataFrame, fmt: str, path):
    """CSV/Excel export of the flattened trade matrix (see flatten_trades)
    - `fmt` is "csv" or "xlsx"."""
    if fmt == "csv":
        df.to_csv(path, index=False)
    elif fmt == "xlsx":
        df.to_excel(path, index=False, engine="openpyxl")
    else:
        raise ValueError(f"Unsupported export format: {fmt}")


def export_analysis_xlsx(payload: dict, scope_label: str, group_a_label: str, group_b_label: str) -> bytes:
    """The full Analysis / Comparison Engine result (see web/app.py's own
    _build_analysis_payload, which computes exactly this dict for both the
    GET /api/telemetry/analysis JSON endpoint and this export) as a
    multi-sheet .xlsx workbook - one sheet per section, real numeric cells
    (sortable/filterable in Excel), same styling convention as src.
    trades_xlsx.build_trades_xlsx. Unlike export_dataframe (the raw
    per-trade snapshot matrix), this is the COMPUTED analysis itself -
    comparison table, predictive ranking, early failure analysis, and
    suggested filters - exactly what the dashboard's own 4 analysis tabs
    show, for offline reference or sharing outside the dashboard."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color="EEF1F5", end_color="EEF1F5", fill_type="solid")
    header_font = Font(bold=True)

    def autosize(ws, ncols):
        for col in range(1, ncols + 1):
            letter = get_column_letter(col)
            width = max((len(str(c.value)) for c in ws[letter] if c.value is not None), default=8)
            ws.column_dimensions[letter].width = min(max(width + 2, 8), 40)

    def write_table(ws, rows: list[dict], columns: list[str] | None = None):
        if not rows:
            ws.append(["No data"])
            return
        cols = columns or list(rows[0].keys())
        ws.append(cols)
        for cell in ws[ws.max_row]:
            cell.font, cell.fill = header_font, header_fill
        for row in rows:
            ws.append([row.get(c) for c in cols])
        ws.freeze_panes = "A2"
        autosize(ws, len(cols))

    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_ws.append(["Trade Telemetry - Analysis / Comparison Engine"])
    summary_ws["A1"].font = Font(bold=True, size=14)
    summary_ws.append([f"Scope: {scope_label} | Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    summary_ws.append([])
    summary_ws.append([EXCLUDED_SNAPSHOT_NOTE])
    summary_ws[summary_ws.max_row][0].font = Font(italic=True)
    summary_ws[summary_ws.max_row][0].alignment = Alignment(wrap_text=True)
    summary_ws.append([])
    summary_ws.append(["Metric", "Value"])
    for cell in summary_ws[summary_ws.max_row]:
        cell.font, cell.fill = header_font, header_fill
    summary_ws.append(["Total trades in scope", payload["trade_count"]])
    summary_ws.append([f"Group A ({group_a_label})", payload["group_a_count"]])
    summary_ws.append([f"Group B ({group_b_label})", payload["group_b_count"]])
    summary_ws.append([])
    summary_ws.append(["All group counts"])
    summary_ws[summary_ws.max_row][0].font = Font(bold=True)
    for key, count in payload["group_counts"].items():
        summary_ws.append([key, count])
    autosize(summary_ws, 2)

    write_table(wb.create_sheet("Comparison"), payload["comparison"])
    write_table(wb.create_sheet("Predictive Ranking"), payload["predictive_ranking"])

    early_ws = wb.create_sheet("Early Failure Analysis")
    early_rows = [
        {"offset": offset, **row}
        for offset, rows in payload["early_failure_analysis"].items()
        for row in rows
    ]
    write_table(early_ws, early_rows, columns=["offset", "metric", "feature_importance", "mutual_information"] if early_rows else None)

    write_table(wb.create_sheet("Suggested Filters"), payload["suggested_filters"])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_rule_evaluation_xlsx(payload: dict, scope_label: str) -> bytes:
    """One Rule Evaluation result (see evaluate_rule) as a downloadable
    multi-sheet .xlsx workbook - Summary (confusion matrix, classification
    metrics, Net Benefit Engine), Outcome Breakdown, and the full Candidate
    Trades table (every flagged trade - "Export all flagged trades" per
    the feature's own spec). Same styling convention as export_analysis_
    xlsx/src.trades_xlsx.build_trades_xlsx."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color="EEF1F5", end_color="EEF1F5", fill_type="solid")
    header_font = Font(bold=True)

    def autosize(ws, ncols):
        for col in range(1, ncols + 1):
            letter = get_column_letter(col)
            width = max((len(str(c.value)) for c in ws[letter] if c.value is not None), default=8)
            ws.column_dimensions[letter].width = min(max(width + 2, 8), 40)

    def write_table(ws, rows: list[dict], columns: list[str] | None = None):
        if not rows:
            ws.append(["No data"])
            return
        cols = columns or list(rows[0].keys())
        ws.append(cols)
        for cell in ws[ws.max_row]:
            cell.font, cell.fill = header_font, header_fill
        for row in rows:
            ws.append([row.get(c) for c in cols])
        ws.freeze_panes = "A2"
        autosize(ws, len(cols))

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_ws.append([f"Rule Evaluation - {payload['label']}"])
    summary_ws["A1"].font = Font(bold=True, size=14)
    summary_ws.append([payload["description"]])
    summary_ws.append([f"Scope: {scope_label} | Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    summary_ws.append([])

    cm = payload["confusion_matrix"]
    summary_ws.append(["Confusion Matrix (vs actual Hard Stop)"])
    summary_ws[summary_ws.max_row][0].font = Font(bold=True)
    summary_ws.append(["True Positive (flagged & hard stop)", cm["tp"]])
    summary_ws.append(["False Positive (flagged & not hard stop)", cm["fp"]])
    summary_ws.append(["False Negative (not flagged & hard stop)", cm["fn"]])
    summary_ws.append(["True Negative (not flagged & not hard stop)", cm["tn"]])
    summary_ws.append(["Total trades evaluated", cm["total"]])
    summary_ws.append(["Excluded - rule not applicable (trade exited before evaluation point)", payload.get("excluded_not_applicable", 0)])
    summary_ws.append(["Excluded - missing required data", payload.get("excluded_missing_required_data", 0)])
    summary_ws.append([])

    m = payload["metrics"]
    summary_ws.append(["Metric", "Value"])
    for cell in summary_ws[summary_ws.max_row]:
        cell.font, cell.fill = header_font, header_fill
    for label, key in [("Precision", "precision"), ("Recall", "recall"), ("F1 Score", "f1_score"),
                        ("Accuracy", "accuracy"), ("Balanced Accuracy", "balanced_accuracy")]:
        summary_ws.append([label, m[key]])
    summary_ws.append([])

    nb = payload.get("net_benefit")
    if nb:
        summary_ws.append(["Net Benefit Engine"])
        summary_ws[summary_ws.max_row][0].font = Font(bold=True)
        for label, key in [
            ("Trades Flagged", "trades_flagged"), ("Hard Stops Captured", "hard_stops_captured"),
            ("Hard Stops Missed", "hard_stops_missed"), ("Trailing Winners Sacrificed", "trailing_winners_sacrificed"),
            ("End-of-Day Winners Sacrificed", "eod_winners_sacrificed"), ("End-of-Day Losers Improved", "eod_losers_improved"),
            ("Other Outcomes Flagged", "other_outcomes_flagged"),
            ("Net Benefit R (ALL flagged trades)", "net_benefit_r"), ("Net Benefit $ (ALL flagged trades)", "net_benefit_dollars"),
            ("Hard Stop Savings R", "hard_stop_savings_r"), ("Recovery Winner Cost R", "recovery_winner_cost_r"),
            ("Hard Stop Savings $", "hard_stop_savings_dollars"), ("Recovery Winner Cost $", "recovery_winner_cost_dollars"),
            ("Other Outcome trades (EOD Winner/Loser/Other)", "other_outcome_count"),
            ("Other Outcome Delta R", "other_outcome_delta_r"), ("Other Outcome Delta $", "other_outcome_delta_dollars"),
            ("Profit Factor (original)", "profit_factor_original"), ("Profit Factor (modified)", "profit_factor_modified"),
            ("Profit Factor Impact", "profit_factor_impact"),
            ("Max Drawdown $ (original)", "max_drawdown_original"), ("Max Drawdown $ (modified)", "max_drawdown_modified"),
            ("Max Drawdown Impact", "max_drawdown_impact"),
        ]:
            summary_ws.append([label, nb[key]])
    autosize(summary_ws, 2)

    write_table(wb.create_sheet("Outcome Breakdown"), payload["outcome_breakdown"])
    write_table(wb.create_sheet("Candidate Trades"), payload["candidate_trades"])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_rule_matrix_xlsx(results: list[dict], scope_label: str) -> bytes:
    """Rule Evaluation Matrix as a downloadable multi-sheet .xlsx - one
    "Matrix" summary sheet (one row per rule, the same headline numbers
    the on-screen side-by-side comparison table shows) plus one Candidate
    Trades sheet per rule (sheet name truncated to Excel's own 31-
    character limit, de-duplicated if two rule labels collide after
    truncation). Same styling convention as export_rule_evaluation_xlsx/
    export_analysis_xlsx."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color="EEF1F5", end_color="EEF1F5", fill_type="solid")
    header_font = Font(bold=True)

    def autosize(ws, ncols):
        for col in range(1, ncols + 1):
            letter = get_column_letter(col)
            width = max((len(str(c.value)) for c in ws[letter] if c.value is not None), default=8)
            ws.column_dimensions[letter].width = min(max(width + 2, 8), 40)

    def write_table(ws, rows: list[dict], columns: list[str] | None = None):
        if not rows:
            ws.append(["No data"])
            return
        cols = columns or list(rows[0].keys())
        ws.append(cols)
        for cell in ws[ws.max_row]:
            cell.font, cell.fill = header_font, header_fill
        for row in rows:
            ws.append([row.get(c) for c in cols])
        ws.freeze_panes = "A2"
        autosize(ws, len(cols))

    wb = Workbook()
    matrix_ws = wb.active
    matrix_ws.title = "Matrix"
    matrix_ws.append(["Rule Evaluation Matrix"])
    matrix_ws["A1"].font = Font(bold=True, size=14)
    matrix_ws.append([f"Scope: {scope_label} | Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    matrix_ws.append([])

    # "Rule Evaluation Matrix" - Required columns (see "Complete Early
    # Failure Rule Evaluations V5, V6 and V7" spec's own list). vs-V3
    # comparison columns (see evaluate_rule_matrix/_attach_vs_v3) are only
    # meaningful/present when early_failure_v3 is one of the rules in THIS
    # export - a rule's own "vs_v3" is None otherwise, so the columns
    # simply read blank rather than a fabricated number.
    matrix_rows = []
    for r in results:
        cm, m, nb = r["confusion_matrix"], r["metrics"], r.get("net_benefit") or {}
        vs3 = r.get("vs_v3") or {}
        matrix_rows.append({
            "rule": r["label"], "rule_description": r.get("description"), "evaluation_offset": r.get("evaluation_offset"),
            "tp": cm["tp"], "fp": cm["fp"], "fn": cm["fn"], "tn": cm["tn"],
            "total_evaluated": cm["total"], "excluded_not_applicable": r.get("excluded_not_applicable", 0),
            "excluded_missing_required_data": r.get("excluded_missing_required_data", 0),
            "precision": m["precision"], "recall": m["recall"], "f1_score": m["f1_score"],
            "accuracy": m["accuracy"], "balanced_accuracy": m["balanced_accuracy"],
            "trades_flagged": nb.get("trades_flagged"),
            "hard_stops_captured": nb.get("hard_stops_captured"),
            "trailing_winners_sacrificed": nb.get("trailing_winners_sacrificed"),
            "eod_winners_sacrificed": nb.get("eod_winners_sacrificed"),
            "eod_losers_improved": nb.get("eod_losers_improved"),
            "other_outcomes_flagged": nb.get("other_outcomes_flagged"),
            "hard_stop_savings_r": nb.get("hard_stop_savings_r"), "recovery_winner_cost_r": nb.get("recovery_winner_cost_r"),
            "other_outcome_delta_r": nb.get("other_outcome_delta_r"),
            "net_benefit_r": nb.get("net_benefit_r"), "net_benefit_dollars": nb.get("net_benefit_dollars"),
            "net_benefit_label": r.get("net_benefit_label"),
            "original_profit_factor": nb.get("profit_factor_original"), "modified_profit_factor": nb.get("profit_factor_modified"),
            "profit_factor_impact": nb.get("profit_factor_impact"),
            "original_max_drawdown": nb.get("max_drawdown_original"), "modified_max_drawdown": nb.get("max_drawdown_modified"),
            "max_drawdown_impact": nb.get("max_drawdown_impact"),
            "net_benefit_r_vs_v3": vs3.get("net_benefit_r"), "net_benefit_dollars_vs_v3": vs3.get("net_benefit_dollars"),
            "precision_vs_v3": vs3.get("precision"), "recall_vs_v3": vs3.get("recall"),
            "trailing_winners_sacrificed_vs_v3": vs3.get("trailing_winners_sacrificed"),
            "max_drawdown_impact_vs_v3": vs3.get("max_drawdown_impact"),
        })
    write_table(matrix_ws, matrix_rows)

    # "Ranking Requirements" - one row per ranking axis, rules ordered by
    # their rank on that axis (rank 1 first) - read straight off each
    # result's own "rankings" dict (see _rank_rule_matrix), never
    # recomputed here, so the sheet can never disagree with the Matrix
    # sheet's own numbers.
    rank_ws = wb.create_sheet("Rankings")
    rank_axis_labels = {
        "net_benefit_r": "Net Benefit R", "net_benefit_dollars": "Net Benefit Dollars",
        "precision": "Precision", "recall": "Recall", "f1_score": "F1",
        "profit_factor_impact": "Profit Factor Impact", "drawdown_reduction": "Drawdown Reduction",
        "trailing_winners_sacrificed": "Fewest Trailing Winners Sacrificed",
    }
    axis_paths = {axis_key: path for axis_key, path, _higher_is_better in _RANKING_AXES}
    rank_rows = []
    for axis_key, axis_label in rank_axis_labels.items():
        ranked = sorted(
            (r for r in results if (r.get("rankings") or {}).get(axis_key) is not None),
            key=lambda r: r["rankings"][axis_key],
        )
        for r in ranked:
            rank_rows.append({
                "ranking_axis": axis_label, "rank": r["rankings"][axis_key], "rule": r["label"],
                "value": _read_path(r, axis_paths[axis_key]),
            })
    write_table(rank_ws, rank_rows, columns=["ranking_axis", "rank", "rule", "value"] if rank_rows else None)

    used_names = {"Matrix", "Rankings"}
    for r in results:
        base_name = f"{r['label']} Trades"[:31]
        name, n = base_name, 2
        while name in used_names:
            suffix = f" ({n})"
            name = f"{base_name[:31 - len(suffix)]}{suffix}"
            n += 1
        used_names.add(name)
        write_table(wb.create_sheet(name), r["candidate_trades"])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
