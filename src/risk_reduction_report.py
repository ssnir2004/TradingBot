"""ORB Long V8/V9 "Dynamic Risk Reduction" comparison report - a PASSIVE,
read-only report over an ALREADY-FINISHED multi-strategy backtest's own
results (see db.get_backtest), never touching any strategy/entry/exit/
backtest logic itself. Matches trades across a baseline strategy (ORB Long
v4.2) and one or more variant strategies (V8/V9) by (symbol, entry
timestamp) - valid because every variant's own rules_json is a byte-for-
byte copy of the baseline's entry logic/position sizing (see src.db's own
EXTRA_STRATEGY_PRESETS comment for V8/V9), so the SAME (symbol, entry
timestamp) pair always identifies the SAME trade across every strategy in
one multi-strategy run - only the exit/stop management can ever differ
(see "Entry Parity Check" below, which verifies this assumption on real
data rather than just asserting it).

Every field this module reads off a variant's own pairs comes from the
single nested `pair["v6_audit"]` dict src.backtest_engine._v6_audit_record
stamps onto every V8/V9 trade (None for every other strategy) - see that
function's own docstring for the full field list. This module never reads
the OLD flat risk_event_*/hard_stop_tightened keys those functions used to
write directly onto the trade record; that scheme was replaced by the
nested v6_audit dict specifically so a strategy's audit trail can never be
silently truncated by src.perf.pair_trades' own explicit field whitelist.

Same "never silently drop, always account for every trade" discipline as
src.telemetry_engine.py throughout - a variant trade with no baseline
match (or vice versa) is reported separately (unmatched_*), never
silently excluded from a count without saying so.
"""
from __future__ import annotations

import re
from datetime import datetime

import numpy as np

from src import perf


def _round(value, digits=4):
    if value is None:
        return None
    try:
        f = float(value)
    except (TypeError, ValueError):
        return None
    if np.isnan(f) or np.isinf(f):
        return None
    return round(f, digits)


def _v6a(pair: dict) -> dict:
    """A closed pair's own V6 audit dict, or {} for a pair that never
    carried one (any non-V8/V9 strategy, or a dict missing the key
    entirely) - lets every reader below use plain .get() without a
    separate None-check at every call site."""
    return pair.get("v6_audit") or {}


def _pairs_by_key(pairs: list[dict]) -> dict[tuple, dict]:
    """(symbol, entry timestamp) -> pair - the open leg's own timestamp
    (buy_time for a long, matching this whole feature's long-only scope)
    is the SAME real bar timestamp across every strategy sharing the same
    entry logic, so this is an exact, not fuzzy, match key."""
    return {(p["symbol"], p["buy_time"]): p for p in pairs}


def _exit_time_key(pair: dict):
    # short's exit is its buy leg (SELL opens, BUY closes); long's is its
    # sell leg - same chronological-exit-order convention perf.py's own
    # compute_max_drawdown already uses.
    return pair["buy_time"] if pair.get("side") == "short" else pair.get("sell_time")


def _max_drawdown_r(pairs: list[dict]):
    """Same peak-to-trough walk as perf.compute_max_drawdown, but over the
    cumulative FINAL-R curve (not dollars) - size-independent, so it's
    comparable across strategies/position-sizing regimes the way Max
    Drawdown Dollars alone isn't. None (not 0) when no pair has a usable
    final_r at all."""
    usable = [p for p in pairs if p.get("final_r") is not None]
    if not usable:
        return None
    cum = peak = max_dd = 0.0
    for p in sorted(usable, key=_exit_time_key):
        cum += p["final_r"]
        peak = max(peak, cum)
        max_dd = max(max_dd, peak - cum)
    return round(max_dd, 3)


def _longest_losing_streak(pairs: list[dict]) -> int:
    """Longest run of consecutive final_r < 0 trades in chronological exit
    order. 0 for an empty/all-winning set."""
    usable = [p for p in pairs if p.get("final_r") is not None]
    streak = longest = 0
    for p in sorted(usable, key=_exit_time_key):
        if p["final_r"] < 0:
            streak += 1
            longest = max(longest, streak)
        else:
            streak = 0
    return longest


def _outcome_label(pair: dict) -> str:
    """Winner/Loser purely off final_r sign (None-safe) - independent of
    exit_reason (a trailing_stop exit can still be a loser if MFE never
    reached breakeven before trailing kicked in, etc.)."""
    final_r = pair.get("final_r")
    if final_r is None:
        return "unknown"
    return "winner" if final_r > 0 else "loser"


def _core_metrics(pairs: list[dict]) -> dict:
    """"Core Metrics"/"Summary" section - every field the spec's own
    Summary comparison table lists: Total Trades, Wins, Losses, Win Rate,
    Gross Profit, Gross Loss, Gross P&L, Commission, Net P&L, Profit
    Factor (all from perf.aggregate, the same numbers the /backtest
    page's own results table already shows), plus Average/Median Final R,
    Expectancy (both $ and R), Average Winner/Loser R, Max Drawdown
    (dollars AND R), and Longest Losing Streak."""
    agg = perf.aggregate(pairs)
    final_rs = [p["final_r"] for p in pairs if p.get("final_r") is not None]
    winners_r = [r for r in final_rs if r > 0]
    losers_r = [r for r in final_rs if r <= 0]
    return {
        **agg,
        "gross_profit_usd": _round(sum(p["pnl_usd"] for p in pairs if p["pnl_usd"] > 0), 2),
        "gross_loss_usd": _round(sum(p["pnl_usd"] for p in pairs if p["pnl_usd"] <= 0), 2),
        "avg_final_r": _round(sum(final_rs) / len(final_rs), 3) if final_rs else None,
        "median_final_r": _round(float(np.median(final_rs)), 3) if final_rs else None,
        # Expectancy R = mean of every trade's own final_r - mathematically
        # identical to avg_final_r (both are the same weighted-average-
        # outcome-per-trade calculation), reported as its own named field
        # since the spec's Summary table lists it separately.
        "expectancy_r": _round(sum(final_rs) / len(final_rs), 3) if final_rs else None,
        "expectancy_usd": _round(agg["net_pnl_usd"] / agg["total_trades"], 2) if agg["total_trades"] else None,
        "avg_winner_r": _round(sum(winners_r) / len(winners_r), 3) if winners_r else None,
        "avg_loser_r": _round(sum(losers_r) / len(losers_r), 3) if losers_r else None,
        "max_drawdown_usd": perf.compute_max_drawdown(pairs),
        "max_drawdown_r": _max_drawdown_r(pairs),
        "longest_losing_streak": _longest_losing_streak(pairs),
    }


def _summary_deltas(variant_core: dict, baseline_core: dict) -> dict:
    """"V8 minus V4.2" delta columns for the Summary comparison table.
    profit_factor_delta is None whenever either side is "inf"/unratable -
    an infinite-minus-finite (or infinite-minus-infinite) delta is not a
    meaningful number, so it's reported as None rather than a fabricated
    value. total_delta_r is intentionally NOT computed here (core-metric
    subtraction over two independently-filtered trade sets is not the
    same number as summing each MATCHED pair's own Delta R) - the caller
    overwrites it with _delta_analysis' own trade-matched total_delta_r,
    the only version of this number the spec's reconciliation checks can
    actually verify."""
    def _pf(m):
        pf = m.get("profit_factor")
        return pf if isinstance(pf, (int, float)) else (float("inf") if pf == "inf" else None)

    pf_v, pf_b = _pf(variant_core), _pf(baseline_core)
    pf_delta = round(pf_v - pf_b, 3) if (pf_v is not None and pf_b is not None and pf_v != float("inf") and pf_b != float("inf")) else None
    exp_v, exp_b = variant_core.get("expectancy_r"), baseline_core.get("expectancy_r")
    return {
        "net_pnl_delta_usd": _round((variant_core.get("net_pnl_usd") or 0) - (baseline_core.get("net_pnl_usd") or 0), 2),
        "win_rate_delta_pct": _round((variant_core.get("win_rate_pct") or 0) - (baseline_core.get("win_rate_pct") or 0), 2),
        "profit_factor_delta": pf_delta,
        "expectancy_delta_r": _round(exp_v - exp_b, 3) if (exp_v is not None and exp_b is not None) else None,
        "max_drawdown_delta_usd": _round((variant_core.get("max_drawdown_usd") or 0) - (baseline_core.get("max_drawdown_usd") or 0), 2),
        "total_delta_r": None,  # overwritten by the caller - see docstring above
    }


def _entry_parity_check(baseline_pairs: list[dict], variant_pairs: list[dict]) -> dict:
    """"Entry Parity Check" - verifies, on the REAL data of this run (not
    just asserted from the two strategies' identical rules_json), that
    every variant trade really does share the same entry (symbol, entry
    timestamp, entry price, position size) as its baseline counterpart.
    parity_ok is False if there is EVEN ONE unmatched trade on either
    side, or ANY entry price/position size mismatch on a matched pair -
    per the spec, a failed parity check must stop the performance
    comparison from claiming a verdict (see build_risk_reduction_report's
    own "winner" gating), not silently proceed as if nothing happened."""
    baseline_by_key = _pairs_by_key(baseline_pairs)
    variant_by_key = _pairs_by_key(variant_pairs)
    b_keys, v_keys = set(baseline_by_key), set(variant_by_key)
    matched_keys = b_keys & v_keys

    price_mismatches, size_mismatches = [], []
    for key in matched_keys:
        b, v = baseline_by_key[key], variant_by_key[key]
        if b.get("buy_price") is not None and v.get("buy_price") is not None and abs(b["buy_price"] - v["buy_price"]) > 1e-6:
            price_mismatches.append({"symbol": key[0], "entry_time": key[1], "v4_2_price": b["buy_price"], "variant_price": v["buy_price"]})
        if b.get("size") is not None and v.get("size") is not None and b["size"] != v["size"]:
            size_mismatches.append({"symbol": key[0], "entry_time": key[1], "v4_2_size": b["size"], "variant_size": v["size"]})

    baseline_only = sorted(b_keys - v_keys, key=lambda k: (k[1] or "", k[0]))
    variant_only = sorted(v_keys - b_keys, key=lambda k: (k[1] or "", k[0]))
    parity_ok = not baseline_only and not variant_only and not price_mismatches and not size_mismatches
    return {
        "v4_2_total_trades": len(baseline_pairs), "variant_total_trades": len(variant_pairs),
        "matched_trades": len(matched_keys),
        "v4_2_only_trades": len(baseline_only), "variant_only_trades": len(variant_only),
        "v4_2_only_detail": [{"symbol": k[0], "entry_time": k[1]} for k in baseline_only],
        "variant_only_detail": [{"symbol": k[0], "entry_time": k[1]} for k in variant_only],
        "entry_price_mismatches": len(price_mismatches), "entry_price_mismatch_detail": price_mismatches,
        "position_size_mismatches": len(size_mismatches), "position_size_mismatch_detail": size_mismatches,
        "parity_ok": parity_ok,
    }


def _adjusted_stop_hit_classification(variant_pair: dict | None, baseline_pair: dict | None) -> str | None:
    """The spec's own "Adjusted Stop Hit Classification" enum, computed by
    comparing this variant trade's own v6_audit state against its MATCHED
    baseline trade's own actual final outcome (never inferred from
    exit_reason alone). None (not a fabricated bucket) for a trade whose
    V6 evaluation never triggered at all - classification only applies to
    a trade V6 actually acted on."""
    if variant_pair is None:
        return None
    v6a = _v6a(variant_pair)
    if not v6a.get("v6_triggered"):
        return None
    if v6a.get("v6_same_bar_stop_ambiguity"):
        return "AMBIGUOUS_INTRABAR_ORDER"

    if v6a.get("v6_stop_changed"):
        if v6a.get("adjusted_stop_hit"):
            if baseline_pair is not None and baseline_pair.get("final_r") is not None and _outcome_label(baseline_pair) == "winner":
                return "ADJUSTED_STOP_HIT_PREVENTED_RECOVERY"
            return "ADJUSTED_STOP_HIT_SAVED_LOSS"
        if variant_pair.get("trail_activated"):
            return "ADJUSTED_STOP_NOT_HIT_LATER_TRAILING"
        return "ADJUSTED_STOP_NOT_HIT_EOD_EXIT"

    # V6_TRIGGERED_ALREADY_TIGHTER - the pre-existing stop was already
    # tighter than V6's own request, so nothing was ever changed.
    if variant_pair.get("exit_reason") == "hard_stop":
        return "ADJUSTED_STOP_NOT_HIT_ORIGINAL_HARD_STOP"
    return "NO_STOP_CHANGE_ALREADY_TIGHTER"


def _v6_event_summary_metrics(variant_pairs: list[dict]) -> dict:
    """"V6 Event Summary Metrics" - the reason-enum partition (see
    src.backtest_engine._v6_audit_record's own "v6_stop_change_reason")
    counted directly off each of the 6 permitted enum values, so the
    spec's own reconciliation identity (Evaluated + Not Applicable +
    Missing Required Data == Total Trades) holds by construction rather
    than by re-deriving it from several separately-computed booleans."""
    total = len(variant_pairs)
    reasons = [_v6a(p).get("v6_stop_change_reason") for p in variant_pairs]
    count = reasons.count

    excluded_closed = count("V6_NOT_APPLICABLE_TRADE_CLOSED")
    excluded_trailing = count("V6_NOT_APPLICABLE_TRAILING_ACTIVE")
    excluded_missing = count("V6_MISSING_REQUIRED_DATA")
    not_triggered = count("V6_NOT_TRIGGERED")
    triggered_already_tighter = count("V6_TRIGGERED_ALREADY_TIGHTER")
    triggered_tightened = count("V6_TRIGGERED_STOP_TIGHTENED")

    evaluated = not_triggered + triggered_already_tighter + triggered_tightened
    eligible = evaluated + excluded_missing
    triggered = triggered_already_tighter + triggered_tightened

    triggered_pairs = [p for p in variant_pairs if _v6a(p).get("v6_triggered")]
    adjusted_hit = sum(1 for p in variant_pairs if _v6a(p).get("adjusted_stop_hit"))
    later_trailing = sum(1 for p in triggered_pairs if p.get("trail_activated") and not _v6a(p).get("adjusted_stop_hit"))
    ended_winner = sum(1 for p in triggered_pairs if _outcome_label(p) == "winner")
    ended_loser = sum(1 for p in triggered_pairs if _outcome_label(p) == "loser")
    ambiguous = sum(1 for p in variant_pairs if _v6a(p).get("v6_same_bar_stop_ambiguity"))

    return {
        "total_trades": total,
        "trades_eligible_for_evaluation": eligible,
        "excluded_trade_closed": excluded_closed,
        "excluded_trailing_active": excluded_trailing,
        "excluded_missing_data": excluded_missing,
        "trades_evaluated": evaluated,
        "v6_risk_events_triggered": triggered,
        "pct_eligible_triggering": _round(triggered / eligible * 100, 1) if eligible else None,
        "stops_actually_tightened": triggered_tightened,
        "adjusted_stops_hit": adjusted_hit,
        "triggered_later_activated_trailing": later_trailing,
        "triggered_ended_winner": ended_winner,
        "triggered_ended_loser": ended_loser,
        "ambiguous_same_bar_intrabar": ambiguous,
    }


def _hard_stop_impact(variant_pairs: list[dict], baseline_by_key: dict[tuple, dict]) -> dict:
    """"Hard Stop Impact" - Trades Saved (the tightened stop fired AND the
    matched baseline trade's own Delta R was positive, i.e. exiting at
    the tightened level really was objectively better than what the
    baseline actually did) vs Winners Lost (the tightened stop fired but
    the MATCHED BASELINE trade went on to become a real winner - the
    tightening cut off a recovery the baseline captured). This is a
    STRICTER, magnitude-aware pair of buckets than _adjusted_stop_hit_
    classification's own SAVED_LOSS/PREVENTED_RECOVERY split (that enum
    is a binary "did the baseline end a winner or not", used by the
    Triggered Trades/Paired Trades/Audit sheets) - a trade whose baseline
    was ALSO a loser, but whose Delta R was still <= 0 (the tightened
    stop caught a WORSE outcome than the baseline's own eventual loss,
    e.g. a decline that later partially recovered before the baseline's
    own exit), deliberately falls into NEITHER bucket here rather than
    being counted as "saved" - see that trade's own Delta R in the V6
    Triggered Trades sheet for the full picture. A triggered/stop-hit
    trade with no baseline match is reported separately, never silently
    folded into either bucket."""
    saved, winners_lost, unmatched = [], [], []
    for p in variant_pairs:
        if not _v6a(p).get("adjusted_stop_hit"):
            continue
        key = (p["symbol"], p["buy_time"])
        baseline = baseline_by_key.get(key)
        if baseline is None or baseline.get("final_r") is None or p.get("final_r") is None:
            unmatched.append({"symbol": p["symbol"], "entry_time": p["buy_time"]})
            continue
        delta_r = p["final_r"] - baseline["final_r"]
        detail = {"symbol": p["symbol"], "entry_time": p["buy_time"], "delta_r": _round(delta_r, 3)}
        if _outcome_label(baseline) == "winner":
            winners_lost.append(detail)
        elif delta_r > 0:
            saved.append(detail)
    return {
        "trades_saved": len(saved), "trades_saved_detail": saved,
        "winners_lost": len(winners_lost), "winners_lost_detail": winners_lost,
        "unmatched_or_incomplete": len(unmatched), "unmatched_or_incomplete_detail": unmatched,
    }


def _delta_analysis(variant_pairs: list[dict], baseline_by_key: dict[tuple, dict]) -> dict:
    """"Delta Analysis" - Delta R = Version Result - V4.2 Result, for
    every trade this variant shares with the baseline (matched by symbol+
    entry timestamp). unmatched_trades counts variant trades with no
    baseline counterpart (should be 0 for V8/V9 vs v4.2, since they share
    identical entry logic - see Entry Parity Check for the authoritative
    check on this; a non-zero count here is the same diagnostic signal).
    total_delta_r/total_delta_pnl_usd here are the numbers every other
    section's own "Total Delta R"/"Net P&L Delta" is required to
    reconcile against (see _reconciliation_checks)."""
    deltas_r, deltas_pnl = [], []
    unmatched = 0
    for p in variant_pairs:
        if p.get("final_r") is None:
            continue
        key = (p["symbol"], p["buy_time"])
        baseline = baseline_by_key.get(key)
        if baseline is None or baseline.get("final_r") is None:
            unmatched += 1
            continue
        deltas_r.append(p["final_r"] - baseline["final_r"])
        deltas_pnl.append((p.get("pnl_usd") or 0) - (baseline.get("pnl_usd") or 0))
    if not deltas_r:
        return {
            "matched_trades": 0, "unmatched_trades": unmatched,
            "total_delta_r": None, "avg_delta_r": None, "median_delta_r": None, "total_delta_pnl_usd": None,
        }
    return {
        "matched_trades": len(deltas_r), "unmatched_trades": unmatched,
        "total_delta_r": _round(sum(deltas_r), 3),
        "avg_delta_r": _round(sum(deltas_r) / len(deltas_r), 3),
        "median_delta_r": _round(float(np.median(deltas_r)), 3),
        "total_delta_pnl_usd": _round(sum(deltas_pnl), 2),
    }


def _v6_risk_event_audit_rows(variant_pairs: list[dict]) -> list[dict]:
    """"V6 Risk Event Audit" sheet - ONE ROW PER EVERY V8/V9 trade (not
    just triggered ones), identity columns plus the complete v6_audit
    dict (every field src.backtest_engine._v6_audit_record stamps - full
    per-condition breakdown, timestamps, stop-before/after, classification
    inputs) spread flat so each field is its own Excel column."""
    rows = []
    for p in variant_pairs:
        row = {
            "symbol": p["symbol"], "entry_date": (p["buy_time"] or "")[:10],
            "entry_time": p["buy_time"], "exit_time": p.get("sell_time"),
            "exit_reason": p.get("exit_reason"), "final_r": _round(p.get("final_r"), 3),
            "size": p.get("size"),
        }
        row.update(_v6a(p))
        rows.append(row)
    rows.sort(key=lambda r: r["entry_time"] or "")
    return rows


def _v6_triggered_trades(variant_pairs: list[dict], baseline_by_key: dict[tuple, dict]) -> tuple[list[dict], dict]:
    """"V6 Triggered Trades" sheet (only trades where v6_triggered is
    True, sorted by Entry Timestamp ascending) plus its own required
    summary block - Number of triggers, Number of actual stop changes,
    Number hitting the adjusted stop, Number later activating trailing
    without hitting the adjusted stop, Number Saved Loss, Number
    Prevented Recovery, Number ambiguous, Total Delta R, Total Delta
    P&L (over whichever triggered trades have a valid baseline match -
    matched_for_delta reports exactly how many that was)."""
    triggered = [p for p in variant_pairs if _v6a(p).get("v6_triggered")]
    triggered.sort(key=lambda p: p["buy_time"] or "")

    rows = []
    delta_r_total = delta_pnl_total = 0.0
    n_delta = 0
    saved = prevented = ambiguous = later_trailing_no_hit = hit_adjusted = stop_changed_n = 0
    for p in triggered:
        v6a = _v6a(p)
        key = (p["symbol"], p["buy_time"])
        baseline = baseline_by_key.get(key)
        classification = _adjusted_stop_hit_classification(p, baseline)

        delta_r = delta_pnl = None
        if baseline is not None and baseline.get("final_r") is not None and p.get("final_r") is not None:
            delta_r = p["final_r"] - baseline["final_r"]
            delta_pnl = (p.get("pnl_usd") or 0) - (baseline.get("pnl_usd") or 0)
            delta_r_total += delta_r
            delta_pnl_total += delta_pnl
            n_delta += 1

        if v6a.get("v6_stop_changed"):
            stop_changed_n += 1
        if v6a.get("adjusted_stop_hit"):
            hit_adjusted += 1
        if classification == "ADJUSTED_STOP_NOT_HIT_LATER_TRAILING":
            later_trailing_no_hit += 1
        elif classification == "ADJUSTED_STOP_HIT_SAVED_LOSS":
            saved += 1
        elif classification == "ADJUSTED_STOP_HIT_PREVENTED_RECOVERY":
            prevented += 1
        elif classification == "AMBIGUOUS_INTRABAR_ORDER":
            ambiguous += 1

        rows.append({
            "symbol": p["symbol"], "entry_time": p["buy_time"], "exit_time": p.get("sell_time"),
            "exit_reason": p.get("exit_reason"), "v6_stop_change_reason": v6a.get("v6_stop_change_reason"),
            "v6_stop_changed": v6a.get("v6_stop_changed"), "adjusted_stop_hit": v6a.get("adjusted_stop_hit"),
            "classification": classification,
            "v4_2_final_r": _round(baseline.get("final_r"), 3) if baseline else None,
            "variant_final_r": _round(p.get("final_r"), 3),
            "delta_r": _round(delta_r, 3) if delta_r is not None else None,
            "delta_pnl_usd": _round(delta_pnl, 2) if delta_pnl is not None else None,
        })

    summary = {
        "triggers": len(triggered), "actual_stop_changes": stop_changed_n,
        "hit_adjusted_stop": hit_adjusted, "later_activated_trailing_without_hit": later_trailing_no_hit,
        "saved_loss": saved, "prevented_recovery": prevented, "ambiguous": ambiguous,
        "matched_for_delta": n_delta,
        "total_delta_r": _round(delta_r_total, 3) if n_delta else None,
        "total_delta_pnl_usd": _round(delta_pnl_total, 2) if n_delta else None,
    }
    return rows, summary


def _paired_trades_rows(baseline_pairs: list[dict], variant_pairs: list[dict]) -> list[dict]:
    """"V4.2 vs V8 Paired Trades" sheet - every (symbol, entry timestamp)
    key seen on EITHER side, one row each, sorted by entry timestamp -
    MATCHED for a real pair, V4_2_ONLY/VARIANT_ONLY for the unmatched
    remainder (the same trades Entry Parity Check already counts, shown
    here at full per-trade detail)."""
    baseline_by_key = _pairs_by_key(baseline_pairs)
    variant_by_key = _pairs_by_key(variant_pairs)
    keys = sorted(set(baseline_by_key) | set(variant_by_key), key=lambda k: (k[1] or "", k[0]))

    rows = []
    for key in keys:
        b = baseline_by_key.get(key)
        v = variant_by_key.get(key)
        match_status = "MATCHED" if (b is not None and v is not None) else ("V4_2_ONLY" if b is not None else "VARIANT_ONLY")
        delta_r = None
        if b is not None and v is not None and b.get("final_r") is not None and v.get("final_r") is not None:
            delta_r = v["final_r"] - b["final_r"]
        v6a = _v6a(v) if v is not None else {}
        rows.append({
            "symbol": key[0], "entry_time": key[1], "match_status": match_status,
            "v4_2_exit_reason": b.get("exit_reason") if b else None,
            "v4_2_final_r": _round(b.get("final_r"), 3) if b else None,
            "variant_exit_reason": v.get("exit_reason") if v else None,
            "variant_final_r": _round(v.get("final_r"), 3) if v else None,
            "delta_r": _round(delta_r, 3) if delta_r is not None else None,
            "v6_triggered": v6a.get("v6_triggered"), "v6_stop_changed": v6a.get("v6_stop_changed"),
            "adjusted_stop_hit": v6a.get("adjusted_stop_hit"),
            "adjusted_stop_hit_classification": _adjusted_stop_hit_classification(v, b) if v is not None else None,
        })
    return rows


def _reconciliation_checks(total_trades: int, v6_summary: dict, delta_analysis: dict) -> dict:
    """The spec's own 4 count-identity Reconciliation Checks (the 2
    Delta-R/Delta-P&L sum identities hold by construction here, since
    every "Total Delta R"/"Net P&L Delta" reported anywhere in this
    module is always _delta_analysis' own total_delta_r/total_delta_pnl_
    usd, never re-derived a second way - see _summary_deltas' own
    docstring)."""
    evaluated = v6_summary["trades_evaluated"]
    not_applicable = v6_summary["excluded_trade_closed"] + v6_summary["excluded_trailing_active"]
    missing = v6_summary["excluded_missing_data"]
    triggered = v6_summary["v6_risk_events_triggered"]
    stop_changed = v6_summary["stops_actually_tightened"]
    adjusted_hit = v6_summary["adjusted_stops_hit"]
    checks = {
        "partition_holds": (evaluated + not_applicable + missing) == total_trades,
        "triggered_le_evaluated": triggered <= evaluated,
        "stop_changed_le_triggered": stop_changed <= triggered,
        "adjusted_hit_le_stop_changed": adjusted_hit <= stop_changed,
    }
    checks["all_checks_passed"] = all(checks.values())
    return checks


def build_risk_reduction_report(
    results_by_strategy: dict, strategy_labels: dict, baseline_id: str, variant_ids: list[str],
) -> dict:
    """The full comparison report. `results_by_strategy` is {strategy_id_
    str: {"pairs": [...], ...}} - a subset of one already-finished multi-
    strategy backtest's own results_json (see db.get_backtest). `baseline_
    id` should be the ORB Long v4.2 strategy_id (str), `variant_ids` the
    V8/V9 strategy_id(s) (str) - all as they appear as keys of `results_
    by_strategy`. `strategy_labels` maps every id (baseline + variants) to
    its display name, for the report's own headers/exports.

    Returns:
      {"baseline": {"strategy_id", "label", "core_metrics", "trade_count"},
       "variants": [{"strategy_id", "label", "core_metrics", "trade_count",
                     "entry_parity", "v6_event_summary_metrics",
                     "hard_stop_impact", "delta_analysis", "summary_deltas",
                     "reconciliation_checks", "v6_risk_event_audit_rows",
                     "v6_triggered_trades_rows", "v6_triggered_trades_summary",
                     "paired_trades_rows"}, ...],
       "winner": {...} | None  (only when exactly 2 variants are compared;
                 "verdict_blocked": True instead of a verdict when either
                 variant's own entry_parity check failed - see below)}

    Every section handles a missing/empty strategy result gracefully
    (empty pairs list) rather than raising - a variant strategy that
    genuinely produced zero trades in this scope is a valid, reportable
    outcome, not an error."""
    baseline_result = results_by_strategy.get(baseline_id) or {}
    baseline_pairs = baseline_result.get("pairs") or []
    baseline_by_key = _pairs_by_key(baseline_pairs)
    baseline_core = _core_metrics(baseline_pairs)

    baseline_section = {
        "strategy_id": baseline_id, "label": strategy_labels.get(baseline_id, f"Strategy #{baseline_id}"),
        "core_metrics": baseline_core, "trade_count": len(baseline_pairs),
    }

    variants_section = []
    for vid in variant_ids:
        variant_result = results_by_strategy.get(vid) or {}
        variant_pairs = variant_result.get("pairs") or []
        variant_core = _core_metrics(variant_pairs)

        entry_parity = _entry_parity_check(baseline_pairs, variant_pairs)
        v6_summary = _v6_event_summary_metrics(variant_pairs)
        triggered_rows, triggered_summary = _v6_triggered_trades(variant_pairs, baseline_by_key)
        v6_summary = {
            **v6_summary,
            "saved_loss_trades": triggered_summary["saved_loss"],
            "prevented_recovery_trades": triggered_summary["prevented_recovery"],
        }
        delta_analysis = _delta_analysis(variant_pairs, baseline_by_key)
        hard_stop_impact = _hard_stop_impact(variant_pairs, baseline_by_key)
        reconciliation = _reconciliation_checks(len(variant_pairs), v6_summary, delta_analysis)
        summary_deltas = _summary_deltas(variant_core, baseline_core)
        summary_deltas["total_delta_r"] = delta_analysis["total_delta_r"]

        variants_section.append({
            "strategy_id": vid, "label": strategy_labels.get(vid, f"Strategy #{vid}"),
            "core_metrics": variant_core, "trade_count": len(variant_pairs),
            "entry_parity": entry_parity,
            "v6_event_summary_metrics": v6_summary,
            "hard_stop_impact": hard_stop_impact,
            "delta_analysis": delta_analysis,
            "summary_deltas": summary_deltas,
            "reconciliation_checks": reconciliation,
            "v6_risk_event_audit_rows": _v6_risk_event_audit_rows(variant_pairs),
            "v6_triggered_trades_rows": triggered_rows,
            "v6_triggered_trades_summary": triggered_summary,
            "paired_trades_rows": _paired_trades_rows(baseline_pairs, variant_pairs),
        })

    winner = None
    if len(variants_section) == 2:
        a, b = variants_section
        if not (a["entry_parity"]["parity_ok"] and b["entry_parity"]["parity_ok"]):
            winner = {
                "verdict_blocked": True,
                "reason": (
                    "Entry parity check failed for at least one variant (see its own entry_parity section) - "
                    "refusing to claim either variant outperformed or underperformed until the mismatch is resolved."
                ),
            }
        else:
            am, bm = a["core_metrics"], b["core_metrics"]

            def _pf(m):
                pf = m.get("profit_factor")
                return pf if isinstance(pf, (int, float)) else (float("inf") if pf == "inf" else None)

            winner = {
                "verdict_blocked": False,
                "better_net_pnl": a["label"] if (am["net_pnl_usd"] or 0) > (bm["net_pnl_usd"] or 0) else b["label"],
                "better_profit_factor": a["label"] if (_pf(am) or -1) > (_pf(bm) or -1) else b["label"],
                "better_drawdown": a["label"] if (am["max_drawdown_usd"] or 0) < (bm["max_drawdown_usd"] or 0) else b["label"],
                "better_expectancy": a["label"] if (am.get("expectancy_r") or -1e18) > (bm.get("expectancy_r") or -1e18) else b["label"],
            }

    return {"baseline": baseline_section, "variants": variants_section, "winner": winner}


def _short_tag(label: str) -> str:
    """A short, Excel-sheet-name-safe tag for one strategy's own label
    ("ORB Long V8 (Dynamic Risk Reduction 2.0R)" -> "V8") - falls back to
    the first 6 characters for a label that doesn't carry a "V<n>" tag."""
    m = re.search(r"\bV(\d+)\b", label)
    return f"V{m.group(1)}" if m else label[:6]


def export_risk_reduction_report_xlsx(report: dict, scope_label: str) -> bytes:
    """Downloadable multi-sheet .xlsx:
      - Summary: core metrics for baseline + every variant side by side,
        Delta columns (variant minus baseline), reconciliation-check
        status, and the V8-vs-V9 Winner verdict when present.
      - Per variant: "{tag} V6 Audit" (one row per EVERY trade),
        "{tag} Triggered" (summary block + triggered-only rows),
        "{tag} Parity" (Entry Parity Check detail), "{tag} Paired"
        (V4.2-vs-variant paired trades).
    Same styling convention as src.telemetry_engine's own export_rule_
    matrix_xlsx/export_rule_evaluation_xlsx."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color="EEF1F5", end_color="EEF1F5", fill_type="solid")
    header_font = Font(bold=True)

    def autosize(ws, ncols):
        for col in range(1, ncols + 1):
            letter = get_column_letter(col)
            width = max((len(str(c.value)) for c in ws[letter] if c.value is not None), default=8)
            ws.column_dimensions[letter].width = min(max(width + 2, 8), 40)

    def write_table(ws, rows: list[dict], columns: list[str] | None = None):
        if not rows:
            ws.append(["No data"])
            return
        cols = columns or list(rows[0].keys())
        ws.append(cols)
        for cell in ws[ws.max_row]:
            cell.font, cell.fill = header_font, header_fill
        for row in rows:
            ws.append([row.get(c) for c in cols])
        ws.freeze_panes = "A2"
        autosize(ws, len(cols))

    def write_kv(ws, rows: list[tuple]):
        for label, value in rows:
            ws.append([label, value])
        autosize(ws, 2)

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_ws.append(["ORB Long V8/V9 - Dynamic Risk Reduction Report"])
    summary_ws["A1"].font = Font(bold=True, size=14)
    summary_ws.append([f"Scope: {scope_label} | Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    summary_ws.append([])

    core_cols = [
        "total_trades", "wins", "losses", "win_rate_pct",
        "gross_profit_usd", "gross_loss_usd", "gross_pnl_usd", "total_commission_usd", "net_pnl_usd",
        "profit_factor", "avg_final_r", "median_final_r", "expectancy_r", "avg_winner_r", "avg_loser_r",
        "max_drawdown_usd", "max_drawdown_r", "longest_losing_streak",
    ]
    core_rows = [{"strategy": report["baseline"]["label"], **{c: report["baseline"]["core_metrics"].get(c) for c in core_cols}}]
    for v in report["variants"]:
        core_rows.append({"strategy": v["label"], **{c: v["core_metrics"].get(c) for c in core_cols}})
    write_table(summary_ws, core_rows, columns=["strategy"] + core_cols)
    summary_ws.append([])

    delta_cols = ["net_pnl_delta_usd", "total_delta_r", "profit_factor_delta", "expectancy_delta_r", "max_drawdown_delta_usd", "win_rate_delta_pct"]
    summary_ws.append(["Variant vs V4.2 - Delta Columns"])
    summary_ws[summary_ws.max_row][0].font = Font(bold=True)
    delta_rows = [{"strategy": v["label"], **{c: v["summary_deltas"].get(c) for c in delta_cols}} for v in report["variants"]]
    write_table(summary_ws, delta_rows, columns=["strategy"] + delta_cols)
    summary_ws.append([])

    summary_ws.append(["Reconciliation Checks"])
    summary_ws[summary_ws.max_row][0].font = Font(bold=True)
    recon_cols = ["partition_holds", "triggered_le_evaluated", "stop_changed_le_triggered", "adjusted_hit_le_stop_changed", "all_checks_passed"]
    recon_rows = [{"strategy": v["label"], **{c: v["reconciliation_checks"].get(c) for c in recon_cols}} for v in report["variants"]]
    write_table(summary_ws, recon_rows, columns=["strategy"] + recon_cols)
    summary_ws.append([])

    if report.get("winner"):
        summary_ws.append(["V8 vs V9 Winner"])
        summary_ws[summary_ws.max_row][0].font = Font(bold=True)
        w = report["winner"]
        if w.get("verdict_blocked"):
            summary_ws.append(["VERDICT BLOCKED", w["reason"]])
        else:
            for label, key in [
                ("Better Net P&L", "better_net_pnl"), ("Better Profit Factor", "better_profit_factor"),
                ("Better Max Drawdown", "better_drawdown"), ("Better Expectancy", "better_expectancy"),
            ]:
                summary_ws.append([label, w[key]])
    autosize(summary_ws, len(core_cols) + 1)

    for v in report["variants"]:
        tag = _short_tag(v["label"])

        audit_ws = wb.create_sheet(f"{tag} V6 Audit"[:31])
        write_table(audit_ws, v["v6_risk_event_audit_rows"])

        trig_ws = wb.create_sheet(f"{tag} Triggered"[:31])
        s = v["v6_triggered_trades_summary"]
        write_kv(trig_ws, [
            ("Number of Triggers", s["triggers"]),
            ("Number of Actual Stop Changes", s["actual_stop_changes"]),
            ("Number Hitting Adjusted Stop", s["hit_adjusted_stop"]),
            ("Number Later Activating Trailing Without Hitting Adjusted Stop", s["later_activated_trailing_without_hit"]),
            ("Number Saved Loss", s["saved_loss"]),
            ("Number Prevented Recovery", s["prevented_recovery"]),
            ("Number Ambiguous (Same-Bar)", s["ambiguous"]),
            ("Matched For Delta", s["matched_for_delta"]),
            ("Total Delta R", s["total_delta_r"]),
            ("Total Delta P&L ($)", s["total_delta_pnl_usd"]),
        ])
        trig_ws.append([])
        write_table(trig_ws, v["v6_triggered_trades_rows"])

        parity_ws = wb.create_sheet(f"{tag} Parity"[:31])
        ep = v["entry_parity"]
        write_kv(parity_ws, [
            ("V4.2 Total Trades", ep["v4_2_total_trades"]), ("Variant Total Trades", ep["variant_total_trades"]),
            ("Matched Trades", ep["matched_trades"]),
            ("V4.2-Only Trades", ep["v4_2_only_trades"]), ("Variant-Only Trades", ep["variant_only_trades"]),
            ("Entry Price Mismatches", ep["entry_price_mismatches"]), ("Position Size Mismatches", ep["position_size_mismatches"]),
            ("Parity OK", ep["parity_ok"]),
        ])
        parity_ws.append([])
        if ep["v4_2_only_detail"]:
            parity_ws.append(["V4.2-Only Trades"])
            write_table(parity_ws, ep["v4_2_only_detail"])
            parity_ws.append([])
        if ep["variant_only_detail"]:
            parity_ws.append(["Variant-Only Trades"])
            write_table(parity_ws, ep["variant_only_detail"])
            parity_ws.append([])
        if ep["entry_price_mismatch_detail"]:
            parity_ws.append(["Entry Price Mismatches"])
            write_table(parity_ws, ep["entry_price_mismatch_detail"])
            parity_ws.append([])
        if ep["position_size_mismatch_detail"]:
            parity_ws.append(["Position Size Mismatches"])
            write_table(parity_ws, ep["position_size_mismatch_detail"])

        paired_ws = wb.create_sheet(f"{tag} Paired"[:31])
        write_table(paired_ws, v["paired_trades_rows"])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
