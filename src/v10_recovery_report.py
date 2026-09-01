"""ORB Long V10 "Dynamic Recovery" audit/comparison report - a PASSIVE,
read-only report over an ALREADY-FINISHED multi-strategy backtest's own
results (see db.get_backtest), never touching any strategy/entry/exit/
backtest logic itself - same discipline as src.risk_reduction_report.py,
whose fully-generic helpers (core metrics, entry parity, max drawdown,
delta analysis, summary deltas) this module reuses directly rather than
re-implementing, since none of them are actually V6/V8-specific.

Every field this module reads off a V10 trade comes from the single
nested pair["v10_audit"] dict src.backtest_engine._v10_audit_record
stamps onto every V10 trade (None for every other strategy) - see that
function's own docstring for the full field list, and _evaluate_v10_
recovery's own docstring for the "checkpoint_log" list each closed V10
trade's own audit carries (one entry per checkpoint actually reached).
"""
from __future__ import annotations

import re
from datetime import datetime

from src.risk_reduction_report import (
    _core_metrics, _delta_analysis, _entry_parity_check, _exit_time_key,
    _outcome_label, _pairs_by_key, _round, _summary_deltas,
)

# The spec's own required, exhaustive classification enum - every V6-
# triggered V10 trade gets EXACTLY one of these (see classify_v10_outcome).
V10_OUTCOME_CLASSIFICATIONS = (
    "RECOVERED_AND_LATER_WINNER", "RECOVERED_AND_LATER_LOSER",
    "PERSISTENT_FAILURE_STOP_SAVED_LOSS", "PERSISTENT_FAILURE_STOP_PREVENTED_RECOVERY",
    "PERSISTENT_FAILURE_STOP_NO_MATERIAL_CHANGE", "TRAILING_ACTIVATED_DURING_MONITORING",
    "CLOSED_BEFORE_NEXT_CHECKPOINT", "NO_PERSISTENT_FAILURE_ACTION",
    "MISSING_REQUIRED_DATA", "AMBIGUOUS_INTRABAR_ORDER",
)


def _v10a(pair: dict) -> dict:
    return pair.get("v10_audit") or {}


def classify_v10_outcome(variant_pair: dict, baseline_pair: dict | None) -> str | None:
    """The spec's own "Outcome Classifications" - one of the 10 values in
    V10_OUTCOME_CLASSIFICATIONS for every V6-triggered trade, None for a
    trade the warning never fired on at all (not a V10-relevant trade in
    the first place). recovery_confirmed_by_score (see backtest_engine's
    own _evaluate_v10_recovery) is what separates a REAL R1-R8-confirmed
    recovery from the 45m "neither confirmed" fallthrough, which reuses
    the same runtime RECOVERED state operationally but is semantically
    NO_PERSISTENT_FAILURE_ACTION here."""
    v10a = _v10a(variant_pair)
    if not v10a.get("v10_warning_triggered"):
        return None
    if v10a.get("v10_same_bar_ambiguity"):
        return "AMBIGUOUS_INTRABAR_ORDER"

    state = v10a.get("v10_state")
    if state == "MISSING_REQUIRED_DATA":
        return "MISSING_REQUIRED_DATA"
    if state == "TRAILING_ACTIVE":
        return "TRAILING_ACTIVATED_DURING_MONITORING"
    if state == "RECOVERED":
        if not v10a.get("v10_recovery_confirmed_by_score"):
            return "NO_PERSISTENT_FAILURE_ACTION"
        return "RECOVERED_AND_LATER_WINNER" if _outcome_label(variant_pair) == "winner" else "RECOVERED_AND_LATER_LOSER"
    if state == "PERSISTENT_FAILURE":
        if not v10a.get("adjusted_stop_hit"):
            return "PERSISTENT_FAILURE_STOP_NO_MATERIAL_CHANGE"
        if baseline_pair is not None and baseline_pair.get("final_r") is not None and _outcome_label(baseline_pair) == "winner":
            return "PERSISTENT_FAILURE_STOP_PREVENTED_RECOVERY"
        if baseline_pair is not None and baseline_pair.get("final_r") is not None and variant_pair.get("final_r") is not None:
            delta_r = variant_pair["final_r"] - baseline_pair["final_r"]
            return "PERSISTENT_FAILURE_STOP_SAVED_LOSS" if delta_r > 0 else "PERSISTENT_FAILURE_STOP_NO_MATERIAL_CHANGE"
        return "PERSISTENT_FAILURE_STOP_NO_MATERIAL_CHANGE"
    # MONITORING_RECOVERY at close time (e.g. closed via the ORIGINAL
    # untouched hard stop, or eod_close, before ever reaching a terminal
    # checkpoint decision) or CLOSED_BEFORE_EVALUATION reached here only
    # if warning_triggered was somehow True without a terminal state -
    # both map to the same "never reached a real decision" bucket.
    return "CLOSED_BEFORE_NEXT_CHECKPOINT"


def _checkpoint_confirmed_at(v10a: dict, action: str) -> int | None:
    """The checkpoint (minutes) whose own checkpoint_log row recorded
    `action` (e.g. "RECOVERED" or "PERSISTENT_FAILURE_STOP_TIGHTENED") -
    None if that action never appears in this trade's own log."""
    for row in v10a.get("checkpoint_log") or []:
        if row.get("action_taken") == action:
            return row.get("checkpoint_minutes")
    return None


def _v10_state_summary(variant_pairs: list[dict], baseline_by_key: dict[tuple, dict]) -> dict:
    total = len(variant_pairs)
    evaluated = sum(1 for p in variant_pairs if _v10a(p).get("v10_warning_evaluated"))
    warned = sum(1 for p in variant_pairs if _v10a(p).get("v10_warning_triggered"))
    recovered_at = {20: 0, 30: 0, 45: 0}
    failure_confirmed_at = {30: 0, 45: 0}
    for p in variant_pairs:
        v10a = _v10a(p)
        cp = _checkpoint_confirmed_at(v10a, "RECOVERED")
        if cp in recovered_at:
            recovered_at[cp] += 1
        cp = _checkpoint_confirmed_at(v10a, "PERSISTENT_FAILURE_STOP_TIGHTENED")
        if cp in failure_confirmed_at:
            failure_confirmed_at[cp] += 1

    trailing_during_monitoring = sum(1 for p in variant_pairs if _v10a(p).get("v10_state") == "TRAILING_ACTIVE")
    candidates_20m = sum(1 for p in variant_pairs if _v10a(p).get("v10_persistent_failure_candidate_at_20m"))
    stops_tightened = sum(1 for p in variant_pairs if _v10a(p).get("v10_stop_changed"))
    stops_hit = sum(1 for p in variant_pairs if _v10a(p).get("adjusted_stop_hit"))

    saved_loss = prevented_recovery = 0
    deltas_r, deltas_pnl = [], []
    for p in variant_pairs:
        key = (p["symbol"], p["buy_time"])
        baseline = baseline_by_key.get(key)
        classification = classify_v10_outcome(p, baseline)
        if classification == "PERSISTENT_FAILURE_STOP_SAVED_LOSS":
            saved_loss += 1
        elif classification == "PERSISTENT_FAILURE_STOP_PREVENTED_RECOVERY":
            prevented_recovery += 1
        if baseline is not None and baseline.get("final_r") is not None and p.get("final_r") is not None:
            deltas_r.append(p["final_r"] - baseline["final_r"])
            deltas_pnl.append((p.get("pnl_usd") or 0) - (baseline.get("pnl_usd") or 0))

    return {
        "total_trades": total,
        "v6_evaluations": evaluated,
        "v6_warnings": warned,
        "pct_receiving_warning": _round(warned / total * 100, 1) if total else None,
        "recovered_at_20m": recovered_at[20], "recovered_at_30m": recovered_at[30], "recovered_at_45m": recovered_at[45],
        "trailing_activated_during_monitoring": trailing_during_monitoring,
        "persistent_failure_candidates_at_20m": candidates_20m,
        "confirmed_persistent_failures_at_30m": failure_confirmed_at[30],
        "confirmed_persistent_failures_at_45m": failure_confirmed_at[45],
        "stops_tightened": stops_tightened,
        "adjusted_stops_hit": stops_hit,
        "saved_loss_trades": saved_loss,
        "prevented_recovery_trades": prevented_recovery,
        "total_delta_r": _round(sum(deltas_r), 3) if deltas_r else None,
        "total_delta_net_pnl_usd": _round(sum(deltas_pnl), 2) if deltas_pnl else None,
    }


def _reconciliation_checks(variant_pairs: list[dict], state_summary: dict) -> dict:
    """The spec's own reconciliation identities, verified directly
    against the counted data (never assumed)."""
    total = state_summary["total_trades"]
    closed_before_eval = sum(1 for p in variant_pairs if _v10a(p).get("v10_state") == "CLOSED_BEFORE_EVALUATION")
    not_triggered = sum(1 for p in variant_pairs if _v10a(p).get("v10_warning_evaluated") and not _v10a(p).get("v10_warning_triggered"))
    missing_data = sum(1 for p in variant_pairs if _v10a(p).get("v10_state") == "MISSING_REQUIRED_DATA")
    warned = state_summary["v6_warnings"]
    partition_holds = (not_triggered + warned + closed_before_eval + missing_data) == total

    recovered = sum(1 for p in variant_pairs if _v10a(p).get("v10_state") == "RECOVERED")
    persistent_failure = sum(1 for p in variant_pairs if _v10a(p).get("v10_state") == "PERSISTENT_FAILURE")
    trailing_during = state_summary["trailing_activated_during_monitoring"]
    closed_during_monitoring = sum(
        1 for p in variant_pairs
        if _v10a(p).get("v10_warning_triggered") and _v10a(p).get("v10_state") == "MONITORING_RECOVERY"
    )
    warning_partition_holds = (recovered + persistent_failure + trailing_during + closed_during_monitoring) == warned

    return {
        "no_v6_warning": not_triggered, "v6_warning_trades": warned,
        "closed_before_evaluation": closed_before_eval, "missing_or_error": missing_data,
        "total_trades": total, "partition_holds": partition_holds,
        "warning_recovered": recovered, "warning_persistent_failure": persistent_failure,
        "warning_trailing_during_monitoring": trailing_during, "warning_closed_during_monitoring": closed_during_monitoring,
        "warning_partition_holds": warning_partition_holds,
        "stop_changed_le_persistent_failure": state_summary["stops_tightened"] <= persistent_failure,
        "adjusted_hit_le_stop_changed": state_summary["adjusted_stops_hit"] <= state_summary["stops_tightened"],
        "all_checks_passed": partition_holds and warning_partition_holds
            and state_summary["stops_tightened"] <= persistent_failure
            and state_summary["adjusted_stops_hit"] <= state_summary["stops_tightened"],
    }


def _base_identity_row(p: dict, i: int) -> dict:
    size = p.get("size")
    open_price = p.get("open_price")
    initial_stop = p.get("initial_stop")
    risk_width = abs(open_price - initial_stop) if open_price is not None and initial_stop is not None else None
    return {
        "Trade ID": i, "Symbol": p.get("symbol"),
        "Entry Date": (p.get("buy_time") or "")[:10], "Entry Time": p.get("buy_time"),
        "Exit Date": (p.get("sell_time") or "")[:10], "Exit Time": p.get("sell_time"),
        "Entry Price": open_price, "Position Size": size,
        "Initial Risk Dollars": _round(risk_width * size, 2) if risk_width is not None and size else None,
        "Initial Risk Width": _round(risk_width, 4),
    }


def _warning_block(v10a: dict) -> dict:
    cp10 = (v10a.get("checkpoint_log") or [{}])[0] if v10a.get("checkpoint_log") else {}
    conditions10 = cp10.get("conditions") or {}
    return {
        "V6 Evaluated?": v10a.get("v10_warning_evaluated"), "V6 Triggered?": v10a.get("v10_warning_triggered"),
        "V6 Evaluation Timestamp": cp10.get("actual_ts").isoformat() if hasattr(cp10.get("actual_ts"), "isoformat") else cp10.get("actual_ts"),
        "V10 Warning State Activated": v10a.get("v10_warning_triggered"),
        "10m Current R": v10a.get("10m_current_r"), "10m RSI": v10a.get("10m_rsi"), "10m RSI Delta": v10a.get("10m_rsi_delta"),
        "10m MFE R So Far": v10a.get("10m_mfe_r"), "10m MAE R So Far": v10a.get("10m_mae_r"),
        "10m Price Above EMA9": v10a.get("10m_above_ema9"), "10m Price Above EMA20": v10a.get("10m_above_ema20"),
        "10m Price Above VWAP": v10a.get("10m_above_vwap"),
        "10m Returned Inside Opening Range": conditions10.get("returned_inside_or"),
    }


def _checkpoint_compact_columns(v10a: dict) -> dict:
    by_cp = {row["checkpoint_minutes"]: row for row in (v10a.get("checkpoint_log") or [])}
    out = {}
    for cp in (15, 20, 30, 45):
        row = by_cp.get(cp) or {}
        out[f"{cp}m Current R"] = row.get("current_r")
        out[f"{cp}m Recovery Score"] = row.get("recovery_score")
        out[f"{cp}m Deterioration Score"] = row.get("deterioration_score")
        out[f"{cp}m Action Taken"] = row.get("action_taken")
    return out


def _stop_audit_block(v10a: dict) -> dict:
    return {
        "Stop Before V10 R": v10a.get("stop_before_v10_r"), "Stop Before V10 Price": v10a.get("stop_before_v10_price"),
        "Requested V10 Stop R": v10a.get("requested_v10_stop_r"), "Requested V10 Stop Price": v10a.get("requested_v10_stop_price"),
        "Stop After V10 R": v10a.get("stop_after_v10_r"), "Stop After V10 Price": v10a.get("stop_after_v10_price"),
        "Stop Changed?": v10a.get("v10_stop_changed"), "Stop Change Timestamp": v10a.get("v10_stop_change_timestamp"),
        "Adjusted Stop Hit?": v10a.get("adjusted_stop_hit"), "Adjusted Stop Hit Timestamp": v10a.get("adjusted_stop_hit_timestamp"),
        "Adjusted Stop Fill Price": v10a.get("adjusted_stop_fill_price"),
    }


def _outcome_block(p: dict, baseline_by_key: dict[tuple, dict], baseline_id_by_key: dict[tuple, int]) -> dict:
    key = (p["symbol"], p["buy_time"])
    baseline = baseline_by_key.get(key)
    delta_r = delta_pnl = None
    if baseline is not None and baseline.get("final_r") is not None and p.get("final_r") is not None:
        delta_r = p["final_r"] - baseline["final_r"]
        delta_pnl = (p.get("pnl_usd") or 0) - (baseline.get("pnl_usd") or 0)
    net_pnl = None
    if p.get("pnl_usd") is not None:
        net_pnl = _round(float(p["pnl_usd"]) - float(p.get("commission_usd") or 0), 2)
    return {
        "Actual Exit Timestamp": p.get("sell_time"), "Actual Exit Reason": p.get("exit_reason"),
        "Actual Final R": _round(p.get("final_r"), 3), "Actual Gross P&L": _round(p.get("pnl_usd"), 2),
        "Commission": _round(p.get("commission_usd"), 2), "Actual Net P&L": net_pnl,
        # Pairs carry no native per-trade id (perf.pair_trades never
        # assigns one) - this is the SAME sequential (by entry time) id
        # the "V4.2 vs V10" baseline sheet would use if it listed its own
        # trades by number, computed once in build_v10_recovery_report so
        # it's consistent across every row referencing a baseline trade.
        "Matched V4.2 Trade ID": baseline_id_by_key.get(key) if baseline else None,
        "V4.2 Final R": _round(baseline.get("final_r"), 3) if baseline else None,
        "V4.2 Net P&L": _round((baseline.get("pnl_usd") or 0) - (baseline.get("commission_usd") or 0), 2) if baseline else None,
        "Delta R Versus V4.2": _round(delta_r, 3) if delta_r is not None else None,
        "Delta Net P&L Versus V4.2": _round(delta_pnl, 2) if delta_pnl is not None else None,
        "Outcome Classification": classify_v10_outcome(p, baseline),
    }


def _all_trades_audit_rows(variant_pairs: list[dict], baseline_by_key: dict[tuple, dict], baseline_id_by_key: dict[tuple, int]) -> list[dict]:
    ordered = sorted(variant_pairs, key=lambda p: p.get("buy_time") or "")
    rows = []
    for i, p in enumerate(ordered, start=1):
        v10a = _v10a(p)
        row = {**_base_identity_row(p, i), **_warning_block(v10a), **_checkpoint_compact_columns(v10a),
               **_stop_audit_block(v10a), **_outcome_block(p, baseline_by_key, baseline_id_by_key)}
        rows.append(row)
    return rows


def _checkpoint_log_rows(variant_pairs: list[dict]) -> list[dict]:
    """Bonus sheet (not one of the spec's own 11 named ones, but a direct,
    necessary elaboration of "Mandatory Runtime Audit -> Checkpoint
    Results") - one row per (trade, checkpoint actually evaluated), every
    field _evaluate_v10_recovery's own checkpoint_log already carries."""
    ordered = sorted(variant_pairs, key=lambda p: p.get("buy_time") or "")
    rows = []
    for i, p in enumerate(ordered, start=1):
        for cp_row in (_v10a(p).get("checkpoint_log") or []):
            row = {"Trade ID": i, "Symbol": p.get("symbol"), "Entry Time": p.get("buy_time")}
            for k, v in cp_row.items():
                if k == "conditions":
                    continue
                row[k] = v.isoformat() if hasattr(v, "isoformat") else v
            rows.append(row)
    return rows


def _filtered_rows(all_rows: list[dict], variant_pairs: list[dict], predicate) -> list[dict]:
    ordered = sorted(variant_pairs, key=lambda p: p.get("buy_time") or "")
    return [row for row, p in zip(all_rows, ordered) if predicate(p)]


def _errors_exclusions_rows(all_rows: list[dict], variant_pairs: list[dict]) -> list[dict]:
    def is_exclusion(p):
        state = _v10a(p).get("v10_state")
        return state in ("CLOSED_BEFORE_EVALUATION", "MISSING_REQUIRED_DATA") or _v10a(p).get("v10_same_bar_ambiguity")

    ordered = sorted(variant_pairs, key=lambda p: p.get("buy_time") or "")
    rows = []
    for row, p in zip(all_rows, ordered):
        if not is_exclusion(p):
            continue
        v10a = _v10a(p)
        reason = "AMBIGUOUS_INTRABAR_ORDER" if v10a.get("v10_same_bar_ambiguity") else v10a.get("v10_state")
        rows.append({
            "Trade ID": row["Trade ID"], "Symbol": row["Symbol"], "Entry Time": row["Entry Time"],
            "Reason": reason,
            "V6 Evaluation Timestamp": row.get("V6 Evaluation Timestamp"),
        })
    return rows


_SCORE_BUCKETS = [("0-2", 0, 2), ("3-4", 3, 4), ("5-6", 5, 6), ("7-8", 7, 8)]


def _checkpoint_analysis(variant_pairs: list[dict]) -> list[dict]:
    """"Checkpoint Analysis" - descriptive-only aggregates per 15/20/30/45m
    checkpoint (never used to retune V10's own thresholds - see the
    spec's own "Anti-Overfitting Rules")."""
    import statistics as stats

    out = []
    for cp in (15, 20, 30, 45):
        cp_rows = []
        for p in variant_pairs:
            for row in (_v10a(p).get("checkpoint_log") or []):
                if row.get("checkpoint_minutes") == cp and "recovery_score" in row:
                    cp_rows.append({**row, "_final_r": p.get("final_r"), "_exit_reason": p.get("exit_reason")})
        n = len(cp_rows)
        current_rs = [r["current_r"] for r in cp_rows if r.get("current_r") is not None]
        rec_scores = [r["recovery_score"] for r in cp_rows]
        det_scores = [r["deterioration_score"] for r in cp_rows]
        out.append({
            "Checkpoint (minutes)": cp, "Number Evaluated": n,
            "Mean Current R": _round(stats.mean(current_rs), 3) if current_rs else None,
            "Median Current R": _round(stats.median(current_rs), 3) if current_rs else None,
            "Mean Recovery Score": _round(stats.mean(rec_scores), 2) if rec_scores else None,
            "Median Recovery Score": _round(stats.median(rec_scores), 2) if rec_scores else None,
            "Mean Deterioration Score": _round(stats.mean(det_scores), 2) if det_scores else None,
            "Median Deterioration Score": _round(stats.median(det_scores), 2) if det_scores else None,
        })
        for label, lo, hi in _SCORE_BUCKETS:
            bucket_rows = [r for r in cp_rows if lo <= r["recovery_score"] <= hi]
            winners = sum(1 for r in bucket_rows if r["_final_r"] is not None and r["_final_r"] > 0)
            hard_stops = sum(1 for r in bucket_rows if r["_exit_reason"] in ("hard_stop", "v10_persistent_failure_stop"))
            final_rs = [r["_final_r"] for r in bucket_rows if r["_final_r"] is not None]
            out.append({
                "Checkpoint (minutes)": cp, "Recovery Score Bucket": label,
                "Trades In Bucket": len(bucket_rows),
                "Later-Winner Rate %": _round(winners / len(bucket_rows) * 100, 1) if bucket_rows else None,
                "Later-Hard-Stop Rate %": _round(hard_stops / len(bucket_rows) * 100, 1) if bucket_rows else None,
                "Final Average R": _round(sum(final_rs) / len(final_rs), 3) if final_rs else None,
            })
    return out


def build_v10_recovery_report_from_pairs(
    baseline_pairs: list[dict], variant_pairs: list[dict], baseline_label: str, variant_label: str,
    baseline_id: str | None = None, v10_id: str | None = None,
) -> dict:
    """The full V10 Phase-2 report, built directly off two already-
    resolved pairs lists - the shared core behind build_v10_recovery_
    report below (one backtest's own two strategies) AND
    audit_v10_recovery.py's own --list/pooling path (v4.2 and V10 pairs
    POOLED across many separate weekly-chunked backtests, matching
    audit_v6_risk_event.py's own --backtest-id pooling for V8/V9 - real
    deployments produce one backtest row per week, so a full multi-month
    audit needs many ids pooled together, not just one pair of strategy
    ids from a single backtest). An empty pairs list is a valid,
    reportable outcome, never an error."""
    baseline_by_key = _pairs_by_key(baseline_pairs)
    baseline_ordered = sorted(baseline_pairs, key=lambda p: p.get("buy_time") or "")
    baseline_id_by_key = {(p["symbol"], p["buy_time"]): i for i, p in enumerate(baseline_ordered, start=1)}

    baseline_core = _core_metrics(baseline_pairs)
    variant_core = _core_metrics(variant_pairs)
    entry_parity = _entry_parity_check(baseline_pairs, variant_pairs)
    delta_analysis = _delta_analysis(variant_pairs, baseline_by_key)
    summary_deltas = _summary_deltas(variant_core, baseline_core)
    summary_deltas["total_delta_r"] = delta_analysis["total_delta_r"]

    state_summary = _v10_state_summary(variant_pairs, baseline_by_key)
    reconciliation = _reconciliation_checks(variant_pairs, state_summary)

    all_rows = _all_trades_audit_rows(variant_pairs, baseline_by_key, baseline_id_by_key)
    warning_rows = _filtered_rows(all_rows, variant_pairs, lambda p: _v10a(p).get("v10_warning_triggered"))
    recovered_rows = _filtered_rows(all_rows, variant_pairs, lambda p: _v10a(p).get("v10_state") == "RECOVERED")
    persistent_failure_rows = _filtered_rows(all_rows, variant_pairs, lambda p: _v10a(p).get("v10_state") == "PERSISTENT_FAILURE")
    changed_outcome_rows = _filtered_rows(all_rows, variant_pairs, lambda p: _v10a(p).get("adjusted_stop_hit"))
    errors_rows = _errors_exclusions_rows(all_rows, variant_pairs)
    checkpoint_log_rows = _checkpoint_log_rows(variant_pairs)
    checkpoint_analysis_rows = _checkpoint_analysis(variant_pairs)

    return {
        "baseline": {"strategy_id": baseline_id, "label": baseline_label,
                     "core_metrics": baseline_core, "trade_count": len(baseline_pairs)},
        "variant": {"strategy_id": v10_id, "label": variant_label,
                    "core_metrics": variant_core, "trade_count": len(variant_pairs)},
        "entry_parity": entry_parity, "delta_analysis": delta_analysis, "summary_deltas": summary_deltas,
        "state_summary": state_summary, "reconciliation": reconciliation,
        "all_trades_audit_rows": all_rows, "warning_trades_rows": warning_rows,
        "recovered_trades_rows": recovered_rows, "persistent_failure_rows": persistent_failure_rows,
        "changed_outcomes_rows": changed_outcome_rows, "errors_exclusions_rows": errors_rows,
        "checkpoint_log_rows": checkpoint_log_rows, "checkpoint_analysis_rows": checkpoint_analysis_rows,
    }


def build_v10_recovery_report(results_by_strategy: dict, strategy_labels: dict, baseline_id: str, v10_id: str) -> dict:
    """Single-backtest entry point (used by web/app.py's own routes) -
    both strategies' pairs come from the SAME already-finished multi-
    strategy backtest's own results dict. See build_v10_recovery_report_
    from_pairs for the shared implementation."""
    baseline_pairs = (results_by_strategy.get(baseline_id) or {}).get("pairs") or []
    variant_pairs = (results_by_strategy.get(v10_id) or {}).get("pairs") or []
    baseline_label = strategy_labels.get(baseline_id, f"Strategy #{baseline_id}")
    variant_label = strategy_labels.get(v10_id, f"Strategy #{v10_id}")
    return build_v10_recovery_report_from_pairs(baseline_pairs, variant_pairs, baseline_label, variant_label, baseline_id, v10_id)


def export_v10_recovery_report_xlsx(report: dict, scope_label: str) -> bytes:
    """The spec's own required 11 sheets, plus one bonus 12th ("V10
    Checkpoint Log", the full raw per-checkpoint data "V10 All Trades
    Audit" only carries a compact summary of)."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill(start_color="EEF1F5", end_color="EEF1F5", fill_type="solid")
    header_font = Font(bold=True)

    def autosize(ws, ncols):
        for col in range(1, ncols + 1):
            letter = get_column_letter(col)
            width = max((len(str(c.value)) for c in ws[letter] if c.value is not None), default=8)
            ws.column_dimensions[letter].width = min(max(width + 2, 8), 40)

    def write_table(ws, rows: list[dict], columns: list[str] | None = None):
        if not rows:
            ws.append(["No data"])
            return
        cols = columns or list(rows[0].keys())
        ws.append(cols)
        for cell in ws[ws.max_row]:
            cell.font, cell.fill = header_font, header_fill
        for row in rows:
            ws.append([row.get(c) for c in cols])
        ws.freeze_panes = "A2"
        autosize(ws, len(cols))

    def write_kv(ws, rows: list[tuple]):
        for label, value in rows:
            ws.append([label, value])
        autosize(ws, 2)

    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Executive Summary"
    summary_ws.append(["ORB Long V10 - Dynamic Recovery Report"])
    summary_ws["A1"].font = Font(bold=True, size=14)
    summary_ws.append([f"Scope: {scope_label} | Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    summary_ws.append([])
    core_cols = [
        "total_trades", "wins", "losses", "win_rate_pct", "gross_pnl_usd", "total_commission_usd", "net_pnl_usd",
        "profit_factor", "avg_final_r", "median_final_r", "expectancy_r", "avg_winner_r", "avg_loser_r",
        "max_drawdown_usd", "max_drawdown_r", "longest_losing_streak",
    ]
    core_rows = [
        {"strategy": report["baseline"]["label"], **{c: report["baseline"]["core_metrics"].get(c) for c in core_cols}},
        {"strategy": report["variant"]["label"], **{c: report["variant"]["core_metrics"].get(c) for c in core_cols}},
    ]
    write_table(summary_ws, core_rows, columns=["strategy"] + core_cols)
    autosize(summary_ws, len(core_cols) + 1)

    v42_v10_ws = wb.create_sheet("V4.2 vs V10")
    v42_v10_ws.append(["V10 minus V4.2 - Delta Columns"])
    v42_v10_ws[1][0].font = Font(bold=True)
    delta_cols = ["net_pnl_delta_usd", "total_delta_r", "profit_factor_delta", "expectancy_delta_r", "max_drawdown_delta_usd", "win_rate_delta_pct"]
    write_table(v42_v10_ws, [{"strategy": report["variant"]["label"], **{c: report["summary_deltas"].get(c) for c in delta_cols}}],
                columns=["strategy"] + delta_cols)
    v42_v10_ws.append([])
    v42_v10_ws.append(["Delta Analysis (trade-matched)"])
    v42_v10_ws[v42_v10_ws.max_row][0].font = Font(bold=True)
    write_kv(v42_v10_ws, list(report["delta_analysis"].items()))

    parity_ws = wb.create_sheet("Entry Parity Check")
    ep = report["entry_parity"]
    write_kv(parity_ws, [
        ("V4.2 Total Trades", ep["v4_2_total_trades"]), ("V10 Total Trades", ep["variant_total_trades"]),
        ("Matched Trades", ep["matched_trades"]), ("V4.2-Only Trades", ep["v4_2_only_trades"]),
        ("V10-Only Trades", ep["variant_only_trades"]), ("Entry Price Mismatches", ep["entry_price_mismatches"]),
        ("Position Size Mismatches", ep["position_size_mismatches"]), ("Parity OK", ep["parity_ok"]),
    ])
    if not ep["parity_ok"]:
        parity_ws.append([])
        parity_ws.append(["PARITY FAILED - do not claim V10 outperformed/underperformed V4.2 until resolved"])
        parity_ws[parity_ws.max_row][0].font = Font(bold=True, color="CC0000")

    state_ws = wb.create_sheet("V10 State Summary")
    write_kv(state_ws, list(report["state_summary"].items()))
    state_ws.append([])
    state_ws.append(["Reconciliation Checks"])
    state_ws[state_ws.max_row][0].font = Font(bold=True)
    write_kv(state_ws, list(report["reconciliation"].items()))

    write_table(wb.create_sheet("V10 All Trades Audit"), report["all_trades_audit_rows"])
    write_table(wb.create_sheet("V10 Warning Trades"), report["warning_trades_rows"])
    write_table(wb.create_sheet("V10 Recovered Trades"), report["recovered_trades_rows"])
    write_table(wb.create_sheet("V10 Persistent Failures"), report["persistent_failure_rows"])
    write_table(wb.create_sheet("V10 Changed Outcomes"), report["changed_outcomes_rows"])
    write_table(wb.create_sheet("V10 Errors and Exclusions"), report["errors_exclusions_rows"])
    write_table(wb.create_sheet("Checkpoint Analysis"), report["checkpoint_analysis_rows"])
    write_table(wb.create_sheet("V10 Checkpoint Log"), report["checkpoint_log_rows"])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
