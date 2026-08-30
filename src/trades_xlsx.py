"""Renders a strategy's pooled backtest trade log (see
perf.pooled_trades_for_strategy) as a downloadable .xlsx workbook - unlike
trades_pdf.py (a print/read-only blotter with everything pre-formatted to
strings), this keeps every numeric column a REAL number so the trader can
sort/filter/pivot/chart it directly in Excel instead of retyping the data.

Timestamps in a pair (buy_time/sell_time) are already tz-aware ISO strings
in America/New_York (see fetch_backtest_data.py's tz_convert(ET) on every
bar it fetches) - written here as naive datetimes (tzinfo stripped, same
wall-clock convention trades_pdf.py's _fmt_dt already uses), since Excel's
datetime cells don't carry timezone info at all.
"""
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src import entry_metrics, perf
from src.trades_pdf import EXIT_REASON_LABELS

_HEADER_FILL = PatternFill(start_color="EEF1F5", end_color="EEF1F5", fill_type="solid")
_HEADER_FONT = Font(bold=True)
_MONEY_FMT = "$#,##0.00"
_R_FMT = "0.00"
_PCT_FMT = "0.0%"


def _naive_et(iso: str | None) -> datetime | None:
    if not iso:
        return None
    try:
        return datetime.fromisoformat(iso).replace(tzinfo=None)
    except ValueError:
        return None


def _autosize(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        letter = get_column_letter(col)
        width = max((len(str(c.value)) for c in ws[letter] if c.value is not None), default=8)
        ws.column_dimensions[letter].width = min(max(width + 2, 8), 40)


def _write_header(ws, header: list[str]) -> None:
    ws.append(header)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def build_trades_xlsx(strategy_name: str, direction: str, backtests_included: int,
                       aggregate_stats: dict, pairs: list[dict], diagnostics: dict | None = None,
                       description: str | None = None) -> bytes:
    """`pairs` must already be enriched (see trade_diagnostics.enrich_all) -
    same contract as trades_pdf.build_trades_pdf, which this mirrors in
    scope (same source data, same "Exit Reason Breakdown"/summary numbers)
    but not in intent - one sheet per trade-by-trade log entry with REAL
    numbers, plus a Summary sheet, instead of a single print layout."""
    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_ws.append([f"{strategy_name} — Backtest Trade Log"])
    summary_ws["A1"].font = Font(bold=True, size=14)
    summary_ws.append([f"Direction: {direction or '-'} | {backtests_included} backtest(s) pooled | "
                        f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    if description:
        summary_ws.append([])
        summary_ws.append([description])
        summary_ws[summary_ws.max_row][0].alignment = Alignment(wrap_text=True, vertical="top")
        summary_ws.row_dimensions[summary_ws.max_row].height = 60
    summary_ws.append([])

    a = aggregate_stats
    summary_rows = [
        ("Total trades", a["total_trades"]), ("Wins", a["wins"]), ("Losses", a["losses"]),
        ("Win rate %", a["win_rate_pct"]), ("Profit factor", a["profit_factor"]),
        ("Gross P&L $", a["gross_pnl_usd"]), ("Commission $", a["total_commission_usd"]),
        ("Net P&L $", a["net_pnl_usd"]), ("Avg winner $", a["avg_winner"]), ("Avg loser $", a["avg_loser"]),
    ]
    if diagnostics and diagnostics.get("summary", {}).get("diagnosable_trades"):
        s = diagnostics["summary"]
        summary_rows += [
            ("Avg MFE $", s["avg_mfe_usd"]), ("Avg MFE R", s["avg_mfe_r"]),
            ("Avg MAE $", s["avg_mae_usd"]), ("Avg MAE R", s["avg_mae_r"]),
            ("Avg Final R", s["avg_final_r"]), ("Median Final R", s["median_final_r"]),
            ("Avg Capture %", s["avg_capture_pct"]), ("Median Capture %", s["median_capture_pct"]),
        ]
    summary_ws.append(["Metric", "Value"])
    for cell in summary_ws[summary_ws.max_row]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    for label, value in summary_rows:
        summary_ws.append([label, value])

    breakdown = diagnostics.get("exit_reason_breakdown") if diagnostics else None
    if breakdown:
        summary_ws.append([])
        summary_ws.append(["Exit Reason Breakdown"])
        summary_ws[summary_ws.max_row][0].font = Font(bold=True, size=12)
        rb_header = ["Category", "Trades", "% of Total", "Win Rate %", "Net P&L $",
                     "Avg Final R", "Median Final R", "Avg MFE R", "Avg Capture %"]
        summary_ws.append(rb_header)
        for cell in summary_ws[summary_ws.max_row]:
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
        for row in breakdown:
            summary_ws.append([
                EXIT_REASON_LABELS.get(row["category"], row["category"]),
                row["trade_count"], row["pct_of_total"], row["win_rate_pct"], row["net_pnl_usd"],
                row["avg_final_r"], row["median_final_r"], row["avg_mfe_r"], row["avg_capture_pct"],
            ])
    _autosize(summary_ws, 9)

    trades_ws = wb.create_sheet("Trades")
    base_header = [
        "#", "Symbol", "Side", "Model", "Entry $", "Stop", "Risk $", "Exit $", "Size",
        "MFE $", "MFE R", "MAE $", "MAE R", "P&L $", "Final R", "Capture %", "Commission $",
        "Exit Reason", "Profit Lock Activated", "Profit Lock Triggered At (R)",
        "Trail Activated", "Trail Triggered At (R)", "Entry Time (ET)", "Exit Time (ET)",
    ]
    # Point-in-time entry metrics (see src/entry_metrics.py) appended after
    # the existing columns - one column per ENTRY_METRICS_KEYS entry, raw
    # values (None for any pre-feature/non-ORB pair), no per-cell number
    # formatting given how many there are - still fully usable as numbers/
    # booleans/strings in Excel as-is.
    metrics_header = [k.replace("_", " ").title() for k in entry_metrics.ENTRY_METRICS_KEYS]
    header = base_header + metrics_header
    _write_header(trades_ws, header)
    metrics_start_col = len(base_header) + 1

    money_cols = {5, 6, 7, 8, 10, 12, 14, 17}  # Entry $/Stop/Risk $/Exit $/MFE $/MAE $/P&L $/Commission $
    r_cols = {11, 13, 15, 20, 22}
    pct_cols = {16}
    dt_cols = {23, 24}

    for i, p in enumerate(pairs, start=1):
        entry_time = p["sell_time"] if p["side"] == "short" else p["buy_time"]
        exit_time = p["buy_time"] if p["side"] == "short" else p["sell_time"]
        risk = perf.initial_risk_per_share(p)
        risk_usd = risk * p["size"] if risk is not None else None
        row = [
            i, p["symbol"], p["side"], p.get("model"),
            p["open_price"], p.get("initial_stop"), risk_usd, p["close_price"], p["size"],
            p.get("mfe_usd"), p.get("mfe_r"), p.get("mae_usd"), p.get("mae_r"), p["pnl_usd"],
            p.get("final_r"), (p.get("capture_pct") / 100 if p.get("capture_pct") is not None else None),
            p.get("commission_usd"),
            EXIT_REASON_LABELS.get(p.get("exit_reason"), p.get("exit_reason") or "-"),
            "Yes" if p.get("profit_lock_activated") else ("No" if p.get("profit_lock_activated") is not None else "-"),
            p.get("profit_lock_activated_at_r"),
            "Yes" if p.get("trail_activated") else ("No" if p.get("trail_activated") is not None else "-"),
            p.get("trail_activated_at_r"),
            _naive_et(entry_time), _naive_et(exit_time),
        ]
        row += [p.get(k) for k in entry_metrics.ENTRY_METRICS_KEYS]
        trades_ws.append(row)
        r = trades_ws.max_row
        for col in money_cols:
            trades_ws.cell(r, col).number_format = _MONEY_FMT
        for col in r_cols:
            trades_ws.cell(r, col).number_format = _R_FMT
        for col in pct_cols:
            trades_ws.cell(r, col).number_format = _PCT_FMT
        for col in dt_cols:
            trades_ws.cell(r, col).number_format = "yyyy-mm-dd hh:mm"
    _autosize(trades_ws, len(header))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
